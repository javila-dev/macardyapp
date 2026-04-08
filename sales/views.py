import datetime
import locale
import numpy_financial

from django.conf import settings
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse, JsonResponse
from django.http.response import HttpResponseRedirect
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.urls.conf import path
from django.db.models import Sum, Count, Exists, OuterRef, Max, Value, DecimalField as ModelDecimalField
from django.db.models.query import F
from django.db.models.query_utils import Q
from django.contrib.auth.models import User
from finance.models import Collection_feed, Collector_per_sale, Credit_info, Incomes, Incomes_detail, Sales_extra_info, Incomes_return
from finance.views import apply_income
from mcd_site.models import Counters, Parameters, Projects, Timeline
from mcd_site.utils import JsonRender, link_callback, parse_semantic_date, pdf_gen, project_permission, user_check_perms, user_permission
from terceros.models import Clients, Sellers
from dateutil.relativedelta import relativedelta
from decimal import Decimal, InvalidOperation
from django.db.models.functions import Coalesce

from sales.forms import adjudicate_saleForm, change_plan_Form, change_property_Form, collectionfeed_Form, newsaleForm, SalesPlanForm
from sales.models import (Assigned_comission, Comission_position, Payment_plans, Paid_comissions, Properties, Sales,
                          Sales_history, Sales_plans, backup_payment_plans, IncomeDetailsBackup)
from terceros.models import Clients
from sales.utils import backup_plan_pagos, recalcular_plan_pagos
# Create your views here.

locale.setlocale(locale.LC_ALL,'es_CO.UTF-8')
#messages.success(request,'<div class="header">¡Lo hicimos!</div>Aprobaste el contrato '+sale)


def get_positions_queryset(project_obj, group, include_default=False):
    filters = {'group': group, 'is_active': True}
    if include_default:
        filters['include_default'] = True
    qs = Comission_position.objects.filter(project=project_obj, **filters).order_by('name')
    if not qs.exists():
        qs = Comission_position.objects.filter(project__isnull=True, **filters).order_by('name')
    return qs


def get_position_for_project(project_obj, name):
    try:
        return Comission_position.objects.get(project=project_obj, name=name)
    except Comission_position.DoesNotExist:
        return Comission_position.objects.get(project__isnull=True, name=name)


def project_selection(request):

    next_url = request.GET.get('next')
    menu = request.GET.get('menu')
    if next_url is None:
        messages.error(
            request, '<div class="header">Encontramos un error</div>No existe una url asociada a esta seleccion')
        HttpResponseRedirect('/landing')

    context = {
        'projects': Projects.objects.all(),
        'next': next_url,
        'menu':menu,
    }

    return render(request, 'project_selection.html', context)

@login_required
@project_permission
@user_permission('crear ventas')
def new_sale(request, project):
    obj_project = Projects.objects.get(name=project)
    if request.method == 'POST':
        rq = request.POST
        first_owner = rq.get('id_first_owner')
        second_owner = rq.get('id_second_owner')
        third_owner = rq.get('id_third_owner')
        fourth_owner = rq.get('id_fourth_owner')
        id_property = rq.get('id_property')
        sale_value = rq.get('sale_value').replace(',', '')
        sale_plan = rq.get('sale_plan')
        rate = rq.get('rate')
        to_finance = rq.get('to_finance')
        quanty_ci_quota = rq.getlist('quanty_ci_quota')
        date_ci_quota = rq.getlist('date_ci_quota')
        value_ci_quota = rq.getlist('value_ci_quota')
        quanty_to_finance_quota = rq.get('quanty_to_finance_quota')
        initial_date_to_finance_quota = rq.get('initial_date_to_finance_quota')
        value_to_finance_quota = rq.get('value_to_finance_quota')
        periodicity_extra_quota = rq.get('periodicity_extra_quota')
        quanty_extra_quota = rq.get('quanty_extra_quota')
        initial_date_extra_quota = rq.get('initial_date_extra_quota')
        value_extra_quota = rq.get('value_extra_quota')
        observations = rq.get('observations')
        club = True if rq.get('club') == 'on' else False
        
        

        consecutive = Counters.objects.get(name='contratos', project=project)

        sale_exists= Sales.objects.filter(project=project,property_sold__description=id_property,add_date=datetime.date.today(),
                                          status='Pendiente')
        
        if not sale_exists.exists():
            
        
            sale = Sales.objects.create(
                project=obj_project, contract_number=consecutive.value, first_owner=Clients.objects.get(
                    pk=first_owner),
                second_owner=Clients.objects.get(pk=second_owner),
                third_owner=Clients.objects.get(pk=third_owner),
                fourth_owner=Clients.objects.get(pk=fourth_owner),
                property_sold=Properties.objects.get(
                    description=id_property, project=obj_project.pk),
                value=sale_value, comission_base=sale_value, sale_plan=Sales_plans.objects.get(
                    pk=sale_plan),
                observations=observations, status='Pendiente', club = club
            )

            consecutive.value += 1
            consecutive.save()

            prop = Properties.objects.get(
                description=id_property, project=obj_project.pk)
            prop.state = 'Asignado'
            prop.save()

            # create a payment plan
            # initial
            counter = 1
            for i in range(0, len(quanty_ci_quota)):
                quanty = quanty_ci_quota[i]
                initial_date = date_ci_quota[i]
                dt_initial_date = datetime.datetime.strptime(
                    initial_date, '%B %d, %Y')
                value = value_ci_quota[i].replace(',', '')
                date = dt_initial_date

                for j in range(0, int(quanty)):
                    id_quota = f'CI{counter}CTR{sale.contract_number}'
                    Payment_plans.objects.create(
                        id_quota=id_quota, sale=sale, pay_date=date,
                        capital=value, interest=0, others=0,
                        project=obj_project, quota_type='CI'
                    )
                    counter += 1
                    date += relativedelta(months=1)

            # to_finance
            if quanty_to_finance_quota:
                quanty = int(quanty_to_finance_quota)
                initial_date = initial_date_to_finance_quota
                dt_initial_date = datetime.datetime.strptime(initial_date, '%B %d, %Y')
                quota = int(value_to_finance_quota.replace(',', ''))
                remaining_value = int(to_finance.replace(',', ''))
                rate_mv = float(rate)/100
                vp_regular = 0
                if periodicity_extra_quota:
                    vp_regular = int(numpy_financial.pv(rate_mv, quanty, quota))*-1
                    remaining_value = vp_regular

                date = dt_initial_date
                for i in range(1, quanty+1):
                    interest = int(remaining_value*rate_mv)
                    capital = quota - interest
                    if capital + 1000 > remaining_value:
                        capital = remaining_value
                    others = 0
                    id_quota = f'SCR{i}CTR{sale.contract_number}'
                    Payment_plans.objects.create(
                        id_quota=id_quota, sale=sale, pay_date=date,
                        capital=capital, interest=interest, others=others,
                        project=obj_project, quota_type='SCR'
                    )
                    date += relativedelta(months=1)
                    remaining_value -= capital

            if periodicity_extra_quota:
                periodicity = int(periodicity_extra_quota)
                quanty = int(quanty_extra_quota)
                initial_date = initial_date_extra_quota
                dt_initial_date = datetime.datetime.strptime(
                    initial_date, '%B %d, %Y')
                quota = int(value_extra_quota.replace(',', ''))
                remaining_value = int(to_finance.replace(',', '')) - vp_regular
                date = dt_initial_date

                for i in range(1, quanty+1):
                    interest = int(remaining_value*rate_mv*periodicity)
                    capital = quota - interest
                    if capital + 1000 > remaining_value:
                        capital = remaining_value
                    if i == quanty and capital != remaining_value:
                        capital = remaining_value
                    others = 0
                    id_quota = f'SCE{i}CTR{sale.contract_number}'
                    Payment_plans.objects.create(
                        id_quota=id_quota, sale=sale, pay_date=date,
                        capital=capital, interest=interest, others=others,
                        project=obj_project, quota_type='SCE'
                    )
                    date += relativedelta(months=periodicity)
                    remaining_value -= capital

            Sales_history.objects.create(
                sale=sale,
                action='Creó el contrato de venta',
                user=request.user
            )

            messages.success(
                request, '<div class="header">¡Excelente!</div>Creaste un contrato nuevo, puedes verlo en la sección VENTAS SIN APROBAR')
    if request.method == 'GET':
        if request.is_ajax():
            todo = request.GET.get('todo')
            
            if todo == 'getporopsforsale':
                # Obtener IDs de propiedades que están siendo usadas en ventas activas
                used_properties = Sales.objects.filter(
                    project=project,
                    status__in=['Pendiente', 'Aprobado', 'Adjudicado']
                ).values_list('property_sold_id', flat=True)

                # Filtrar propiedades: estado Libre Y no usadas en ventas activas
                prop = Properties.objects.filter(
                    project=project,
                    state='Libre'
                ).exclude(
                    id_property__in=used_properties
                ).order_by('block','location')

                json_prop = JsonRender(prop,query_functions=['description_to_search','property_price']).render()

                data = {
                    'data':json_prop
                }

                return JsonResponse(data)
            
            
    obj_project = Projects.objects.get(name=project)

    # Obtener IDs de propiedades que están siendo usadas en ventas activas
    used_properties = Sales.objects.filter(
        project=project,
        status__in=['Pendiente', 'Aprobado', 'Adjudicado']
    ).values_list('property_sold_id', flat=True)

    # Filtrar propiedades: estado Libre Y no usadas en ventas activas
    available_properties = Properties.objects.filter(
        project=project,
        state='Libre'
    ).exclude(
        id_property__in=used_properties
    ).order_by('block','location')

    context = {
        'project': obj_project,
        'properties': available_properties,
        'form_sale': newsaleForm
    }
    return render(request, 'new_sale.html', context)

@login_required
@project_permission
@user_permission('ver ventas sin aprobar')
def non_approved_sales(request, project):
    obj_project = Projects.objects.get(name=project)
    context = {
        'is_selected': False,
        'project': obj_project,
        'nonapprovedsales': Sales.objects.filter(status='Pendiente', project=project),
        'sale_has_scale': False,
    }
    if request.method == 'GET' and request.GET:
        sale = request.GET.get('sale')
        obj_sale = Sales.objects.get(contract_number=sale, project=project)
        payment_plan = Payment_plans.objects.filter(sale=obj_sale.pk)
        if obj_sale.status != 'Pendiente':
            messages.error(
                request, '<div class="header">Encontramos un problema</div>Este contrato ya no está disponible, su estado actual es '+obj_sale.status)
            return HttpResponseRedirect(f'/sales/{project}/nonapprovedsales')
        type_of_payment = 'Normal'
        if payment_plan.filter(quota_type='SCE').exists():
            type_of_payment = 'Extraordinario'

        initial_quotas = payment_plan.filter(
            quota_type='CI').order_by('pay_date', 'pk')
        initial = initial_quotas.aggregate(Sum('capital')).get('capital__sum')
        if initial == None:
            initial = 0
        initial_perc = initial*100/obj_sale.value
        initial_perc = float(f'{initial_perc:.2f}')
        to_finance = obj_sale.value - initial
        to_finance_perc = 100 - initial_perc

        # calculate  initial payment blocks
        ci_form = []
        i = 0
        block_ci_quanty = 0
        initial_block_date = ''
        block_value = ''
        for q in initial_quotas:
            if i == 0:
                block_ci_quanty = 1
                initial_block_date = q.pay_date
                block_value = f'{q.capital:,.0f}'
                if initial_quotas.count() == 1:
                    ci_form.append({
                        'quanty': block_ci_quanty,
                        'initial_date': initial_block_date,
                        'value': block_value,
                    })
            else:
                if q.capital == initial_quotas[i-1].capital:
                    block_ci_quanty += 1
                else:
                    ci_form.append({
                        'quanty': block_ci_quanty,
                        'initial_date': initial_block_date,
                        'value': block_value,
                    })

                    block_ci_quanty = 1
                    initial_block_date = q.pay_date
                    block_value = f'{q.capital:,.0f}'

                if i + 1 == initial_quotas.count():
                    ci_form.append({
                        'quanty': block_ci_quanty,
                        'initial_date': initial_block_date,
                        'value': block_value,
                    })
            i += 1

        # calculate regular quotas
        ctas_scr = payment_plan.filter(quota_type='SCR').order_by('pay_date')
        ctas_scr_q = ctas_scr.count()
        value_scr = ''
        frmt_date_scr = ''
        if ctas_scr.exists():
            value_scr = ctas_scr[0].total_payment()
            value_scr = f'{value_scr:,.0f}'
            initial_date = ctas_scr[0].pay_date
            frmt_date_scr = datetime.datetime.strftime(
                initial_date, '%B %d, %Y')

        # calculate extra quotas
        ctas_sce = payment_plan.filter(quota_type='SCE').order_by('pay_date')
        ctas_sce_q = ''
        value_sce = ''
        frmt_date_sce = ''
        periodicity = ''
        if ctas_sce.exists():
            ctas_sce_q = ctas_sce.count()
            value_sce = ctas_sce[0].total_payment()
            value_sce = f'{value_sce:,.0f}'
            initial_date = ctas_sce[0].pay_date
            frmt_date_sce = datetime.datetime.strftime(
                initial_date, '%B %d, %Y')
            periodicity = 1
            if ctas_sce_q > 1:
                second_date = ctas_sce[1].pay_date
                periodicity = relativedelta(second_date, initial_date).months

        obj_incomes = Incomes.objects.filter(
            sale=obj_sale.pk,
        ).order_by('-receipt', '-add_date')

        form = newsaleForm()
        form.initial = {
            'id_first_owner': obj_sale.first_owner.pk,
            'id_second_owner': obj_sale.second_owner.pk,
            'id_third_owner': obj_sale.third_owner.pk,
            'id_fourth_owner': obj_sale.fourth_owner.pk,
            'id_property': obj_sale.property_sold.description,
            'sale_value': f'{obj_sale.value:,}',
            'sale_plan': obj_sale.sale_plan.pk,
            'type_of_payment': type_of_payment,
            'initial_value': f'{initial:,.0f}',
            'initial_rate': initial_perc,
            'to_finance': f'{to_finance:,.0f}',
            'to_finance_rate': to_finance_perc,
            'rate': obj_sale.sale_plan.rate,
            'quanty_to_finance_quota': ctas_scr_q,
            'initial_date_to_finance_quota': frmt_date_scr,
            'value_to_finance_quota': value_scr,
            'quanty_extra_quota': ctas_sce_q,
            'initial_date_extra_quota': frmt_date_sce,
            'periodicity_extra_quota': periodicity,
            'value_extra_quota': value_sce,
            'observations': obj_sale.observations,
            'club': obj_sale.club,
        }
        form.helper.layout[0].pop(8)

        comission_assign = Assigned_comission.objects.filter(project=obj_project.pk, sale=obj_sale.pk)

        context.update({
            'is_selected': True,
            'form': form,
            'sale': obj_sale,
            'ci_form': ci_form,
            'comission_position': get_positions_queryset(obj_project, 'Publico'),
            'comission_assign': comission_assign,
            'sellers': Sellers.objects.filter(seller_state='Activo',projects__name=project).order_by('-first_name'),
            'receipts': obj_incomes,
            'sale_has_scale': comission_assign.exists(),
        })
    if request.is_ajax():
        if request.method == 'POST':
            sale = request.POST.get('sale')
            obj_sale = Sales.objects.get(contract_number=sale, project=project)
            action = request.POST.get('action')
            if action == 'modify':
                if not user_check_perms(request, 'modificar venta'):
                    data = {
                        'type': 'danger',
                        'title': 'Faltan privilegios',
                        'msj': 'Tu usuario no puede modificar un contrato',
                    }
                    return JsonResponse(data)

                parameter = Parameters.objects.filter(
                    name='modificar/reimprimir contrato', project=obj_project.pk)
                if not parameter.exists():
                    parameter = Parameters.objects.create(name='modificar/reimprimir contrato',
                                                          project=obj_project, state=0)
                else:
                    parameter = parameter[0]
                # verificate
                safe_date = obj_sale.add_date < datetime.date.today()
                permission = request.user.user_profile.has_permission(
                    'modificar venta extemporaneo')

                if safe_date and not (parameter.state or permission):
                    data = {
                        'type': 'danger',
                        'title': 'Faltan privilegios',
                        'msj': 'Tu usuario no puede modificar un contrato luego de la fecha de creación',
                    }
                    return JsonResponse(data)

                rq = request.POST
                first_owner = rq.get('id_first_owner')
                second_owner = rq.get('id_second_owner')
                third_owner = rq.get('id_third_owner')
                fourth_owner = rq.get('id_fourth_owner')
                sale_value = rq.get('sale_value').replace(',', '')
                sale_plan = rq.get('sale_plan')
                to_finance = rq.get('to_finance')
                rate = rq.get('rate')
                quanty_ci_quota = rq.getlist('quanty_ci_quota')
                date_ci_quota = rq.getlist('date_ci_quota')
                value_ci_quota = rq.getlist('value_ci_quota')
                quanty_to_finance_quota = rq.get('quanty_to_finance_quota')
                initial_date_to_finance_quota = rq.get(
                    'initial_date_to_finance_quota')
                value_to_finance_quota = rq.get('value_to_finance_quota')
                periodicity_extra_quota = rq.get('periodicity_extra_quota')
                quanty_extra_quota = rq.get('quanty_extra_quota')
                initial_date_extra_quota = rq.get('initial_date_extra_quota')
                value_extra_quota = rq.get('value_extra_quota')
                observations = rq.get('observations')
                club = True if rq.get('club') == 'on' else False

                obj_sale.first_owner = Clients.objects.get(pk=first_owner)
                obj_sale.second_owner = Clients.objects.get(pk=second_owner)
                obj_sale.third_owner = Clients.objects.get(pk=third_owner)
                obj_sale.fourth_owner = Clients.objects.get(pk=fourth_owner)
                obj_sale.value = sale_value
                obj_sale.comission_base = sale_value
                obj_sale.sale_plan = Sales_plans.objects.get(pk=sale_plan)
                obj_sale.observations = observations
                obj_sale.club = club
                obj_sale.save()

                Sales_history.objects.create(
                    sale=obj_sale,
                    action='Modificó el contrato de venta',
                    user=request.user
                )

                payment_plan = Payment_plans.objects.filter(
                    sale=obj_sale.pk, project=project
                )

                for quota in payment_plan:
                    quota.delete()

                # create a new payment plan
                # initial
                counter = 1
                for i in range(0, len(quanty_ci_quota)):
                    quanty = quanty_ci_quota[i]
                    initial_date = date_ci_quota[i]
                    dt_initial_date = datetime.datetime.strptime(
                        initial_date, '%B %d, %Y')
                    value = value_ci_quota[i].replace(',', '')
                    date = dt_initial_date

                    for j in range(0, int(quanty)):
                        id_quota = f'CI{counter}CTR{obj_sale.contract_number}'
                        Payment_plans.objects.create(
                            id_quota=id_quota, sale=obj_sale, pay_date=date,
                            capital=value, interest=0, others=0,
                            project=obj_project, quota_type='CI'
                        )
                        counter += 1
                        date += relativedelta(months=1)

                # to_finance
                quanty = int(quanty_to_finance_quota)
                initial_date = initial_date_to_finance_quota
                dt_initial_date = datetime.datetime.strptime(
                    initial_date, '%B %d, %Y')
                quota = int(value_to_finance_quota.replace(',', ''))
                remaining_value = int(to_finance.replace(',', ''))
                rate_mv = float(rate)/100
                vp_regular = 0
                if periodicity_extra_quota:
                    vp_regular = int(numpy_financial.pv(
                        rate_mv, quanty, quota))*-1
                    remaining_value = vp_regular

                date = dt_initial_date
                for i in range(1, quanty+1):
                    interest = int(remaining_value*rate_mv)
                    capital = quota - interest
                    if capital > remaining_value:
                        capital = remaining_value
                    others = 0
                    id_quota = f'SCR{i}CTR{obj_sale.contract_number}'
                    Payment_plans.objects.create(
                        id_quota=id_quota, sale=obj_sale, pay_date=date,
                        capital=capital, interest=interest, others=others,
                        project=obj_project, quota_type='SCR'
                    )
                    date += relativedelta(months=1)
                    remaining_value -= capital

                if periodicity_extra_quota:
                    periodicity = int(periodicity_extra_quota)
                    quanty = int(quanty_extra_quota)
                    initial_date = initial_date_extra_quota
                    dt_initial_date = datetime.datetime.strptime(
                        initial_date, '%B %d, %Y')
                    quota = int(value_extra_quota.replace(',', ''))
                    remaining_value = int(
                        to_finance.replace(',', '')) - vp_regular
                    date = dt_initial_date

                    for i in range(1, quanty+1):
                        interest = int(remaining_value*rate_mv*periodicity)
                        capital = quota - interest
                        if capital > remaining_value:
                            capital = remaining_value
                        if i == quanty and capital != remaining_value:
                            capital = remaining_value
                        others = 0
                        id_quota = f'SCE{i}CTR{obj_sale.contract_number}'
                        Payment_plans.objects.create(
                            id_quota=id_quota, sale=obj_sale, pay_date=date,
                            capital=capital, interest=interest, others=others,
                            project=obj_project, quota_type='SCE'
                        )
                        date += relativedelta(months=periodicity)
                        remaining_value -= capital

                data = {
                    'type': 'success',
                    'title': '¡Lo hicimos!',
                    'msj': 'Se modificó el contrato sin problemas',
                }

                return JsonResponse(data)
            elif action == 'approve':
                if not user_check_perms(request, 'aprobar venta'):
                    data = {
                        'type': 'danger',
                        'title': 'Faltan privilegios',
                        'msj': 'Tu usuario no puede aprobar un contrato',
                    }
                    return JsonResponse(data,status=403)
                obj_sale.status = 'Aprobado'
                obj_sale.save()

                Sales_history.objects.create(
                    sale=obj_sale,
                    action='Aprobó el contrato de venta',
                    user=request.user
                )
                messages.success(
                    request, '<div class="header">¡Lo hicimos!</div>Aprobaste el contrato '+sale)
                data = {
                    'type': 'success',
                    'title': '¡Lo hicimos!',
                    'msj': 'Se Aprobó el contrato sin problemas',
                }

                return JsonResponse(data)
            elif action == 'nullify':
                if not user_check_perms(request, 'anular venta'):
                    messages.error(
                        request, '<div class="header">Falta de privilegios</div>Tu usuario no tiene privilegios suficientes para anular un contrato')
                    data = {
                        'type': 'danger',
                        'title': 'Faltan privilegios',
                        'msj': 'Tu usuario no puede anular un contrato',
                    }
                    return JsonResponse(data)

                incomes = Incomes.objects.filter(
                    sale=obj_sale.pk
                )
                total_incomes = incomes.aggregate(
                    Sum('value')).get('value__sum')
                if total_incomes == None:
                    total_incomes = 0
                if total_incomes > 0:
                    messages.error(
                        request, '<div class="header">Error al anular</div>No se puede anular un contrato si tiene ingresos registrados')
                    data = {
                        'type': 'danger',
                        'title': 'Error al anular',
                        'msj': 'No se puede anular un contrato si tiene ingresos registrados',
                    }
                    return JsonResponse(data)

                obj_sale.status = 'Anulado'
                obj_sale.save()

                prop = Properties.objects.get(
                    pk=obj_sale.property_sold.pk)
                prop.state = 'Libre'
                prop.save()

                Sales_history.objects.create(
                    sale=obj_sale,
                    action='Anuló el contrato de venta',
                    user=request.user
                )
                messages.success(
                    request, '<div class="header">¡Lo hicimos!</div>Anulaste el contrato '+sale)
                data = {
                    'type': 'success',
                    'title': '¡Lo hicimos!',
                    'msj': 'Se Anuló el contrato sin problemas',
                }

                return JsonResponse(data)

    return render(request, 'non-approved.html', context)

@login_required
@project_permission
@user_permission('ver ventas por adjudicar')
def to_adjudicate_sales(request, project):
    obj_project = Projects.objects.get(pk=project)
    context = {
        'project': obj_project,
        'nonapprovedsales': Sales.objects.filter(
            project=project, status='Aprobado'
        ),
    }

    if request.method == 'GET' and request.GET:
        sale = request.GET.get('sale')
        obj_sale = Sales.objects.get(contract_number=sale, project=project)
        payment_plan = Payment_plans.objects.filter(sale=obj_sale.pk)
        if obj_sale.status != 'Aprobado':
            messages.error(
                request, '<div class="header">Encontramos un problema</div>Este contrato ya no está disponible, su estado actual es '+obj_sale.status)
            return HttpResponseRedirect(f'/sales/{project}/toadjudicatesales')
        type_of_payment = 'Normal'
        if payment_plan.filter(quota_type='SCE').exists():
            type_of_payment = 'Extraordinario'

        initial_quotas = payment_plan.filter(
            quota_type='CI').order_by('pay_date', 'pk')
        initial = initial_quotas.aggregate(Sum('capital')).get('capital__sum')
        if initial == None:
            initial = 0
        initial_perc = initial*100/obj_sale.value
        initial_perc = float(f'{initial_perc:.2f}')
        to_finance = obj_sale.value - initial
        to_finance_perc = 100 - initial_perc

        # calculate  initial payment blocks
        ci_form = []
        i = 0
        block_ci_quanty = 0
        initial_block_date = ''
        block_value = ''
        for q in initial_quotas:
            if i == 0:
                block_ci_quanty = 1
                initial_block_date = q.pay_date
                block_value = f'{q.capital:,.0f}'
                if initial_quotas.count() == 1:
                    ci_form.append({
                        'quanty': block_ci_quanty,
                        'initial_date': initial_block_date,
                        'value': block_value,
                    })
            else:
                if q.capital == initial_quotas[i-1].capital:
                    block_ci_quanty += 1
                else:
                    ci_form.append({
                        'quanty': block_ci_quanty,
                        'initial_date': initial_block_date,
                        'value': block_value,
                    })

                    block_ci_quanty = 1
                    initial_block_date = q.pay_date
                    block_value = f'{q.capital:,.0f}'

                if i + 1 == initial_quotas.count():
                    ci_form.append({
                        'quanty': block_ci_quanty,
                        'initial_date': initial_block_date,
                        'value': block_value,
                    })
            i += 1

        # calculate regular quotas
        ctas_scr = payment_plan.filter(quota_type='SCR').order_by('pay_date')
        ctas_scr_q = ctas_scr.count()
        value_scr = ''
        frmt_date_scr = ''
        if ctas_scr.exists():
            value_scr = ctas_scr[0].total_payment()
            value_scr = f'{value_scr:,.0f}'
            initial_date = ctas_scr[0].pay_date
            frmt_date_scr = datetime.datetime.strftime(
                initial_date, '%B %d, %Y')

        # calculate extra quotas
        ctas_sce = payment_plan.filter(quota_type='SCE').order_by('pay_date')
        ctas_sce_q = ''
        value_sce = ''
        frmt_date_sce = ''
        periodicity = ''
        if ctas_sce.exists():
            ctas_sce_q = ctas_sce.count()
            value_sce = ctas_sce[0].total_payment()
            value_sce = f'{value_sce:,.0f}'
            initial_date = ctas_sce[0].pay_date
            frmt_date_sce = datetime.datetime.strftime(
                initial_date, '%B %d, %Y')
            periodicity = 1
            if ctas_sce_q > 1:
                second_date = ctas_sce[1].pay_date
                periodicity = relativedelta(second_date, initial_date).months

        obj_incomes = Incomes.objects.filter(
            sale=obj_sale.pk,
        ).order_by('-receipt', '-add_date')

        form = newsaleForm()
        form.initial = {
            'id_first_owner': obj_sale.first_owner.pk,
            'id_second_owner': obj_sale.second_owner.pk,
            'id_third_owner': obj_sale.third_owner.pk,
            'id_fourth_owner': obj_sale.fourth_owner.pk,
            'id_property': obj_sale.property_sold.description,
            'sale_value': f'{obj_sale.value:,}',
            'sale_plan': obj_sale.sale_plan.pk,
            'type_of_payment': type_of_payment,
            'initial_value': f'{initial:,.0f}',
            'initial_rate': initial_perc,
            'to_finance': f'{to_finance:,.0f}',
            'to_finance_rate': to_finance_perc,
            'rate': obj_sale.sale_plan.rate,
            'quanty_to_finance_quota': ctas_scr_q,
            'initial_date_to_finance_quota': frmt_date_scr,
            'value_to_finance_quota': value_scr,
            'quanty_extra_quota': ctas_sce_q,
            'initial_date_extra_quota': frmt_date_sce,
            'periodicity_extra_quota': periodicity,
            'value_extra_quota': value_sce,
            'observations': obj_sale.observations,
            'club': obj_sale.club,
        }
        form.helper.layout[0].pop(8)

        context.update({
            'is_selected': True,
            'form': form,
            'sale': obj_sale,
            'ci_form': ci_form,
            'comission_position': get_positions_queryset(obj_project, 'Publico'),
            'comission_assign': Assigned_comission.objects.filter(project=obj_project.pk, sale=obj_sale.pk),
            'sellers': Sellers.objects.filter(seller_state='Activo').order_by('-first_name'),
            'receipts': obj_incomes,
            'adjudicate_form': adjudicate_saleForm,
        })

    if request.method == 'POST':
        user_check_perms(request, 'adjudicar venta', raise_exception=True)
        sale = request.POST.get('sale')
        obj_sale = Sales.objects.get(pk=sale)
        action = request.POST.get('action')
        if action == 'approve':
            comission_base = request.POST.get(
                'comission_base').replace(',', '')
            collector_user = request.POST.get('collector_user')
            obj_collector = User.objects.get(user_profile=collector_user)

            obj_incomes = Incomes.objects.filter(sale=obj_sale.pk)
            for income in obj_incomes:
                apply_income(income, 100)

            obj_sale.status = 'Adjudicado'
            obj_sale.comission_base = comission_base
            obj_sale.save()

            Sales_history.objects.create(
                sale=obj_sale,
                action='Adjudicó el contrato',
                user=request.user
            )

            Collector_per_sale.objects.create(
                sale=obj_sale,
                collector_user=obj_collector
            )

            Timeline.objects.create(
                user=request.user,
                action=f'Asignó a {request.user.username} como gestor de cobro del contrato Nº {obj_sale.contract_number}',
                aplication='finance',
                project=Projects.objects.get(name=project)
            )

            messages.success(
                request, f'<div class="header">¡Lo hicimos!</div>Se adjudicó el contrato {obj_sale.contract_number}')

            return HttpResponseRedirect('/sales/'+project+'/adjudicatesales')

        if action == 'disapprove':
            user_check_perms(request, 'desaprobar venta',
                             raise_exception=True)
            obj_sale.status = 'Pendiente'
            obj_sale.save()
            Sales_history.objects.create(
                sale=obj_sale,
                action='Desaprobó el contrato',
                user=request.user
            )

            messages.success(
                request, f'<div class="header">¡Lo hicimos!</div>Se desaprobó el contrato {obj_sale.contract_number}')

            return HttpResponseRedirect('/sales/'+project+'/nonapprovedsales')

    return render(request, 'to_adjudicate.html', context)

@login_required
@project_permission
@user_permission('ver ventas adjudicadas')
def adjudicate_sales(request, project):
    obj_project = Projects.objects.get(name=project)
    obj_sale = Sales_extra_info.objects.filter(
            project=project
        )
    context = {
        'project': obj_project,
        'adjudicatedsales': obj_sale.filter(status='Adjudicado'),
        'desistsales':obj_sale.filter(status='Desistido')
    }
    if request.GET:
        sale_identifier = request.GET.get('sale')
        obj_sale = None
        # Permitir identificar por PK o número de contrato, priorizando el PK para evitar colisiones
        try:
            obj_sale = Sales.objects.get(project=project, pk=int(sale_identifier))
        except (Sales.DoesNotExist, ValueError, TypeError):
            try:
                obj_sale = Sales.objects.get(project=project, contract_number=sale_identifier)
            except Sales.DoesNotExist:
                obj_sale = None
        if obj_sale is None:
            return redirect(f'/sales/{project}/adjudicatesales')
        obj_sale_credit = Sales_extra_info.objects.get(pk=obj_sale.pk)
        if user_check_perms(request,'ver comisiones adjudicadas'):
            obj_asign_comiss = Assigned_comission.objects.filter(sale=obj_sale.pk)
            obj_paid_comiss = Paid_comissions.objects.filter(
                assign_paid__sale=obj_sale
            ).select_related(
                'assign_paid__position',
                'assign_paid__seller',
                'user'
            ).order_by('-pay_date', '-pk')
            # Agrupar pagos por fecha de liquidación para mostrar en el modal
            paid_grouped = []
            current_group = None
            for pc in obj_paid_comiss:
                if current_group is None or current_group['date'] != pc.pay_date:
                    current_group = {
                        'date': pc.pay_date,
                        'items': [],
                        'total_comission': Decimal('0'),
                        'total_provision': Decimal('0'),
                        'total_net': Decimal('0'),
                    }
                    paid_grouped.append(current_group)
                current_group['items'].append(pc)
                current_group['total_comission'] += Decimal(str(pc.comission or 0))
                current_group['total_provision'] += Decimal(str(pc.provision or 0))
                current_group['total_net'] += Decimal(str(pc.net_payment or 0))
            paid_grouped = [
                dict(group, count=len(group['items']))
                for group in paid_grouped
            ]
        else:
            obj_asign_comiss = None
            obj_paid_comiss = []
            paid_grouped = []
        obj_history = Sales_history.objects.filter(
            sale=obj_sale.pk).order_by('-add_date')
        incomes_applied_qs = Incomes_detail.objects.filter(income=OuterRef('pk'))
        obj_incomes = (
            Incomes.objects.filter(sale=obj_sale.pk)
            .annotate(is_applied=Exists(incomes_applied_qs))
            .order_by('-add_date')
        )
        obj_credit_info = Credit_info.objects.filter(
            sale=obj_sale.pk).order_by('pay_date', 'pk')
        obj_collection_feed = Collection_feed.objects.filter(
            sale=obj_sale.pk).order_by('-add_date')
        obj_inc_detail = Incomes_detail.objects.filter(
                quota__sale__exact=obj_sale.pk
            ).order_by('-income__payment_date')

        orphan_details = []
        total_orphan_value = Decimal('0.00')
        if request.user.is_superuser:
            orphan_details_qs = (
                Incomes_detail.objects
                .filter(quota__sale_id=obj_sale.pk, income__sale_id__isnull=False)
                .exclude(income__sale_id=F('quota__sale_id'))
                .select_related(
                    'income__sale__first_owner',
                    'income__sale__second_owner',
                    'income__sale__third_owner',
                    'income__sale__project'
                )
            )

            for detail in orphan_details_qs:
                income_sale = detail.income.sale if detail.income_id else None

                if income_sale:
                    receipt = detail.income.receipt
                    contract = income_sale.pk
                    contract_number = income_sale.contract_number
                    sale_status = income_sale.status or 'Sin venta'
                    sale_client = income_sale.first_owner.full_name()
                    project_name = income_sale.project.name_to_show
                else:
                    receipt = 'Sin recibo'
                    contract = None
                    contract_number = None
                    sale_status = 'Sin venta'
                    sale_client = 'N/A'
                    project_name = ''

                value = detail.total_income()
                total_orphan_value += value

                orphan_details.append({
                    'detail_id': detail.pk,
                    'receipt': receipt,
                    'value': value,
                    'contract': contract,
                    'contract_number': contract_number,
                    'sale_status': sale_status,
                    'sale_client': sale_client,
                    'project_name': project_name,
                })
        
        pendient_quota_by_type = obj_sale_credit.rv_by_type_of_quota()
        extra_orphans = max(len(orphan_details) - 5, 0)

        context.update({
            'is_selected': True,
            'clients': Clients.objects.all().order_by('first_name'),
            'sale': obj_sale,
            'extra_info': obj_sale_credit,
            'comissions_scale': obj_asign_comiss,
            'paid_comissions': obj_paid_comiss,
            'paid_comissions_grouped': paid_grouped,
            'history': obj_history,
            'incomes': obj_incomes,
            'credit_info': obj_credit_info,
            'collection_feed': obj_collection_feed,
            'form_collection_feed': collectionfeed_Form,
            'form_change_prop': change_property_Form(project=project),
            'incomes_detail': obj_inc_detail,
            'form_change_plan': change_plan_Form(
                initial={
                    'actual_capital': f"{obj_sale_credit.basic_info().get('remaining_value',0):,.0f}",
                    'actual_plan': obj_sale.sale_plan.name,
                    'actual_value': f'{obj_sale.value:,.0f}',
                    'ci_to_change': f"{pendient_quota_by_type.get('pendient_ci'):,.0f}",
                    'finance_to_change': f"{pendient_quota_by_type.get('pendient_finance'):,.0f}",
                    'rate': obj_sale.sale_plan.rate,
                }
            ),
            'orphan_details': orphan_details,
            'orphan_details_total': total_orphan_value,
            'orphan_details_count': len(orphan_details),
            'orphan_details_extra': extra_orphans,
        })

    return render(request, 'adjudicate_sales.html', context)


@login_required
@project_permission
@user_permission('ver inventario')
def properties_for_sales(request,project):
    obj_project = Projects.objects.get(name=project)
    obj_properties = Properties.objects.filter(project=project)
    stages = obj_properties.values('stage').order_by('stage').annotate(total=Count('stage'))
    blocks = obj_properties.values('block').order_by('block').annotate(total=Count('block'))
    
    if request.is_ajax():
        if request.method == 'GET':
            search_by = request.GET.get('search_by')
            aggrupate_value = request.GET.get('value')
            history_action = ""
            
            if search_by == 'stage':
                obj_properties_filtered = obj_properties.filter(stage=aggrupate_value)
            elif search_by == 'block':
                obj_properties_filtered = obj_properties.filter(block=aggrupate_value)
            elif search_by == 'property':
                obj_properties_filtered = obj_properties.filter(description=aggrupate_value)
            else:
                obj_properties_filtered = obj_properties
            data = {
                'data':JsonRender(obj_properties_filtered).render()
            }
            return JsonResponse(data)
        
        if request.method == 'POST':
            to_change = request.POST.get('to_change')
            value = request.POST.get('value')
            to_do = request.POST.get('to_do')
            
            history_action = ""
            type_of = 'error'
            title = 'Ups'
            msj= 'Los parametros enviados no coinciden con los requeridos'
            
            translat = {
                'stage':'la Etapa',
                'block': 'la Manzana',
                'property':'el Lote'
            }
            
            obj_prop = Properties.objects.filter(project=project)
            if to_change == 'stage':
                properties_to_change = obj_prop.filter(stage = value)
            elif to_change == 'block':
                properties_to_change = obj_prop.filter(block = value)
            elif to_change == 'property':
                properties_to_change = obj_prop.filter(description = value)
            
            cant_do = ''
            if to_do  == 'relase':
                user_check_perms(request,'liberar inventario',raise_exception=True)

                # Obtener IDs de propiedades que están siendo usadas en ventas activas
                used_properties_ids = Sales.objects.filter(
                    project=project,
                    status__in=['Pendiente', 'Aprobado', 'Adjudicado']
                ).values_list('property_sold_id', flat=True)

                for prop in properties_to_change:
                    # Doble check: verificar estado 'Asignado' O si está en ventas activas
                    if prop.state == 'Asignado' or prop.id_property in used_properties_ids:
                        cant_do += prop.description +' '
                    else:
                        prop.state = 'Libre'
                        prop.save()
                        history_action = f'Liberó {translat[to_change]} {value}'

                if cant_do == '':
                    title = 'Petición lista!'
                    msj = 'Listo, los lotes fueron liberados sin problemas'
                    type_of = 'success'

                else:
                    title = 'Peticion con novedades'
                    msj = 'Listo, tu petición fue realizada pero los siguientes lotes no se liberaron porque están asignados o en ventas activas: '+cant_do
                    type_of = 'yellow'                
                
            elif to_do == 'block':
                check = user_check_perms(request,'bloquear inventario',raise_exception=True)
                
                
                for prop in properties_to_change:
                    if prop.state == 'Asignado':
                        cant_do += prop.description +' '
                    else:
                        prop.state = 'Bloqueado'
                        prop.save()
                        history_action = f'Bloqueó {translat[to_change]} {value}'
                        
                if cant_do == '':
                    title = 'Petición lista!'
                    msj = 'Listo, los lotes fueron bloqueados sin problemas'
                    type_of = 'success'
                else:
                    title = 'Peticion con novedades'
                    msj = 'Tu petición fue realizada pero los siguientes lotes no se bloquearon porque ya estan asignados: '+cant_do
                    type_of = 'yellow'
                
                
            elif to_do == 'changeprice':
                user_check_perms(request,'cambiar precio inventario',raise_exception=True)
                new_value = request.POST.get('new_price').replace(',','')
                for prop in properties_to_change:
                    prop.m2_price = new_value
                    prop.save()
                    
                title = 'Petición lista!'
                msj = 'Listo, el valor de los lotes fué actualizado sin problemas'
                type_of = 'success'
                history_action = f'Cambio el precio del m2 de {translat[to_change]} {value} a {int(new_value):,}'
            
            if history_action != "":
                Timeline.objects.create(
                        user = request.user,
                        action = history_action,
                        project = obj_project,
                        aplication = 'sales'
                    )
            
            data = {
                'type':type_of,
                'title':title,
                'msj':msj
            }
            
            return JsonResponse(data)
        
    
    context = {
        'project':obj_project,
        'properties':obj_properties.order_by('description'),
        'stages':stages,
        'blocks':blocks,
    }
    
    return render(request,'properties.html',context)

from .models import Sales, SalesFiles
from .forms import SalesFileForm

@login_required
@project_permission
def get_sales_files(request, project, contract_number):
    """
    Endpoint para obtener los documentos asociados a una venta por contract_number y proyecto.
    """
    if request.method == 'GET' and request.is_ajax():
        obj_project = Projects.objects.get(name=project)
        try:
            sale = Sales.objects.get(contract_number=contract_number, project=obj_project)
        except Sales.DoesNotExist:
            return JsonResponse({
                'status': 'error', 
                'message': f'No se encontró la venta con contrato #{contract_number} en el proyecto {project}'
            }, status=404)
        
        files = SalesFiles.objects.filter(sale=sale, is_active=True).order_by('-upload_date')

        files_data = [
            {
                'id': file.id_file,
                'description': file.description,
                'file_url': file.file.url,
                'file_type': file.file_type,
                'observations': file.observations,
                'uploaded_by': file.uploaded_by.username,
                'upload_date': file.upload_date.strftime('%Y-%m-%d %H:%M:%S'),
            }
            for file in files
        ]

        return JsonResponse({'status': 'success', 'files': files_data})
    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

@login_required
@project_permission
def add_sales_file(request, project, contract_number):
    if request.method == 'POST' and request.is_ajax():
        obj_project = Projects.objects.get(name=project)
        sale = get_object_or_404(Sales, contract_number=contract_number, project=obj_project)
        form = SalesFileForm(request.POST, request.FILES)
        if form.is_valid():
            sales_file = form.save(commit=False)
            sales_file.sale = sale
            sales_file.uploaded_by = request.user
            sales_file.save()
            return JsonResponse({
                'status': 'success',
                'message': 'Archivo cargado correctamente.',
                'file': {
                    'id': sales_file.id_file,
                    'description': sales_file.description.capitalize(),
                    'file_url': sales_file.file.url,
                    'uploaded_by': sales_file.uploaded_by.username,
                    'upload_date': sales_file.upload_date.strftime('%Y-%m-%d %H:%M:%S'),
                    'file_type': sales_file.file_type.capitalize(),
                    'observations': sales_file.observations,
                }
            })
        return JsonResponse({'status': 'error', 'message': 'Formulario inválido.'}, status=400)

@login_required
@project_permission
@user_permission('eliminar archivos de ventas')
def delete_sales_file(request, project, file_id):
    if request.method == 'POST' and request.is_ajax():
        obj_project = Projects.objects.get(name=project)
        sales_file = get_object_or_404(SalesFiles, pk=file_id, sale__project=obj_project)
        sales_file.is_active = False
        sales_file.save()
        return JsonResponse({
            'status': 'success',
            'message': 'Archivo eliminado correctamente.',
            'file_id': file_id
        })
        
@login_required
def sales_plans_list(request):
    """
    Vista para listar los planes de venta.
    """
    plans = Sales_plans.objects.filter(status=True).order_by('name')
    return render(request, 'sales/sales_plans_list.html', {'plans': plans})

@login_required
def sales_plan_create(request):
    """
    Vista para crear un nuevo plan de venta.
    """
    if request.method == 'POST':
        form = SalesPlanForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('sales_plans_list')
    else:
        form = SalesPlanForm()
    return render(request, 'sales/sales_plan_form.html', {'form': form})

@login_required
def sales_plan_edit(request, plan_id):
    """
    Vista para editar un plan de venta existente.
    """
    plan = get_object_or_404(Sales_plans, pk=plan_id)
    if request.method == 'POST':
        form = SalesPlanForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            return redirect('sales_plans_list')
    else:
        form = SalesPlanForm(instance=plan)
    return render(request, 'sales/sales_plan_form.html', {'form': form})

@login_required
def sales_plan_delete(request, plan_id):
    """
    Vista para eliminar un plan de venta.
    """
    plan = get_object_or_404(Sales_plans, pk=plan_id)
    plan.status = False  # Eliminación lógica
    plan.save()
    return redirect('sales_plans_list')

@login_required
@user_permission('ver ventas adjudicadas')
def plan_pagos_detalle(request, sale_id):
    sale = get_object_or_404(Sales, pk=sale_id)
    
    # Verifica si hay una reestructuración pendiente para esta venta
    if PaymentPlanRestructuring.objects.filter(sale=sale, status='Pendiente').exists():
        messages.error(request, "No puedes modificar el plan de pagos de esta venta porque existe una reestructuración pendiente para este contrato.")
        return redirect('listado_reestructuraciones', project=sale.project.name)

    cuotas = Payment_plans.objects.filter(sale=sale).order_by('pay_date', 'quota_type')
    fecha_actual = datetime.datetime.now().date()

    saldo_teorico = Decimal(str(sale.value))
    saldo_real = Decimal(str(sale.value))
    cuotas_con_saldo = []

    for cuota in cuotas:
        pagado = Decimal(str(cuota.paid()))
        total = Decimal(str(cuota.total_payment()))
        pendiente = total - pagado
        capital = cuota.capital  # Decimal
        capital_pagado = cuota.capital_paid()  # Solo lo efectivamente pagado a capital
        capital_vigente = max(capital - Decimal(str(capital_pagado)), Decimal('0.00'))

        saldo_teorico -= capital
        saldo_real -= capital_pagado
        
        saldo_teorico_actual = saldo_teorico
        saldo_real_actual = saldo_real

        cuotas_con_saldo.append({
            'cuota': cuota,
            'pagado': pagado,
            'pendiente': pendiente,
            'capital_vigente': capital_vigente,
            'saldo_teorico': max(saldo_teorico_actual, Decimal('0.00')),
            'saldo_real': max(saldo_real_actual, Decimal('0.00')),
            'es_vencida': cuota.pay_date < fecha_actual and pendiente > 0,
            'es_pagada': pagado >= total
        })

    # Ordenar: pagadas arriba (True primero), luego por fecha
    cuotas_con_saldo.sort(key=lambda x: (not x['es_pagada'], x['cuota'].pay_date))

    rate_decimal = Decimal(str(sale.sale_plan.rate or 0))
    latest_restruct = PaymentPlanRestructuring.objects.filter(
        sale=sale, status='Aprobado'
    ).order_by('-approved_at', '-id_restructuring').first()
    if latest_restruct and latest_restruct.tasa and Decimal(str(latest_restruct.tasa)) > Decimal('0'):
        rate_decimal = Decimal(str(latest_restruct.tasa))
    elif sale.tasa and Decimal(str(sale.tasa)) > Decimal('0'):
        rate_decimal = Decimal(str(sale.tasa))

    context = {
        'sale': sale,
        'cuotas': cuotas_con_saldo,
        'saldo_total': saldo_real,
        'preferred_rate': rate_decimal,
        'preferred_rate_js': float(rate_decimal),
    }
    
    return render(request, 'plan_pagos_detalle.html', context)

@login_required
@user_permission('modificar plan de pagos')
def editar_cuota(request, cuota_id):
    """Vista AJAX para editar una cuota específica"""
    if not request.is_ajax() or request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
    
    cuota = get_object_or_404(Payment_plans, pk=cuota_id)
    sale = cuota.sale
    
    # Creamos backup antes de modificar
    backup_plan_pagos(sale)
    
    try:
        # Obtenemos los valores del formulario
        nuevos_valores = {
            'capital': request.POST.get('capital'),
            'interest': request.POST.get('interest', 0) if cuota.quota_type != 'CI' else 0,
            'others': request.POST.get('others', 0),
        }
        
        # Recalculamos el plan
        recalcular_plan_pagos(sale.pk, cuota.pk, nuevos_valores)
        
        # Obtenemos las cuotas actualizadas para devolver al frontend
        cuotas_actualizadas = Payment_plans.objects.filter(
            sale=sale, 
            pay_date__gte=cuota.pay_date
        ).order_by('pay_date')
        
        # Preparamos los datos para devolver como JSON
        data = []
        saldo_acumulado = 0
        for c in cuotas_actualizadas:
            pendiente = c.total_payment() - c.paid()
            saldo_acumulado += pendiente
            data.append({
                'id': c.pk,
                'capital': float(c.capital),
                'interest': float(c.interest),
                'others': float(c.others),
                'total': float(c.total_payment()),
                'pendiente': float(pendiente),
                'saldo_acumulado': float(saldo_acumulado)
            })
        
        return JsonResponse({
            'status': 'success',
            'message': 'Cuota actualizada correctamente',
            'cuotas': data
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Error al actualizar cuota: {str(e)}'
        }, status=400)

from django.utils.decorators import method_decorator
from django.views import View
import json
from .models import Sales, PaymentPlanRestructuring, PaymentPlanRestructuringDetail

@method_decorator([login_required, user_permission('crear reestructuracion')], name='dispatch')
class GuardarReestructuracionView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            sale_id = data.get('sale_id')
            cuotas = data.get('cuotas', [])
            observations = data.get('observations', '')
            tasa = data.get('tasa', None)
            nuevo_valor_venta = data.get('nuevo_valor_venta', None)

            sale = Sales.objects.get(pk=sale_id)
            # Asegura que la tasa se guarde como Decimal si viene informada
            tasa_decimal = None
            if tasa is not None and tasa != '':
                try:
                    tasa_decimal = Decimal(str(tasa))
                except InvalidOperation:
                    return JsonResponse({'success': False, 'error': 'Formato de tasa inválido'})
            
            # Procesar el nuevo valor de venta - CORREGIDO
            nuevo_valor_decimal = None  # Definir la variable antes de usarla
            if nuevo_valor_venta is not None and nuevo_valor_venta != '':
                try:
                    nuevo_valor_decimal = Decimal(str(nuevo_valor_venta))
                except (InvalidOperation, ValueError):
                    return JsonResponse({'success': False, 'error': 'Formato de valor de venta inválido'})

            restructuring = PaymentPlanRestructuring.objects.create(
                sale=sale,
                created_by=request.user,
                observations=observations,
                status='Pendiente',
                tasa=tasa_decimal,
                nuevo_valor_venta=nuevo_valor_decimal
            )
            detalles = []
            for c in cuotas:
                # Determina el tipo según el campo es_pagada
                if c.get('es_pagada', False):
                    tipo = 'pagada'
                else:
                    tipo = 'pendiente'
                detalles.append(PaymentPlanRestructuringDetail(
                    restructuring=restructuring,
                    id_quota=c['id_quota'],
                    quota_type=c['quota_type'],
                    pay_date=c['pay_date'],
                    capital=c['capital'] or 0,
                    interest=c['interest'] or 0,
                    others=c['others'] or 0,
                    tipo=tipo
                ))
            PaymentPlanRestructuringDetail.objects.bulk_create(detalles)
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
from django.views.decorators.http import require_POST

""" @login_required
@require_POST
def aprobar_reestructuracion(request, id):
    try:
        restructuring = PaymentPlanRestructuring.objects.get(pk=id)
        if restructuring.status == 'Pendiente':
            restructuring.status = 'Aprobado'
            restructuring.save()
            return JsonResponse({'success': True})
        return JsonResponse({'success': False, 'error': 'Ya procesada'})
    except PaymentPlanRestructuring.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No encontrada'}) """

@login_required
@require_POST
@user_permission('aprobar reestructuracion')
def rechazar_reestructuracion(request, id):
    try:
        restructuring = PaymentPlanRestructuring.objects.get(pk=id)
        if restructuring.status == 'Pendiente':
            restructuring.status = 'Rechazado'
            restructuring.save()
            return JsonResponse({'success': True})
        return JsonResponse({'success': False, 'error': 'Ya procesada'})
    except PaymentPlanRestructuring.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No encontrada'})

from django.db import transaction
from django.utils import timezone

@login_required
@require_POST
@user_permission('aprobar reestructuracion')
def aprobar_reestructuracion(request, id):
    try:
        with transaction.atomic():
            restructuring = PaymentPlanRestructuring.objects.select_related('sale').prefetch_related('details').get(pk=id)
            if restructuring.status != 'Pendiente':
                return JsonResponse({'success': False, 'error': 'Ya procesada'})

            sale = restructuring.sale

            # 1. Backup del plan actual
            planes_actuales = list(Payment_plans.objects.filter(sale=sale).order_by('pay_date', 'pk'))
            backup_objs = [
                backup_payment_plans(
                    backup_date=timezone.now(),
                    id_payment=plan.id_payment,
                    id_quota=plan.id_quota,
                    quota_type=plan.quota_type,
                    sale=plan.sale,
                    pay_date=plan.pay_date,
                    capital=plan.capital,
                    interest=plan.interest,
                    others=plan.others,
                    project=plan.project
                )
                for plan in planes_actuales
            ]
            backup_payment_plans.objects.bulk_create(backup_objs)

            # 2. Actualizar tasa y valor de venta si están definidos en la reestructuración
            if hasattr(restructuring, 'tasa') and restructuring.tasa is not None:
                sale.tasa = restructuring.tasa
            
            # Actualizar el valor de la venta si se especificó un nuevo valor
            if (hasattr(restructuring, 'nuevo_valor_venta') and 
                restructuring.nuevo_valor_venta is not None and 
                abs(restructuring.nuevo_valor_venta - sale.value) > 0.01):  # Diferencia mayor a 1 centavo
                
                # Guardar el valor original para el historial
                valor_original = sale.value
                
                # Actualizar el valor de la venta
                sale.value = restructuring.nuevo_valor_venta
                
                # Registrar el cambio en el historial solo si realmente cambió
                Sales_history.objects.create(
                    sale=sale,
                    action=f'Se modificó el valor de la venta de ${valor_original:,.2f} a ${restructuring.nuevo_valor_venta:,.2f} mediante reestructuración #{restructuring.id_restructuring}',
                    user=request.user
                )
            
            # Guardar los cambios en la venta
            sale.save()
            
            # 3. Actualizar, eliminar o crear cuotas según corresponda (por id_quota)
            detalles = list(restructuring.details.all().order_by('pay_date', 'pk'))
            planes_actuales_dict = {p.id_quota: p for p in planes_actuales}
            detalles_dict = {d.id_quota: d for d in detalles}
            
            # Actualizar las cuotas que existen en ambos
            for id_quota, detalle in detalles_dict.items():
                if id_quota in planes_actuales_dict:
                    plan = planes_actuales_dict[id_quota]
                    plan.quota_type = detalle.quota_type
                    plan.pay_date = detalle.pay_date
                    plan.capital = detalle.capital
                    plan.interest = detalle.interest
                    plan.others = detalle.others
                    plan.project = sale.project
                    plan.save()
                else:
                    # Crear nuevas cuotas si no existen en el plan original
                    Payment_plans.objects.create(
                        id_quota=detalle.id_quota,
                        quota_type=detalle.quota_type,
                        sale=sale,
                        pay_date=detalle.pay_date,
                        capital=detalle.capital,
                        interest=detalle.interest,
                        others=detalle.others,
                        project=sale.project
                    )

            # Eliminar cuotas que están en el plan original pero no en la reestructuración
            for id_quota, plan in planes_actuales_dict.items():
                if id_quota not in detalles_dict:
                    plan.delete()

            # 4. Marcar como aprobada la reestructuración
            restructuring.status = 'Aprobado'
            restructuring.approved_by = request.user
            restructuring.approved_at = timezone.now()
            restructuring.save()
            
            # 5. Agregar entrada al historial de ventas
            Sales_history.objects.create(
                sale=sale,
                action=f'Se aprobó una reestructuración del plan de pagos para el contrato CTR{sale.contract_number}',
                user=request.user
            )

        return JsonResponse({'success': True})
    except PaymentPlanRestructuring.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No encontrada'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@project_permission
@user_permission('ver graficos')
def graphs(request,project):
    obj_project = Projects.objects.get(pk=project)
       
    context = {
        'project':obj_project    
    }
    
    if request.is_ajax():
        if request.method == 'GET':
            graph_type = request.GET.get('type')
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            
            dt_from = datetime.datetime.strptime(date_from,'%Y-%m-%d')
            dt_to = datetime.datetime.strptime(date_to,'%Y-%m-%d')
            
            
            period = None
            counter = 0
            
            if graph_type == 'sales':
                initial = dt_from + relativedelta(months=counter)
                #final_date = intial.
                """ obj_sale = Sales.objects.filter(
                    project = project,add_date__gte=dt_from,
                    add_date__lte=dt_to
                ).annotate(
                    total_pendientes= Count('pk',Q(status='Pendiente')),
                    valor_pendientes= Sum('value',Q(status='Pendiente')),
                    total_aprobado= Count('pk',Q(status='Aprobado')),
                    valor_aprobado= Sum('value',Q(status='Aprobado')),
                    total_adjudicado= Count('pk',Q(status='Adjudicado')),
                    valor_adjudicado= Sum('value',Q(status='Adjudicado')),
                    total_desistido= Count('pk',Q(status='Desistido')),
                    valor_desistido= Sum('value',Q(status='Desistido')),
                ) """
                
                
                
                data = {
                    'data':list()#obj_sale)
                }
                
                
                return JsonResponse(data)
                
                
                
    
    return render(request,'graphs.html',context)

# ajax

def ajax_change_owners(request,project):
    if request.is_ajax():
        if request.method == 'POST':
            user_check_perms(request,'aplicar cesion de titulares', raise_exception=True)
            
            contract_number = request.POST.get('contract_number')
            
            first_prop = request.POST.get('new_first_client')
            sec_prop = request.POST.get('new_second_client')
            third_prop = request.POST.get('new_third_client')
            fourth_prop = request.POST.get('new_fourth_client')
            
            obj_sale = Sales.objects.get(pk = contract_number)
            de = f'[{obj_sale.first_owner.full_name()}, {obj_sale.second_owner.full_name()}, {obj_sale.third_owner.full_name()}]'
            
            obj_first_prop = Clients.objects.get(pk=first_prop)
            obj_sec_prop = Clients.objects.get(pk=sec_prop)
            obj_third_prop = Clients.objects.get(pk=third_prop)
            obj_fourth_prop = Clients.objects.get(pk=fourth_prop)
            
            obj_sale.first_owner = obj_first_prop
            obj_sale.second_owner = obj_sec_prop
            obj_sale.third_owner = obj_third_prop
            obj_sale.fourth_owner = obj_fourth_prop
            
            obj_sale.save()

            a = f'[{obj_first_prop.full_name()}, {obj_sec_prop.full_name()},{obj_third_prop.full_name()},{obj_fourth_prop.full_name()}]'

            
            action = f'Realizó un cambio en uno o mas titulares del contrato así: de {de} a {a}'
            
            Sales_history.objects.create(
                sale = obj_sale, user = request.user,
                action = action
            )
            
            data ={
                
            } 
            
            return JsonResponse(data)

def ajax_get_plans_info(request):
    if request.method == 'GET':
        plan = request.GET.get('plan')
        plan_name = request.GET.get('plan_name')
        obj_plan = Sales_plans.objects.filter(pk=plan)
        if plan_name:
            obj_plan = Sales_plans.objects.filter(name=plan_name)
        data = {
            'data': JsonRender(obj_plan).render()
        }

        return JsonResponse(data)
    return JsonResponse({'status', 'error'})

def ajax_comissions(request, project, sale):
    obj_project = Projects.objects.get(name=project)
    obj_sale = Sales.objects.get(project=obj_project.pk, contract_number=sale)
    obj_comissions = Assigned_comission.objects.filter(
        project=obj_project.pk, sale=obj_sale.pk,
    )
    if request.is_ajax():
        if request.method == 'POST':
            positions = request.POST.getlist('position')
            sellers_id = request.POST.getlist('seller_id')
            comissions = request.POST.getlist('comission')

            action = 'Asignó escala comisionable'

            if obj_comissions.exists():
                action = 'Modificó escala comisionable'
                for i in obj_comissions:
                    i.delete()

            private_positions = get_positions_queryset(obj_project, 'Privado', include_default=True)
            msj = ''
            for pos in private_positions:
                id_comission = f'{pos.pk}CTR{sale}'
                if pos.default == None:
                    msj += f'El cargo privado <b>{pos.name}</b> no tiene asesor por defecto asignado.<br>'
                    continue
                
                Assigned_comission.objects.create(
                    id_comission=id_comission,
                    project=obj_project,
                    sale=obj_sale,
                    position=pos,
                    seller=pos.default,
                    comission=pos.rate,
                )

            for i in range(0, len(positions)):
                obj_position = get_position_for_project(obj_project, positions[i])
                obj_seller = Sellers.objects.get(pk=sellers_id[i])
                id_comission = f'{obj_position.pk}CTR{sale}'

                Assigned_comission.objects.create(
                    id_comission=id_comission,
                    project=obj_project,
                    sale=obj_sale,
                    position=obj_position,
                    seller=obj_seller,
                    comission=comissions[i]
                )

            Sales_history.objects.create(
                sale=obj_sale,
                action=action,
                user=request.user,
            )

            data = {
                'type': 'success',
                'title': '¡Lo hicimos!',
                'msj': 'Se cargó la escala registrada'
            }
            
            if msj != '':
                data = {
                    'type': 'warning',
                    'title': '¡Lo hicimos!, pero tuvimos algunos inconvenientes',
                    'msj': f'Se cargó la escala registrada, pero algunos cargos privados no se pudieron registrar de forma automatica:<br>{msj}<b>Consejo: </b>Puedes asignarlos en el sitio administrativo y volver a grabar la escala en esta pagina.'
                }

            return JsonResponse(data)

    return JsonResponse({'status': 'No data'})

def ajax_print_documents(request, project):
    obj_project = Projects.objects.get(name=project)
    if request.method == "GET" and request.GET:
        sale = request.GET.get('sale')
        doc_type = request.GET.get('type')
        obj_sale = Sales.objects.get(project=project, contract_number=sale)
        
        if doc_type == 'comissions-report':
            comissions_executive_object = Assigned_comission.objects.filter(sale=obj_sale.pk,
                                                            position__group='Publico'
                                                    ).exclude(
                                                        Q(position = 1) | Q(position = 2) | Q(position = 3)
                                                    )
            
            total_exec = comissions_executive_object.aggregate(total=Sum('comission')).get('total')
            total_exec = 0 if total_exec==None else total_exec
            
            total_comiss_exec = f'{total_exec * obj_sale.comission_base/100:,.0f}'
            
            comissions_basic_object = Assigned_comission.objects.filter(
                                                            Q(position = 1) | Q(position = 2) | Q(position = 3),
                                                            sale=obj_sale.pk,
                                                            position__group='Publico'
                                                    )
            
            total_basic = comissions_basic_object.aggregate(total=Sum('comission')).get('total')
            total_basic = 0 if total_basic==None else total_basic

            total_comiss_basic = f'{total_basic * obj_sale.comission_base/100:,.0f}'
            
            template = f'pdf/sales_comissions_report.html'
            context = {
                'sale': obj_sale,
                'project':obj_project,
                'comissions_exec':comissions_executive_object,
                'total_exec':total_comiss_exec,
                'comissions_basic':comissions_basic_object,
                'total_basic':total_comiss_basic,
                'now':datetime.datetime.now(),
                'user':request.user.username,
                
            }
            
            filename = f'Hoja_de_radicacion_CTR{sale}_{obj_project.name_to_show}.pdf'.replace('ñ','n')
            pdf = pdf_gen(template, context, filename)

            pdf_url = pdf.get('url')
            href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'msj': 'Puedes descargarlo haciendo click '+href
            }
            
            return JsonResponse(data)
        
        obj_plan = Payment_plans.objects.filter(sale=obj_sale.pk)
        initial_quotas = obj_plan.filter(
            quota_type='CI').order_by('pay_date', 'pk')
        initial = initial_quotas.aggregate(total=Sum('capital')).get('total')
        if initial == None:
            initial = 0
        ci_form = []
        i = 0
        block_ci_quanty = 0
        initial_block_date = ''
        block_value = ''
        for q in initial_quotas:
            if i == 0:
                block_ci_quanty = 1
                initial_block_date = q.pay_date
                block_value = f'{q.capital:,.0f}'
                if initial_quotas.count() == 1:
                    ci_form.append({
                        'quanty': block_ci_quanty,
                        'initial_date': initial_block_date,
                        'value': block_value,
                    })
            else:
                if q.capital == initial_quotas[i-1].capital:
                    block_ci_quanty += 1
                else:
                    ci_form.append({
                        'quanty': block_ci_quanty,
                        'initial_date': initial_block_date,
                        'value': block_value,
                    })

                    block_ci_quanty = 1
                    initial_block_date = q.pay_date
                    block_value = f'{q.capital:,.0f}'
                if i +1 == initial_quotas.count():
                    ci_form.append({
                        'quanty': block_ci_quanty,
                        'initial_date': initial_block_date,
                        'value': block_value,
                })
            i += 1
            
        total_to_finance = obj_sale.value - initial

        ctas_scr = obj_plan.filter(quota_type='SCR').order_by('pay_date')
        ctas_scr_q = ctas_scr.count()
        value_scr = ''
        frmt_date_scr = ''
        if ctas_scr.exists():
            value_scr = ctas_scr[0].total_payment()
            value_scr = f'{value_scr:,.0f}'
            initial_date = ctas_scr[0].pay_date
            frmt_date_scr = ctas_scr[0].pay_date

        scr_form = {
            'ctas_scr_q': ctas_scr_q,
            'value_scr': value_scr,
            'initial_date': frmt_date_scr
        }

        # calculate extra quotas
        ctas_sce = obj_plan.filter(quota_type='SCE').order_by('pay_date')
        ctas_sce_q = ''
        value_sce = ''
        frmt_date_sce = ''
        periodicity = ''
        if ctas_sce.exists():
            ctas_sce_q = ctas_sce.count()
            value_sce = ctas_sce[0].total_payment()
            value_sce = f'{value_sce:,.0f}'
            initial_date = ctas_sce[0].pay_date
            frmt_date_sce = ctas_sce[0].pay_date

            periodicity = 1
            if ctas_sce_q > 1:
                second_date = ctas_sce[1].pay_date
                periodicity = relativedelta(second_date, frmt_date_sce).months

        periods = {
            1: 'Mensual',
            3: 'Trimestral',
            6: 'Semestral',
            12: 'Anual'
        }

        sce_form = {
            'ctas_sce_q': ctas_sce_q,
            'value_sce': value_sce,
            'initial_date': frmt_date_sce,
            'periodicity': periods.get(periodicity, ''),
        }

        plan = {
            'total_initial_payment': initial,
            'total_to_finance': total_to_finance,
            'ci_form': ci_form,
            'scr_form': scr_form,
            'sce_form': sce_form,
        }

        if doc_type == 'contract':
            template = f'{obj_project.formats_path}/contrato.html'
            if not obj_sale.club and project == 'altoscovenas':
                template = f'{obj_project.formats_path}/contrato_sin_club.html'
            context = {
                
                'sale': obj_sale,
                'plan': plan,
            }
            filename = f'Oferta_Comercial_CTR{sale}_{obj_project.name_to_show}.pdf'.replace('ñ','n')
            pdf = pdf_gen(template, context, filename)

            pdf_url = pdf.get('url')
            href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'msj': 'Puedes descargarlo haciendo click '+href
            }

            return JsonResponse(data)

        elif doc_type == 'verification':
            
            template = f'{obj_project.formats_path}/formulario_verificacion.html'
            if not obj_sale.club and project == 'altoscovenas':
                template = f'{obj_project.formats_path}/formulario_verificacion_sin_club.html'
                
            context = {
                'sale': obj_sale,
                'plan': plan,
            }
            filename = f'Formulario_Verificacion_CTR{sale}_{obj_project.name_to_show}.pdf'.replace('ñ','n')
            pdf = pdf_gen(template, context, filename)

            pdf_url = pdf.get('url')
            href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'msj': 'Puedes descargarlo haciendo click '+href
            }

            return JsonResponse(data)

        elif doc_type == 'pagare':
            template = f'{obj_project.formats_path}/pagare.html'
            context = {
                'sale': obj_sale,
                'plan': plan,
            }
            filename = f'Pagare_CTR{sale}_{obj_project.name_to_show}.pdf'.replace('ñ','n')
            pdf = pdf_gen(template, context, filename)

            pdf_url = pdf.get('url')
            href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'msj': 'Puedes descargarlo haciendo click '+href
            }

            return JsonResponse(data)

        elif doc_type == 'statement_of_account':
            template = f'pdf/statement_of_account.html'

            credit_info_qs = Credit_info.objects.filter(sale=obj_sale.pk).order_by('pay_date', 'pk')
            today = datetime.date.today()

            # Resumen de pagos por cuota para evitar N+1
            income_summary = (
                Incomes_detail.objects
                .filter(quota__sale=obj_sale.pk)
                .values('quota_id')
                .annotate(
                    paid_capital=Coalesce(Sum('capital'), Value(0, output_field=ModelDecimalField(max_digits=20, decimal_places=2))),
                    paid_interest=Coalesce(Sum('interest'), Value(0, output_field=ModelDecimalField(max_digits=20, decimal_places=2))),
                    paid_others=Coalesce(Sum('others'), Value(0, output_field=ModelDecimalField(max_digits=20, decimal_places=2))),
                    paid_arrears=Coalesce(Sum('arrears'), Value(0, output_field=ModelDecimalField(max_digits=20, decimal_places=2))),
                    last_pay=Max('income__payment_date', filter=Q(capital__gt=0) | Q(interest__gt=0)),
                    last_only_arrears=Max('income__payment_date', filter=Q(capital=0, interest=0, arrears__gt=0)),
                )
            )
            income_summary = {
                item['quota_id']: {
                    'paid_capital': Decimal(str(item['paid_capital'] or 0)),
                    'paid_interest': Decimal(str(item['paid_interest'] or 0)),
                    'paid_others': Decimal(str(item['paid_others'] or 0)),
                    'paid_arrears': Decimal(str(item['paid_arrears'] or 0)),
                    'last_pay': item['last_pay'],
                    'last_only_arrears': item['last_only_arrears'],
                }
                for item in income_summary
            }

            mora_param = Parameters.objects.filter(name='tasa de mora mv').first()
            mora_rate = Decimal(str(mora_param.value)) if mora_param and mora_param.value is not None else Decimal('0')

            expired_rows = []
            prox_rows = []
            plan_rows = []
            total_value = Decimal('0')
            total_arrears = Decimal('0')
            total_value_prox = Decimal('0')

            for quota in credit_info_qs:
                summary = income_summary.get(quota.pk, {
                    'paid_capital': Decimal('0'),
                    'paid_interest': Decimal('0'),
                    'paid_others': Decimal('0'),
                    'paid_arrears': Decimal('0'),
                    'last_pay': None,
                    'last_only_arrears': None,
                })
                paid_capital = summary['paid_capital']
                paid_interest = summary['paid_interest']
                paid_others = summary['paid_others']
                paid_arrears = summary['paid_arrears']

                pending_capital = quota.capital - paid_capital
                pending_interest = quota.interest - paid_interest
                pending_others = quota.others - paid_others
                pending_total = pending_capital + pending_interest + pending_others
                if pending_total < 0:
                    pending_total = Decimal('0')

                last_pay = summary['last_pay'] or quota.pay_date
                last_only_arrears = summary['last_only_arrears']

                arrears_days = 0
                arrears_value = Decimal('0')
                if quota.pay_date < today and pending_total > 0:
                    if last_only_arrears and last_only_arrears > last_pay:
                        days = (today - last_pay).days
                        total_arrears_calc = pending_total * Decimal(days) * (mora_rate / Decimal('30')) / Decimal('100')
                        arrears_value = total_arrears_calc - paid_arrears
                        arrears_days = max(days, 0)
                    else:
                        if last_pay < quota.pay_date:
                            last_pay = quota.pay_date
                        days = (today - last_pay).days
                        arrears_days = max(days, 0)
                        arrears_value = pending_total * Decimal(arrears_days) * (mora_rate / Decimal('30')) / Decimal('100')

                    if arrears_value < 0:
                        arrears_value = Decimal('0')

                total_due = pending_total + arrears_value
                paid_total = paid_capital + paid_interest + paid_others

                row = {
                    'pay_date': quota.pay_date,
                    'id_quota': quota.id_quota,
                    'capital': quota.capital,
                    'interest': quota.interest,
                    'paid_capital': paid_capital,
                    'paid_interest': paid_interest,
                    'pending_capital': pending_capital,
                    'pending_interest': pending_interest,
                    'pending_total': pending_total,
                    'arrears_value': arrears_value,
                    'arrears_days': arrears_days,
                    'total_due': total_due,
                    'paid_total': paid_total,
                    'total_payment': quota.total_payment(),
                }

                plan_rows.append(row)

                if pending_total > 0 and quota.pay_date <= today:
                    expired_rows.append(row)
                    total_value += pending_total
                    total_arrears += arrears_value

                days_to_pay = (quota.pay_date - today).days
                if pending_total > 0 and 0 < days_to_pay <= 30:
                    prox_rows.append(row)
                    total_value_prox += pending_total

            totals = {
                'total_value': total_value,
                'total_arrears': total_arrears,
                'total_value_prox': total_value_prox,
                'total_general': total_value + total_arrears,
            }

            context = {
                'sale': obj_sale,
                'plan_resume': plan,
                'sale_extra_info': Sales_extra_info.objects.get(pk=obj_sale.pk),
                'now': datetime.datetime.now(),
                'credit_info': credit_info_qs,
                'plan_rows': plan_rows,
                'expired_rows': expired_rows,
                'prox_quotas': prox_rows,
                'totals': totals,
                'user': request.user
            }

            filename = f'Estado_de_cuenta_CTR{sale}_{obj_project.name_to_show}.pdf'.replace('ñ','n')
            pdf = pdf_gen(template, context, filename)

            pdf_url = pdf.get('url')
            href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'msj': 'Puedes descargarlo haciendo click '+href
            }

            return JsonResponse(data)

        elif doc_type == 'collection_letter':
            template = f'pdf/collection_letter_1.html'
            credit_info = Credit_info.objects.filter(sale=obj_sale.pk)
            total_ci = 0
            for quota in credit_info.filter(quota_type='CI'):
                if quota.is_expired():
                    pending = quota.quota_pending().get('total_pending')
                    if pending > 0:
                        total_ci += pending
            total_s = 0
            for quota in credit_info.exclude(quota_type='CI'):
                if quota.is_expired():
                    pending = quota.quota_pending().get('total_pending')
                    if pending > 0:
                        total_s += pending
            
            context = {
                'sale': obj_sale,
                'plan': plan,
                'credit_info':credit_info,
                'total_ci':total_ci,
                'total_s':total_s,
                'today':datetime.date.today()
            }
            filename = f'Carta_cobro_CTR{sale}_{obj_project.name_to_show}.pdf'.replace('ñ','n')
            pdf = pdf_gen(template, context, filename)

            pdf_url = pdf.get('url')
            href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'msj': 'Puedes descargarlo haciendo click '+href
            }

            return JsonResponse(data)
         
        elif doc_type == 'admission_application':
            template = f'pdf/solicitud_admision.html'
            
            context = {
                'sale': obj_sale,
                'project': obj_project,
                'plan': plan,
                'fecha_actual': datetime.date.today(),
            }
            
            filename = f'Solicitud_Admision_CTR{sale}_{obj_project.name_to_show}.pdf'.replace('ñ','n')
            pdf = pdf_gen(template, context, filename)

            pdf_url = pdf.get('url')
            href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'msj': 'Puedes descargarlo haciendo click '+href
            }

            return JsonResponse(data)

        elif doc_type == 'data_authorization':
            # Generar una hoja por cada titular (hasta 4)
            titulares = [
                {'nombre': obj_sale.first_owner.full_name(), 'cc': obj_sale.first_owner.client_document},
                {'nombre': obj_sale.second_owner.full_name(), 'cc': obj_sale.second_owner.client_document} if obj_sale.second_owner else None,
                {'nombre': obj_sale.third_owner.full_name(), 'cc': obj_sale.third_owner.client_document} if obj_sale.third_owner else None,
                {'nombre': obj_sale.fourth_owner.full_name(), 'cc': obj_sale.fourth_owner.client_document} if obj_sale.fourth_owner else None,
            ]
            titulares = [t for t in titulares if t and t['nombre'] and t['cc']]

            context = {
                'titulares': titulares,
                'fecha': datetime.date.today().strftime('%d/%m/%Y'),
                'ciudad': 'Medellín',
            }
            template = 'pdf/data_authorization.html'
            filename = f'Autorizacion_Datos_CTR{sale}_{obj_project.name_to_show}.pdf'.replace('ñ','n')
            pdf = pdf_gen(template, context, filename)

            pdf_url = pdf.get('url')
            href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

            data = {
                'type': 'success',
                'title': '¡Documento generado!',
                'msj': 'Puedes descargarlo haciendo click ' + href
            }
            return JsonResponse(data)

        return JsonResponse({'response': 'Debes enviar un tipo de documento'})

def ajax_change_property(request, project):
    if request.is_ajax():
        if request.method == 'POST':
            if not user_check_perms(request, 'cambiar inmueble'):
                data = {
                    'type': 'error',
                    'title': 'Error de privilegios',
                    'msj': 'Tu usuario no tiene permisos suficientes para hacer esto.'
                }
                return JsonResponse(data)

            sale = request.POST.get('sale')
            new_prop = request.POST.get('new_property')

            obj_sale = Sales.objects.get(pk=sale)
            old_prop = obj_sale.property_sold
            obj_new_prop = Properties.objects.get(
                pk=new_prop
            )
            obj_sale.property_sold = obj_new_prop
            obj_sale.save()

            old_prop.state = 'Libre'
            old_prop.save()

            obj_new_prop.state = 'Asignado'
            obj_new_prop.save()

            Sales_history.objects.create(
                sale=obj_sale,
                action=f'Cambió el inmueble vendido   , de {old_prop.description} a {obj_new_prop.description}',
                user=request.user
            )

            data = {
                'type': 'success',
                'title': 'Lo hicimos',
                'msj': 'El inmueble vendido fue cambiado'
            }

            messages.success(
                request, '<div class="header">¡Lo hicimos!</div>El inmueble vendido fue cambiado')

            return JsonResponse(data)

def ajax_reestructurate_payment(request, sale):
    obj_sale = Sales.objects.get(pk=sale)
    if request.is_ajax():
        if request.method == 'POST':
            if not user_check_perms(request, 'cambiar plan de pagos'):
                data = {
                    'status': 403,
                    'type': 'error',
                    'title': 'Oh Oh!',
                    'msj': 'Tu usuario no tiene los privilegios suficientes para modificar un plan de pagos.'
                }
                return JsonResponse(data)

            type_of_change = request.POST.get('type_of_change')

            to_finance = request.POST.get('finance_to_change')
            rate = request.POST.get('rate')
            quanty_ci_quota = request.POST.getlist('quanty_ci_quota')
            date_ci_quota = request.POST.getlist('date_ci_quota')
            value_ci_quota = request.POST.getlist('value_ci_quota')
            quanty_to_finance_quota = request.POST.get(
                'quanty_to_finance_quota')
            initial_date_to_finance_quota = request.POST.get(
                'initial_date_to_finance_quota')
            value_to_finance_quota = request.POST.get('value_to_finance_quota')
            periodicity_extra_quota = request.POST.get(
                'periodicity_extra_quota')
            quanty_extra_quota = request.POST.get('quanty_extra_quota')
            initial_date_extra_quota = request.POST.get(
                'initial_date_extra_quota')
            value_extra_quota = request.POST.get('value_extra_quota')

            credit_info = Credit_info.objects.filter(sale=sale)
            ci_info = credit_info.filter(
                quota_type='CI'
            ).order_by('pay_date')
            scr_info = credit_info.filter(
                quota_type='SCR'
            ).order_by('pay_date')
            sce_info = credit_info.filter(
                quota_type='SCE'
            ).order_by('pay_date')

            # Saving actual payment plan in a backup table in error case can be restored
            obj_payment = Payment_plans.objects.filter(sale=sale)
            backup_datetime = datetime.datetime.now()
            for quota in obj_payment:
                backup_payment_plans.objects.create(
                    id_payment=quota.pk, id_quota=quota.id_quota,
                    quota_type=quota.quota_type, sale=obj_sale,
                    pay_date=quota.pay_date, capital=quota.capital,
                    interest=quota.interest, others=quota.others,
                    project=obj_sale.project, backup_date=backup_datetime
                )
            # ci
            if len(quanty_ci_quota) > 0:
                counter = 1
                for q in ci_info:

                    if q.how_paid().get('full_pendient'):
                        q.delete()
                    elif q.how_paid().get('partial_paid'):
                        q.capital = q.quota_balance().get('paid_capital')
                        q.others = q.quota_balance().get('paid_others')
                        q.save()
                        counter += 1
                    elif q.how_paid().get('full_paid'):
                        counter += 1

                for i in range(0, len(quanty_ci_quota)):
                    quanty = quanty_ci_quota[i]
                    initial_date = date_ci_quota[i]
                    dt_initial_date = datetime.datetime.strptime(
                        initial_date, '%B %d, %Y')
                    value = value_ci_quota[i].replace(',', '')
                    date = dt_initial_date

                    for j in range(0, int(quanty)):
                        id_quota = f'CI{counter}CTR{obj_sale.contract_number}'
                        Payment_plans.objects.create(
                            id_quota=id_quota, sale=obj_sale, pay_date=date,
                            capital=value, interest=0, others=0,
                            project=obj_sale.project, quota_type='CI'
                        )
                        counter += 1
                        date += relativedelta(months=1)

            # to finance
            if type_of_change != 'initial':
                
                if quanty_to_finance_quota:
                    counter_scr = 1
                    for q in scr_info:

                        if q.how_paid().get('full_pendient'):
                            q.delete()
                        elif q.how_paid().get('partial_paid'):
                            q.capital = q.quota_balance().get('paid_capital')
                            q.interest = q.quota_balance().get('paid_interest')
                            q.others = q.quota_balance().get('paid_others')
                            q.save()
                            counter_scr += 1
                        elif q.how_paid().get('full_paid'):
                            counter_scr += 1
                            
                    counter_sce = 1
                    for q in sce_info:

                        if q.how_paid().get('full_pendient'):
                            q.delete()
                        elif q.how_paid().get('partial_paid'):
                            q.capital = q.quota_balance().get('paid_capital')
                            q.interest = q.quota_balance().get('paid_interest')
                            q.others = q.quota_balance().get('paid_others')
                            q.save()
                            counter_sce += 1
                        elif q.how_paid().get('full_paid'):
                            counter_sce += 1
                                            
                    """ if type_of_change == 'to_capital_pay': 
                        capital_pay_quota = counter_sce
                        counter_sce += 1 """
                    
                    
                    quanty = int(quanty_to_finance_quota)
                    initial_date = initial_date_to_finance_quota
                    dt_initial_date = datetime.datetime.strptime(
                        initial_date, '%B %d, %Y')
                    quota = int(value_to_finance_quota.replace(',', ''))
                    remaining_value = int(to_finance.replace(',', ''))
                    rate_mv = float(rate)/100

                    if quanty_extra_quota:
                        vp_regular = int(numpy_financial.pv(
                            rate_mv, quanty, quota))*-1
                        remaining_value = vp_regular

                    date = dt_initial_date
                    for i in range(counter_scr, counter_scr+quanty):
                        interest = int(remaining_value*rate_mv)
                        capital = quota - interest
                        if capital + 1000> remaining_value:
                            capital = remaining_value
                        if i == counter_scr+quanty -1 and capital != remaining_value:
                            capital = remaining_value
                        others = 0
                        id_quota = f'SCR{i}CTR{obj_sale.contract_number}'
                        Payment_plans.objects.create(
                            id_quota=id_quota, sale=obj_sale, pay_date=date,
                            capital=capital, interest=interest, others=others,
                            project=obj_sale.project, quota_type='SCR'
                        )
                        date += relativedelta(months=1)
                        remaining_value -= capital

                    if quanty_extra_quota:
                        periodicity = int(periodicity_extra_quota)
                        quanty = int(quanty_extra_quota)
                        initial_date = initial_date_extra_quota
                        dt_initial_date = datetime.datetime.strptime(
                            initial_date, '%B %d, %Y')
                        quota = int(value_extra_quota.replace(',', ''))
                        remaining_value = int(
                            to_finance.replace(',', '')) - vp_regular
                        date = dt_initial_date
                        
                        for i in range(counter_sce, counter_sce+quanty -1):
                            interest = int(remaining_value*rate_mv)
                            capital = quota - interest
                            if capital + 1000> remaining_value:
                                capital = remaining_value
                            if i == counter_sce+quanty -1 and capital != remaining_value:
                                capital = remaining_value
                            others = 0
                            id_quota = f'SCE{i}CTR{obj_sale.contract_number}'
                            Payment_plans.objects.create(
                                id_quota=id_quota, sale=obj_sale, pay_date=date,
                                capital=capital, interest=interest, others=others,
                                project=obj_sale.project, quota_type='SCE'
                            )
                            date += relativedelta(months=periodicity)
                            remaining_value -= capital


            if type_of_change == 'credit_value':
                old_value = obj_sale.value
                new_sale_value = request.POST.get('new_sale_value')
                obj_sale.value = new_sale_value.replace(',', '')
                old_plan = obj_sale.sale_plan.name
                new_plan = request.POST.get('new_payment_plan')
                obj_new_plan = Sales_plans.objects.get(pk=new_plan)
                if old_plan != obj_new_plan.name:
                    obj_sale.sale_plan = obj_new_plan
                    obj_sale.save()
                Sales_history.objects.create(
                    sale=obj_sale,
                    action=f'Modificó el plan de pagos, con cambio en el valor de venta (de ${old_value:,} a {new_sale_value})',
                    user=request.user
                )
            elif type_of_change == 'payment_plan':
                old_plan = obj_sale.sale_plan.name
                new_plan = request.POST.get('new_payment_plan')
                obj_new_plan = Sales_plans.objects.get(pk=new_plan)
                obj_sale.sale_plan = obj_new_plan
                obj_sale.save()
                Sales_history.objects.create(
                    sale=obj_sale,
                    action=f'Modificó el plan de pagos, con cambio en el tipo de plan (de {old_plan} a {obj_new_plan.name})',
                    user=request.user
                )
                
            else:
                Sales_history.objects.create(
                    sale=obj_sale,
                    action=f'Modificó el plan de pagos',
                    user=request.user
                )
            
            messages.success(
                request, '<div class="header">¡Lo hicimos!</div>Se reestructuró el plan de pagos sin problemas.')

            data = {
                'status': 200,
                'type': 'success',
                'title': 'Hecho!',
                'msj': 'Se reestructuró el plan de pagos sin problemas.'
            }
            return JsonResponse(data)

def ajax_change_comissions(request, sale):
    obj_sale = Sales.objects.get(pk=sale)
    if request.is_ajax():
        if request.method == 'POST':
            user_check_perms(request, 'editar comisiones adjudicadas',raise_exception=True)
            positions = request.POST.getlist('position')
            comissions = request.POST.getlist('comission')
            states = request.POST.getlist('state')

            quanty = len(positions)

            for i in range(0, quanty):
                obj_position = Assigned_comission.objects.get(pk=positions[i])
                obj_position.comission = comissions[i]
                obj_position.state = states[i]
                obj_position.save()

            Sales_history.objects.create(
                user=request.user,
                sale=obj_sale,
                action='Realizó una modificación en la escala de comisiones'
            )

            data = {
                'message': {
                    'type': 'success',
                    'title': 'Lo hicimos!',
                    'msj': 'Se actualizó la escala de comisiones'
                }
            }

            return JsonResponse(data, status=200)

def ajax_desist_sale(request, sale):
    if request.is_ajax():
        if request.method == 'POST':
            obj_sale = Sales.objects.get(pk=sale)
            todo = request.POST.get('todo')
            if todo == 'desist':
                if user_check_perms(request, 'desistir venta'):
                    date = request.POST.get('date')
                    if date != None and date != '':
                        date = parse_semantic_date(date,'date')
                    value = int(request.POST.get('value').replace(',','')) * -1
                    
                    obj_sale.status = 'Desistido'
                    obj_sale.save()

                    prop = obj_sale.property_sold
                    prop.state = 'Libre'
                    prop.save()

                    Sales_history.objects.create(
                        user=request.user,
                        sale=obj_sale,
                        action='Desistió el contrato'
                    )
                    
                    
                    if value < 0:
                        Incomes_return.objects.create(sale=obj_sale, date=date,
                                                                  value=value, user=request.user)
                    

                    
                    messages.success(
                        request, '<div class="header">¡Lo hicimos!</div>Se desistió el contrato')

                    data = {

                    }
                    status = 200
                else:
                    data = {
                        'type': 'error',
                        'title': 'Ups!',
                        'msj': 'No tienes los privilegios suficientes para desistir una venta'
                    }
                    status = 403

                return JsonResponse(data, status=status)
            
def ajax_change_dates_adj(request,project,sale):
    if request.method == 'POST':
        if user_check_perms(request, 'cambiar fechas de cuotas'):
            type_of = request.POST.get('type-of-change')
            from_cta = request.POST.get('from_cta')
            value = int(request.POST.get('value_to_change'))
            
            obj_sale = Sales.objects.get(pk=sale)
            
            obj_from_cta = Payment_plans.objects.get(
                id_quota = from_cta, project = project
            )
            
            obj_ctas = Payment_plans.objects.filter(
                sale = sale, project = project,
                pay_date__gte = obj_from_cta.pay_date
            ).order_by('pay_date')
            action = ''
            if type_of == 'change-day':
                for cta in obj_ctas:
                    cta.pay_date = cta.pay_date.replace(day=value)
                    cta.save()
                action=f'Cambió el dia de pago a partir de la cuota {from_cta} al dia {value}'
                
            elif type_of == 'add-months':
                for cta in obj_ctas:
                    cta.pay_date = cta.pay_date + relativedelta(months=value)
                    cta.save()
                action=f'Movió la fecha de pago en {value} meses a partir de la cuota {from_cta}'

            
            
            Sales_history.objects.create(
                        sale=obj_sale,
                        action=action,
                        user=request.user
                    )
            
            messages.success(request,'<div class="header">¡Lo hicimos!</div>Se aplico un cambio de fechas en cuotas')

            data = {}
            status = 200
        else:
            data = {
                'type': 'error',
                'title': 'Ups!',
                'msj': 'No tienes los privilegios suficientes para cambiar fechas de cuotas'
            }
            status = 403

        return JsonResponse(data, status=status)
    
@login_required
@project_permission
@user_permission('ver ventas adjudicadas')
def generar_plan_pagos_pdf(request, project, sale_id):
    """Generar PDF del plan de pagos"""
    obj_project = Projects.objects.get(name=project)
    obj_sale = Sales.objects.get(pk=sale_id, project=obj_project)
    
    # Obtener plan de pagos
    obj_plan = Payment_plans.objects.filter(sale=obj_sale).order_by('pay_date', 'pk')
    
    # Calcular resumen del plan
    initial_quotas = obj_plan.filter(quota_type='CI').order_by('pay_date', 'pk')
    initial = initial_quotas.aggregate(total=Sum('capital')).get('total') or 0
    
    # Bloques de cuotas iniciales
    ci_form = []
    i = 0
    block_ci_quanty = 0
    initial_block_date = ''
    block_value = ''
    for q in initial_quotas:
        if i == 0:
            block_ci_quanty = 1
            initial_block_date = q.pay_date
            block_value = f'{q.capital:,.0f}'
            if initial_quotas.count() == 1:
                ci_form.append({
                    'quanty': block_ci_quanty,
                    'initial_date': initial_block_date,
                    'value': block_value,
                })
        else:
            if q.capital == initial_quotas[i-1].capital:
                block_ci_quanty += 1
            else:
                ci_form.append({
                    'quanty': block_ci_quanty,
                    'initial_date': initial_block_date,
                    'value': block_value,
                })
                block_ci_quanty = 1
                initial_block_date = q.pay_date
                block_value = f'{q.capital:,.0f}'
            if i + 1 == initial_quotas.count():
                ci_form.append({
                    'quanty': block_ci_quanty,
                    'initial_date': initial_block_date,
                    'value': block_value,
                })
        i += 1
    
    # Cuotas regulares
    ctas_scr = obj_plan.filter(quota_type='SCR').order_by('pay_date')
    scr_form = {}
    if ctas_scr.exists():
        scr_form = {
            'count': ctas_scr.count(),
            'value': f'{ctas_scr[0].total_payment():,.0f}',
            'initial_date': ctas_scr[0].pay_date
        }
    
    # Cuotas extraordinarias
    ctas_sce = obj_plan.filter(quota_type='SCE').order_by('pay_date')
    sce_form = {}
    if ctas_sce.exists():
        periodicity = 1
        if ctas_sce.count() > 1:
            second_date = ctas_sce[1].pay_date
            periodicity = relativedelta(second_date, ctas_sce[0].pay_date).months
        
        periods = {1: 'Mensual', 3: 'Trimestral', 6: 'Semestral', 12: 'Anual'}
        sce_form = {
            'count': ctas_sce.count(),
            'value': f'{ctas_sce[0].total_payment():,.0f}',
            'initial_date': ctas_sce[0].pay_date,
            'periodicity': periods.get(periodicity, ''),
        }
    
    total_to_finance = obj_sale.value - initial
    
    plan_resume = {
        'total_initial_payment': initial,
        'total_to_finance': total_to_finance,
        'ci_form': ci_form,
        'scr_form': scr_form,
        'sce_form': sce_form,
    }
    import re
    payment_plan = [
        {
            'quota_type': 'Cuota Inicial' if quota.quota_type.startswith('CI') else 'Saldo',
            'id_quota': re.search(r'\d+', quota.id_quota).group(),  # Extrae solo los números
            'pay_date': quota.pay_date,
            'capital': quota.capital,
            'interest': quota.interest,
            'others': quota.others,
            'total_payment': quota.total_payment,
        }
        for quota in obj_plan
    ]
    context = {
        'sale': obj_sale,
        'project': obj_project,
        'plan_resume': plan_resume,
        'payment_plan': payment_plan,
        'now': datetime.datetime.now(),
        'user': request.user
    }
    
    template = 'pdf/plan_pagos.html'
    filename = f'Plan_Pagos_CTR{obj_sale.contract_number}_{obj_project.name_to_show}.pdf'.replace('ñ','n')
    
    pdf = pdf_gen(template, context, filename)
    
    if request.is_ajax():
        return JsonResponse({
            'type': 'success',
            'title': '¡Documento generado!',
            'msj': f'<a href="{pdf.get("url")}" target="_blank"><strong>Descargar Plan de Pagos</strong></a>'
        })
    
    # Leer el archivo generado desde la ruta en 'root'
    pdf_path = pdf.get('root')
    if pdf_path:
        with open(pdf_path, 'rb') as pdf_file:
            return HttpResponse(pdf_file.read(), content_type='application/pdf')
    
    # Si no se encuentra el archivo, retornar un error
    return HttpResponse("Error al generar el PDF", status=500)

@login_required
@project_permission
@user_permission('ver ventas adjudicadas')
def generar_recibos_pdf(request, project, sale_id):
    """Generar PDF del listado de recibos"""
    obj_project = Projects.objects.get(name=project)
    obj_sale = Sales.objects.get(pk=sale_id, project=obj_project)
    
    # Obtener recibos/ingresos
    obj_incomes = Incomes.objects.filter(sale=obj_sale).order_by('payment_date', 'receipt')
    
    # Obtener detalles de ingresos
    obj_inc_detail = Incomes_detail.objects.filter(
        quota__sale=obj_sale
    ).order_by('-income__payment_date', 'quota__pay_date')
    
    # Calcular totales
    total_ingresos = obj_incomes.aggregate(total=Sum('value')).get('total') or 0
    total_capital = obj_inc_detail.aggregate(total=Sum('capital')).get('total') or 0
    total_intereses = obj_inc_detail.aggregate(total=Sum('interest')).get('total') or 0
    total_otros = obj_inc_detail.aggregate(total=Sum('others')).get('total') or 0
    total_mora = obj_inc_detail.aggregate(total=Sum('arrears')).get('total') or 0
    
    # Obtener información extra de la venta
    try:
        sale_extra_info = Sales_extra_info.objects.get(pk=obj_sale.pk)
        saldo_pendiente = sale_extra_info.remain_value()
    except:
        saldo_pendiente = 0
    
    context = {
        'sale': obj_sale,
        'project': obj_project,
        'incomes': obj_incomes,
        'incomes_detail': obj_inc_detail,
        'totals': {
            'total_ingresos': total_ingresos,
            'total_capital': total_capital,
            'total_intereses': total_intereses,
            'total_otros': total_otros,
            'total_mora': total_mora,
            'saldo_pendiente': saldo_pendiente
        },
        'now': datetime.datetime.now(),
        'user': request.user
    }
    
    template = 'pdf/listado_recibos.html'
    filename = f'Listado_Recibos_CTR{obj_sale.contract_number}_{obj_project.name_to_show}.pdf'.replace('ñ','n')
    
    pdf = pdf_gen(template, context, filename)
    
    if request.is_ajax():
        print('response al ajax')
        return JsonResponse({
            'type': 'success',
            'title': '¡Documento generado!',
            'msj': f'<a href="{pdf.get("url")}" target="_blank"><strong>Descargar Listado de Recibos</strong></a>'
        })
        
    # Leer el archivo generado desde la ruta en 'root'
    pdf_path = pdf.get('root')
    if pdf_path:
        with open(pdf_path, 'rb') as pdf_file:
            return HttpResponse(pdf_file.read(), content_type='application/pdf')
    
    # Si no se encuentra el archivo, retornar un error
    return HttpResponse("Error al generar el PDF", status=500)

@login_required
@user_permission('ver reestructuraciones')
def listado_reestructuraciones(request, project):
    obj_project = Projects.objects.get(name=project)
    
    queryset = PaymentPlanRestructuring.objects.select_related('sale', 'created_by').filter(sale__project__name=project)
    

    restructuraciones = queryset.order_by('-created_at')

    return render(request, 'listado_reestructuraciones.html', {
        'restructuraciones': restructuraciones,
        'project': obj_project
    })
    
from django.template.loader import render_to_string

@login_required
def ajax_detalle_reestructuracion(request, id):
    restructuring = PaymentPlanRestructuring.objects.select_related('sale', 'created_by').prefetch_related('details').get(pk=id)
    detalles = restructuring.details.all().order_by('pay_date', 'pk')
    saldo = float(restructuring.sale.value)
    detalles_con_saldo = []
    for d in detalles:
        saldo -= float(d.capital or 0)  # Restar el capital de la cuota actual
        detalles_con_saldo.append({
            'detalle': d,
            'saldo': saldo  # Este es el saldo después de pagar la cuota
        })
    html = render_to_string(
        'detalle_reestructuracion_modal.html',
        {'restructuring': restructuring, 'detalles_con_saldo': detalles_con_saldo}
    )
    return HttpResponse(html)

@login_required
@project_permission
@user_permission('ver entregas y escrituracion')
def delivery_dates_list(request, project):
    """Lista de fechas de entrega organizadas por fecha"""
        
    # ✅ CORRECCIÓN: Usar los campos correctos del modelo Sales
    sales = Sales.objects.select_related('first_owner', 'property_sold').filter(
        property_sold__project__name=project,  # Cambiar de inmueble a property_sold
        status='Adjudicado'  # Cambiar de 'vendido' a 'Adjudicado'
    ).exclude(
        scheduled_delivery_date__isnull=True
    )
    
    # ✅ CORRECCIÓN: El modelo Sales no tiene campo advisor, usar otro criterio
    # if es_gestor and not es_lider:
    #     sales = sales.filter(advisor=request.user)
    
    # Ordenar por fecha programada de entrega (más reciente primero)
    sales = sales.order_by('-scheduled_delivery_date', '-actual_delivery_date')
    
    # Procesar datos para la tabla
    delivery_data = []
    for sale in sales:
        # Determinar estado de entrega
        is_delivered = sale.actual_delivery_date is not None
        delivery_status = 'Entregado' if is_delivered else 'Pendiente'
        
        # Determinar estado de escrituración
        is_deed_signed = sale.actual_deed_date is not None
        deed_status = 'Escriturado' if is_deed_signed else 'Pendiente'
        
        # Calcular días de retraso/adelanto
        delivery_delay = None
        if sale.scheduled_delivery_date:
            if is_delivered:
                delivery_delay = (sale.actual_delivery_date - sale.scheduled_delivery_date).days
            else:
                from datetime import date
                delivery_delay = (date.today() - sale.scheduled_delivery_date).days
        
        deed_delay = None
        if sale.scheduled_deed_date:
            if is_deed_signed:
                deed_delay = (sale.actual_deed_date - sale.scheduled_deed_date).days
            else:
                from datetime import date
                deed_delay = (date.today() - sale.scheduled_deed_date).days
        
        delivery_data.append({
            'sale': sale,
            'is_delivered': is_delivered,
            'delivery_status': delivery_status,
            'delivery_delay': delivery_delay,
            'is_deed_signed': is_deed_signed,
            'deed_status': deed_status,
            'deed_delay': deed_delay,
        })
    
    # ✅ CORRECCIÓN: Obtener el objeto project
    obj_project = Projects.objects.get(name=project)
    
    # Calcular estadísticas
    entregadas = sum(1 for item in delivery_data if item['is_delivered'])
    escrituradas = sum(1 for item in delivery_data if item['is_deed_signed'])
    
    context = {
        'delivery_data': delivery_data,
        'project': obj_project,
        'estadisticas': {
            'total': len(delivery_data),
            'entregadas': entregadas,
            'escrituradas': escrituradas,
        }
    }
    
    return render(request, 'delivery_dates_list.html', context)

@login_required
@project_permission
@user_permission('registrar entregas y escrituracion')
def ajax_delivery_dates(request, project, sale_id):
    try:
        sale = Sales.objects.get(pk=sale_id)
        
        if request.method == 'GET':
            # Retornar datos existentes
            data = {
                'scheduled_delivery_date': sale.scheduled_delivery_date.strftime('%d/%m/%Y') if sale.scheduled_delivery_date else None,
                'actual_delivery_date': sale.actual_delivery_date.strftime('%d/%m/%Y') if sale.actual_delivery_date else None,
                'delivery_observations': sale.delivery_observations or '',
                'scheduled_deed_date': sale.scheduled_deed_date.strftime('%d/%m/%Y') if sale.scheduled_deed_date else None,
                'actual_deed_date': sale.actual_deed_date.strftime('%d/%m/%Y') if sale.actual_deed_date else None,
                'notary': sale.notary or '',
                'deed_number': sale.deed_number or '',
                'deed_observations': sale.deed_observations or '',
            }
            return JsonResponse({'success': True, 'data': data})
        
        elif request.method == 'POST':
            # Función auxiliar para parsear fechas
            def parse_date_safe(date_str):
                """Acepta dd/mm/YYYY, ISO o el formato semantic UI por compatibilidad."""
                if not date_str or date_str.strip() == '':
                    return None
                cleaned = date_str.strip()
                for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
                    try:
                        return datetime.datetime.strptime(cleaned, fmt).date()
                    except ValueError:
                        continue
                # Último intento: formato largo en inglés (Semantic UI)
                return parse_semantic_date(cleaned,'date')
            
            # Actualizar fechas usando la función auxiliar
            sale.scheduled_delivery_date = parse_date_safe(request.POST.get('scheduled_delivery_date'))
            sale.actual_delivery_date = parse_date_safe(request.POST.get('actual_delivery_date'))
            sale.delivery_observations = request.POST.get('delivery_observations', '')
            sale.scheduled_deed_date = parse_date_safe(request.POST.get('scheduled_deed_date'))
            sale.actual_deed_date = parse_date_safe(request.POST.get('actual_deed_date'))
            sale.notary = request.POST.get('notary', '')
            sale.deed_number = request.POST.get('deed_number', '')
            sale.deed_observations = request.POST.get('deed_observations', '')
            
            sale.save()
            
            return JsonResponse({'success': True, 'message': 'Fechas actualizadas correctamente'})
            
    except Sales.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Venta no encontrada'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

from django.views.decorators.http import require_POST

@login_required
@project_permission
@user_permission('ver planes de venta')
def sales_plans_project_list(request, project):
    obj_project = Projects.objects.get(name=project)
    plans = Sales_plans.objects.filter(project=obj_project).order_by('name') if hasattr(Sales_plans, 'project') else Sales_plans.objects.all().order_by('name')
    # Si no tienes campo project en Sales_plans, filtra solo por status=True
    # plans = Sales_plans.objects.filter(status=True).order_by('name')
    for plan in plans:
        plan.adjudicaciones_count = Sales.objects.filter(sale_plan=plan, project=obj_project, status='Adjudicado').count()
    return render(request, 'sales_plans_project_list.html', {
        'plans': plans,
        'project': obj_project
    })

@login_required
@project_permission
@user_permission('modificar plan de venta')
@require_POST
def sales_plan_toggle_status(request, plan_id):
    plan = get_object_or_404(Sales_plans, pk=plan_id)
    plan.status = not plan.status
    plan.save()
    return JsonResponse({'success': True, 'status': plan.status})

@login_required
@project_permission
@user_permission('modificar plan de venta')
@require_POST
def sales_plan_create_for_project(request, project):
    obj_project = Projects.objects.get(name=project)
    name = request.POST.get('name', '').strip()
    initial_payment = request.POST.get('initial_payment')
    to_finance = request.POST.get('to_finance')
    rate = request.POST.get('rate')
    if name and initial_payment and to_finance and rate:
        existing = Sales_plans.objects.filter(name__iexact=name)
        if hasattr(Sales_plans, 'project'):
            existing = existing.filter(project=obj_project)
        if existing.exists():
            messages.error(request, 'Ya existe un plan de pago con ese nombre.')
            return redirect('sales_plans_project_list', project=project)

        plan_kwargs = {
            'name': name,
            'initial_payment': initial_payment,
            'to_finance': to_finance,
            'rate': rate,
            'status': True,
        }
        if hasattr(Sales_plans, 'project'):
            plan_kwargs['project'] = obj_project
        Sales_plans.objects.create(**plan_kwargs)
    return redirect('sales_plans_project_list', project=project)

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@login_required
@require_POST
def generar_plantilla_excel(request):
    """Genera una plantilla Excel para modificar el plan de pagos"""
    try:
        # Importación adicional para resolver el problema
        from openpyxl.utils import get_column_letter
        
        # Intentar leer datos del cuerpo JSON o del POST form
        if request.content_type and 'application/json' in request.content_type:
            data = json.loads(request.body)
        else:
            # Si viene de un formulario normal
            json_data = request.POST.get('data', '{}')
            data = json.loads(json_data)
            
        # Convertir valores numéricos reemplazando comas por puntos
        saldo_total = float(str(data.get('saldo_total', 0)).replace(',', '.'))
        fecha_base = data.get('fecha_base', '')
        tasa = float(str(data.get('tasa', 0)).replace(',', '.'))  # Reemplazar coma por punto
        
        # Crear un libro de trabajo
        wb = Workbook()
        
        # ----- 1. HOJA DE INSTRUCCIONES -----
        ws_instrucciones = wb.create_sheet(title="Instrucciones", index=0)
        
        # Título principal
        ws_instrucciones.merge_cells('A1:H1')
        titulo = ws_instrucciones['A1']
        titulo.value = "GUÍA PARA CREAR TU PLAN DE PAGOS"
        titulo.font = Font(bold=True, size=16)
        titulo.alignment = Alignment(horizontal='center')
        titulo.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        titulo.font = Font(bold=True, size=16, color="FFFFFF")
        
        ws_instrucciones.row_dimensions[6].height = 90
        
        # Sección 1: Qué es un Plan de Amortización
        ws_instrucciones.merge_cells('A3:H3')
        ws_instrucciones['A3'].value = "¿Qué es un Plan de Amortización?"
        ws_instrucciones['A3'].font = Font(bold=True, size=12)
        ws_instrucciones['A3'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        ws_instrucciones.merge_cells('A4:H6')
        ws_instrucciones['A4'].value = (
            "Un plan de amortización es un calendario que muestra cómo se pagará una deuda a lo largo del tiempo. "
            "Te muestra cuánto pagarás en cada fecha y cómo se reduce el saldo pendiente con cada pago.\n\n"
            "Conceptos básicos:\n"
            "• Capital: Es el dinero que debe devolverse.\n"
            "• Interés: Es lo que se paga por usar el dinero prestado.\n"
            "• Cuota: Es el pago que se realiza (incluye capital + interés).\n"
            "• Saldo: Es lo que aún se debe después de cada pago."
        )
        ws_instrucciones['A4'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Sección 2: Cómo elaborar la tabla
        ws_instrucciones.merge_cells('A8:H8')
        ws_instrucciones['A8'].value = "Cómo elaborar tu tabla en Excel"
        ws_instrucciones['A8'].font = Font(bold=True, size=12)
        ws_instrucciones['A8'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        ws_instrucciones.merge_cells('A9:H15')
        ws_instrucciones['A9'].value = (
            "La plantilla ya contiene la estructura correcta. Sigue estos pasos:\n\n"
            "1. No modifiques la primera fila (encabezados).\n"
            "2. La suma de todas las celdas de 'Capital' debe ser igual al saldo pendiente total.\n"
            "3. Completa las columnas principales:\n"
            "   - Tipo Cuota: Puede ser SCR (Cuota Regular), SCE (Cuota Extraordinaria) o CI (Cuota Inicial).\n"
            "   - Fecha: La fecha de pago en formato DD/MM/AAAA.\n"
            "   - Capital: La parte del pago que reduce tu deuda.\n"
            "   - Interés: Lo que pagas por el préstamo.\n"
            "   - Otros: Otros cargos si aplican.\n"
            "4. El Total debe ser igual a la suma de Capital + Interés + Otros.\n"
            "5. La primera fila debe tener el Saldo Restante igual al saldo pendiente total."
        )
        ws_instrucciones['A9'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Sección 3: Validaciones
        ws_instrucciones.merge_cells('A17:H17')
        ws_instrucciones['A17'].value = "Validaciones importantes"
        ws_instrucciones['A17'].font = Font(bold=True, size=12)
        ws_instrucciones['A17'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        ws_instrucciones.merge_cells('A18:H22')
        ws_instrucciones['A18'].value = (
            "Cuando subas tu Excel, el sistema verificará que:\n\n"
            "1. La suma de todos los capitales debe ser igual al saldo pendiente total.\n"
            "2. Los tipos de cuota deben ser únicamente SCR, SCE o CI.\n"
            "3. No puede haber valores negativos o cero en Capital o en el Total.\n"
            "4. En cada cuota, la suma de Capital + Interés + Otros debe ser igual al Total declarado."
        )
        ws_instrucciones['A18'].alignment = Alignment(wrap_text=True, vertical='top')
        
                # Sección 4: Aclaración sobre capital pendiente
        ws_instrucciones.merge_cells('A24:H24')
        ws_instrucciones['A24'].value = "ACLARACIÓN IMPORTANTE"
        ws_instrucciones['A24'].font = Font(bold=True, size=12)
        ws_instrucciones['A24'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        ws_instrucciones.merge_cells('A25:H27')
        ws_instrucciones['A25'].value = (
            "Este plan de pagos se calcula ÚNICAMENTE sobre el CAPITAL PENDIENTE, no sobre el valor total de la venta.\n\n"
            f"Capital pendiente actual: ${saldo_total:,.2f}\n"
            "Este valor corresponde al saldo pendiente de pago en este momento, y es sobre este monto que debe elaborarse el nuevo plan de pagos."
        )
        ws_instrucciones['A25'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_instrucciones['A25'].font = Font(bold=True)
        
        # Sección 5: Ejemplo práctico
        ws_instrucciones.merge_cells('A29:H29')
        ws_instrucciones['A29'].value = "¿Cómo se calcula una cuota?"
        ws_instrucciones['A29'].font = Font(bold=True, size=12)
        ws_instrucciones['A29'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        ws_instrucciones.merge_cells('A30:H37')
        ws_instrucciones['A30'].value = (
            f"Para entender mejor, imagina un préstamo de $1,000,000 a 12 meses con una tasa del {tasa}% mensual:\n\n"
            "Mes 1:\n"
            "• Saldo inicial: $1,000,000\n"
            f"• Interés: $1,000,000 × {tasa/100} = ${1000000 * (tasa/100):,.0f}\n"
            "• Valor de la cuota: $88,849 (calculada con fórmula de amortización)\n"
            f"• Capital: $88,849 - ${1000000 * (tasa/100):,.0f} = ${88849 - 1000000 * (tasa/100):,.0f}\n"
            f"• Saldo después del pago: $1,000,000 - ${88849 - 1000000 * (tasa/100):,.0f} = ${1000000 - (88849 - 1000000 * (tasa/100)):,.0f}\n\n"
            "Y así sucesivamente hasta terminar de pagar."
        )
        ws_instrucciones['A30'].alignment = Alignment(wrap_text=True, vertical='top')
        
        
        # Ajustar anchos de columna para la hoja de instrucciones
        for col in range(1, 9):
            letra_col = get_column_letter(col)
            ws_instrucciones.column_dimensions[letra_col].width = 15

        

        # ----- 2. HOJA DE INFORMACIÓN -----
        ws_info = wb.create_sheet(title="Información")
        
        # Encabezados y estilos
        headers_info = ["Venta", "Contrato", "Proyecto", "Valor Total", "Saldo Pendiente", "Tasa MV"]
        for col_num, header in enumerate(headers_info, 1):
            cell = ws_info.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Datos
        values_info = [
            data.get('venta', ''),
            data.get('contrato', ''),
            data.get('proyecto', ''),
            data.get('valor_total', ''),
            data.get('saldo_pendiente', ''),
            f"{tasa}%"
        ]
        
        for col_num, value in enumerate(values_info, 1):
            ws_info.cell(row=2, column=col_num).value = value
        
        # Ajustar anchos de columna para la hoja de información
        for col_num, header in enumerate(headers_info, 1):
            letra_col = get_column_letter(col_num)
            ws_info.column_dimensions[letra_col].width = max(15, len(header) + 2)
        
        # ----- 3. HOJA DEL PLAN DE PAGOS -----
        ws_plan = wb.create_sheet(title="Plan de Pagos")
        
        # Encabezados
        headers_plan = ["Tipo Cuota", "Fecha", "Capital", "Interés", "Otros", "Total", "Saldo Restante"]
        for col_num, header in enumerate(headers_plan, 1):
            cell = ws_plan.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Agregar 24 filas en blanco
        for i in range(24):
            row_num = i + 2
            for col_num, header in enumerate(headers_plan, 1):
                cell = ws_plan.cell(row=row_num, column=col_num)
                # Para la primera fila, agregar el saldo total en la columna de saldo restante
                if i == 0 and header == "Saldo Restante":
                    cell.value = saldo_total
                else:
                    cell.value = ""
                
                # Para Tipo Cuota, establecer SCR por defecto
                if i == 0 and header == "Tipo Cuota":
                    cell.value = "SCR"
        
        # Establecer el formato de la fecha
        for i in range(24):
            cell = ws_plan.cell(row=i+2, column=2)  # Columna de fecha
            cell.number_format = "DD/MM/YYYY"
        
        # Ajustar anchos de columna para la hoja del plan de pagos
        for col_num, header in enumerate(headers_plan, 1):
            letra_col = get_column_letter(col_num)
            ws_plan.column_dimensions[letra_col].width = max(15, len(header) + 2)
        
        # Establecer la hoja de Instrucciones como activa al abrir
        wb.active = 0
        
        # Guardar el archivo Excel en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Configurar la respuesta con el tipo correcto
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Plantilla_Plan_Pagos_{data.get("contrato", "")}.xlsx'
        
        # Guardar el Excel directamente en la respuesta
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        print(f"Error al generar plantilla Excel: {str(e)}")
        print(traceback.format_exc())
        # Devuelve una respuesta de error que el navegador pueda mostrar
        return HttpResponse(
            f"Error al generar la plantilla Excel: {str(e)}",
            content_type='text/plain',
            status=500
        )

@login_required
@require_POST
def procesar_excel_plan_pagos(request):
    """Procesa un archivo Excel con el plan de pagos modificado"""
    try:
        if 'excel_file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No se ha subido ningún archivo'})
        
        excel_file = request.FILES['excel_file']
        sale_id = request.POST.get('sale_id')
        
        # Obtener la venta para validaciones
        try:
            sale = Sales.objects.get(pk=sale_id)
            # Verificar si tiene plan de pagos existente
            tiene_plan_existente = Payment_plans.objects.filter(sale=sale).exists()
        except Sales.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Venta no encontrada'})
        
        # Leer el archivo Excel
        from openpyxl import load_workbook
        wb = load_workbook(filename=BytesIO(excel_file.read()))
        
        # Verificar que exista la hoja "Plan de Pagos"
        if "Plan de Pagos" not in wb.sheetnames:
            return JsonResponse({
                'success': False, 
                'error': 'No se encontró la hoja "Plan de Pagos" en el archivo'
            })
        
        ws = wb["Plan de Pagos"]
        
        # Obtener encabezados
        headers = [cell.value for cell in ws[1]]
        
        # Verificar columnas requeridas
        columnas_requeridas = ['Tipo Cuota', 'Fecha', 'Capital', 'Interés', 'Otros']
        for col in columnas_requeridas:
            if col not in headers:
                return JsonResponse({
                    'success': False, 
                    'error': f'Columna requerida "{col}" no encontrada en el Excel'
                })
        
        # Mapear índices de columnas
        col_indices = {}
        for idx, header in enumerate(headers):
            col_indices[header] = idx
        
        # Procesar las filas válidas
        cuotas = []
        saldo_anterior = None
        total_capital = Decimal('0.00')
        
        for row in ws.iter_rows(min_row=2):
            # Obtener valores de las celdas
            tipo_cuota = row[col_indices['Tipo Cuota']].value
            fecha = row[col_indices['Fecha']].value
            capital = row[col_indices['Capital']].value
            interes = row[col_indices['Interés']].value
            otros = row[col_indices['Otros']].value
            saldo_restante = row[col_indices['Saldo Restante']].value if 'Saldo Restante' in col_indices else None
            
            # Saltear filas sin tipo de cuota o fecha
            if not tipo_cuota or not fecha:
                continue
            
            # Procesar fecha
            if isinstance(fecha, datetime.datetime):
                fecha_str = fecha.strftime('%Y-%m-%d')
            elif isinstance(fecha, str):
                # Intentar parsear la fecha si es string
                try:
                    # Asumir formato dd/mm/yyyy
                    if '/' in fecha:
                        parts = fecha.split('/')
                        if len(parts) == 3:
                            # Convertir dd/mm/yyyy a yyyy-mm-dd
                            fecha_str = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                        else:
                            continue
                    # O formato yyyy-mm-dd
                    elif '-' in fecha:
                        fecha_str = fecha
                    else:
                        continue  # Formato desconocido
                except:
                    continue  # Error al parsear fecha
            else:
                continue  # Tipo de fecha desconocido
            
            # Si es la primera fila con datos y no hay saldo anterior, usar valor de venta
            if saldo_anterior is None:
                if saldo_restante is not None:
                    try:
                        saldo_anterior = Decimal(str(saldo_restante))
                    except (ValueError, TypeError):
                        saldo_anterior = Decimal(str(sale.value))  # Override: usar valor de venta
                else:
                    saldo_anterior = Decimal(str(sale.value))  # Override: usar valor de venta
            
            # Convertir valores numéricos
            try:
                capital_val = Decimal(str(capital)) if capital is not None else Decimal('0.00')
                interes_val = Decimal(str(interes)) if interes is not None else Decimal('0.00')
                otros_val = Decimal(str(otros)) if otros is not None else Decimal('0.00')
            except (ValueError, TypeError):
                continue  # Valores no numéricos
            
            # Validar que el capital no sea cero o negativo
            if capital_val <= 0:
                return JsonResponse({
                    'success': False, 
                    'error': f'El capital de la cuota {tipo_cuota} no puede ser cero o negativo'
                })
            
            # Acumular total de capital
            total_capital += capital_val
            
            # Añadir cuota válida
            cuotas.append({
                'id_quota': f"{tipo_cuota}1CTR{sale.contract_number}",  # Se ajustará después
                'tipo': tipo_cuota,
                'fecha': fecha_str,
                'capital': float(capital_val),
                'interes': float(interes_val),
                'otros': float(otros_val),
                'saldo_anterior': float(saldo_anterior)
            })
            
            # Actualizar saldo para la siguiente cuota
            saldo_anterior -= capital_val
        
        if not cuotas:
            return JsonResponse({'success': False, 'error': 'No se encontraron cuotas válidas en el Excel'})
        
        # OVERRIDE PARA VENTAS SIN PLAN: Permitir diferencia si no tiene plan existente
        valor_esperado = Decimal(str(sale.value))
        if not tiene_plan_existente:
            # Para ventas sin plan, ser más flexible con la validación
            diferencia = abs(total_capital - valor_esperado)
            tolerancia = valor_esperado * Decimal('0.05')  # 5% de tolerancia
            
            if diferencia > tolerancia:
                return JsonResponse({
                    'success': False, 
                    'error': f'La suma del capital ({total_capital:,.2f}) debe aproximarse al valor de la venta ({valor_esperado:,.2f}). Diferencia: {diferencia:,.2f}'
                })
        else:
            # Para ventas con plan existente, validar exactitud
            if total_capital != valor_esperado:
                return JsonResponse({
                    'success': False, 
                    'error': f'La suma del capital ({total_capital:,.2f}) debe ser igual al saldo pendiente ({valor_esperado:,.2f})'
                })
        
        # Generar IDs únicos para las cuotas
        contadores_tipo = {}
        for cuota in cuotas:
            tipo = cuota['tipo']
            if tipo not in contadores_tipo:
                contadores_tipo[tipo] = 1
            cuota['id_quota'] = f"{tipo}{contadores_tipo[tipo]}CTR{sale.contract_number}"
            contadores_tipo[tipo] += 1
        
        # Registrar en historial si es override
        if not tiene_plan_existente:
            Sales_history.objects.create(
                sale=sale,
                action=f'Cargó plan de pagos desde Excel (override - sin plan previo) - {len(cuotas)} cuotas',
                user=request.user
            )
        
        return JsonResponse({
            'success': True, 
            'data': cuotas,
            'override_sin_plan': not tiene_plan_existente,
            'mensaje': 'Plan cargado correctamente' + (' (aplicado override para venta sin plan)' if not tiene_plan_existente else '')
        })
        
    except Exception as e:
        import traceback
        print(f"Error al procesar Excel: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)})
    
@login_required
@require_POST
@user_permission('gestionar_recaudos')
def desvincular_recaudos(request, project, sale_id):
    try:
        motivo = request.POST.get('motivo', '').strip()
        
        if not motivo:
            return JsonResponse({
                'success': False, 
                'message': 'El motivo es obligatorio para desvincular recaudos'
            })
        
        sale = get_object_or_404(Sales, pk=sale_id, project__name=project)
        
        with transaction.atomic():
            # CORRECCIÓN: Usar el método que funciona según el debug
            income_details = Incomes_detail.objects.filter(quota__sale=sale)
            
            if not income_details.exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'No hay recaudos aplicados para desvincular en esta venta'
                })
            
            # Contar aplicaciones antes de eliminar
            total_aplicaciones = income_details.count()
            total_valor = income_details.aggregate(
                total=Sum(
                    F('capital') + F('interest') + F('others') + F('arrears'),
                    output_field=ModelDecimalField(max_digits=20, decimal_places=2)
                )
            ).get('total') or 0
            
            # Crear backup de las aplicaciones actuales
            backup_details = []
            for detail in income_details:
                backup_details.append({
                    'income_id': detail.income_id,
                    'quota_id': detail.quota_id,
                    'capital': float(detail.capital) if detail.capital else 0,
                    'interest': float(detail.interest) if detail.interest else 0,
                    'others': float(detail.others) if detail.others else 0,
                    'arrears': float(detail.arrears) if detail.arrears else 0,
                    'arrears_days': detail.arrears_days
                })
            
            # Guardar el backup
            backup = IncomeDetailsBackup.objects.create(
                sale=sale,
                user=request.user,
                details=json.dumps(backup_details),
                motivo=motivo
            )
            
            # Eliminar todas las aplicaciones
            income_details.delete()
            
            # Registrar en historial
            Sales_history.objects.create(
                sale=sale,
                action=f"Desvinculó las aplicaciones de todos los recaudos. Motivo: {motivo}",
                user=request.user
            )
            
            # Registrar en timeline del proyecto
            Timeline.objects.create(
                user=request.user,
                action=f'Desvinculó recaudos de CTR{sale.contract_number}',
                project=sale.project,
                aplication='sales'
            )
            
        return JsonResponse({
            'success': True,
            'message': f'Se desvincularon {total_aplicaciones} aplicaciones por un total de ${total_valor:,.2f}',
            'backup_id': backup.id,
            'total_aplicaciones': total_aplicaciones,
            'total_valor': float(total_valor)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Error al desvincular recaudos: {str(e)}'
        })


@login_required
@require_POST
@user_permission('gestionar_recaudos')
def desaplicar_recibos(request, project, sale_id):
    """
    Permite desaplicar los últimos recibos de forma incremental (último en entrar, primero en salir).
    Para desaplicar un recibo intermedio, se desaplican también todos los más recientes.
    """
    motivo = request.POST.get('motivo', '').strip()
    income_id = request.POST.get('income_id')

    if not motivo:
        return JsonResponse({
            'success': False,
            'message': 'Debes ingresar el motivo de la desaplicación'
        }, status=400)

    if not income_id:
        return JsonResponse({
            'success': False,
            'message': 'Debes seleccionar un recibo para desaplicar'
        }, status=400)

    sale = get_object_or_404(Sales, pk=sale_id, project__name=project)

    with transaction.atomic():
        incomes_desc = list(
            Incomes.objects.filter(sale=sale)
            .order_by('-add_date', '-receipt', '-pk')
        )

        selected_index = next(
            (idx for idx, income in enumerate(incomes_desc) if str(income.pk) == str(income_id)),
            None
        )

        if selected_index is None:
            return JsonResponse({
                'success': False,
                'message': 'El recibo seleccionado no pertenece a la venta'
            }, status=400)

        incomes_to_unapply = incomes_desc[:selected_index + 1]
        income_ids = [income.pk for income in incomes_to_unapply]

        income_details = Incomes_detail.objects.filter(income__in=income_ids)

        if not income_details.exists():
            return JsonResponse({
                'success': False,
                'message': 'Los recibos seleccionados no tienen aplicaciones activas'
            }, status=400)

        total_aplicaciones = income_details.count()
        total_valor = income_details.aggregate(
            total=Sum(F('capital') + F('interest') + F('others') + F('arrears'))
        ).get('total') or Decimal('0')

        receipts_list = [str(income.receipt) for income in incomes_to_unapply]

        backup_details = []
        for detail in income_details.select_related('income', 'quota'):
            backup_details.append({
                'income_id': detail.income_id,
                'quota_id': detail.quota_id,
                'capital': float(detail.capital) if detail.capital else 0,
                'interest': float(detail.interest) if detail.interest else 0,
                'others': float(detail.others) if detail.others else 0,
                'arrears': float(detail.arrears) if detail.arrears else 0,
                'arrears_days': detail.arrears_days
            })

        receipts_join = ', '.join(receipts_list)

        IncomeDetailsBackup.objects.create(
            sale=sale,
            user=request.user,
            details=json.dumps(backup_details),
            motivo=f"Desaplicación parcial ({receipts_join}): {motivo}"
        )

        income_details.delete()

        def truncate_text(text, max_length=255):
            return text if len(text) <= max_length else f"{text[:max_length - 3]}..."

        if len(receipts_list) > 6:
            receipts_summary = ', '.join(receipts_list[:6]) + f'... (+{len(receipts_list) - 6} recibos)'
        else:
            receipts_summary = receipts_join

        Sales_history.objects.create(
            sale=sale,
            action=truncate_text(f"Desaplicó los recibos {receipts_summary}. Motivo: {motivo}"),
            user=request.user
        )

        Timeline.objects.create(
            user=request.user,
            action=truncate_text(f'Desaplicó recibos {receipts_summary} de CTR{sale.contract_number}'),
            project=sale.project,
            aplication='sales'
        )

    return JsonResponse({
        'success': True,
        'message': (
            f"Se desaplicaron {len(receipts_list)} recibo(s): {', '.join(receipts_list)}. "
            f"Total aplicaciones afectadas: {total_aplicaciones}"
        ),
        'recibos': receipts_list,
        'total_aplicaciones': total_aplicaciones,
        'total_valor': float(total_valor)
    })


@login_required
@require_POST
@user_permission('gestionar_recaudos')
def eliminar_detalles_orfanos(request, project, sale_id):
    motivo = request.POST.get('motivo', '').strip()
    sale = get_object_or_404(Sales, pk=sale_id, project__name=project)

    orphans_qs = (
        Incomes_detail.objects
        .filter(quota__sale=sale)
        .exclude(income__sale=sale)
        .select_related('income', 'quota')
    )

    if not orphans_qs.exists():
        return JsonResponse({
            'success': False,
            'message': 'No hay detalles huérfanos para eliminar'
        }, status=400)

    total_valor = Decimal('0.00')
    backup_details = []

    for detail in orphans_qs:
        total_valor += detail.total_income()
        backup_details.append({
            'income_id': detail.income_id,
            'quota_id': detail.quota_id,
            'capital': float(detail.capital) if detail.capital else 0,
            'interest': float(detail.interest) if detail.interest else 0,
            'others': float(detail.others) if detail.others else 0,
            'arrears': float(detail.arrears) if detail.arrears else 0,
            'arrears_days': detail.arrears_days,
            'orphan_from_sale': detail.income.sale.contract_number if detail.income_id else None,
        })

    IncomeDetailsBackup.objects.create(
        sale=sale,
        user=request.user,
        details=json.dumps(backup_details),
        motivo=f"Eliminó {len(backup_details)} aplicaciones huérfanas. Motivo: {motivo or 'No indicado'}"
    )

    deleted_count = orphans_qs.count()
    orphans_qs.delete()

    Sales_history.objects.create(
        sale=sale,
        action=f"Eliminó {deleted_count} aplicaciones huérfanas. Motivo: {motivo or 'No indicado'}",
        user=request.user
    )

    Timeline.objects.create(
        user=request.user,
        action=f'Eliminó {deleted_count} aplicaciones huérfanas de CTR{sale.contract_number}',
        project=sale.project,
        aplication='sales'
    )

    return JsonResponse({
        'success': True,
        'message': f'Se eliminaron {deleted_count} aplicaciones huérfanas por ${total_valor:,.2f}',
        'total_valor': float(total_valor),
        'deleted_count': deleted_count
    })

@login_required
@project_permission
def review_properties_status(request, project):
    """
    Revisa inconsistencias entre el estado de los inmuebles y las ventas activas.
    Retorna un preview de cambios propuestos para aprobación/rechazo.
    """
    obj_project = Projects.objects.get(name=project)
    
    if request.method == 'POST' and request.is_ajax():
        action = request.POST.get('action')
        
        if action == 'preview':
            # Revisar inconsistencias y generar preview
            changes = []
            
            # Obtener todos los inmuebles del proyecto
            properties = Properties.objects.filter(project=project)
            
            for prop in properties:
                # Buscar ventas activas (no desistidas) asociadas a este inmueble
                active_sales = Sales.objects.filter(
                    property_sold=prop,
                    project=obj_project
                ).exclude(
                    status__in=['Desistido', 'Anulado']
                ).exists()
                
                change_needed = False
                new_state = prop.state
                reason = ""
                sale_contract = ""
                
                # Verificar inconsistencias
                if prop.state == 'Asignado' and not active_sales:
                    # Inmueble marcado como asignado pero sin venta activa
                    change_needed = True
                    new_state = 'Libre'
                    reason = "Inmueble asignado sin venta activa"
                    
                elif prop.state in ['Libre', 'Bloqueado'] and active_sales:
                    # Inmueble libre/bloqueado pero con venta activa
                    change_needed = True
                    new_state = 'Asignado'
                    reason = "Inmueble libre/bloqueado con venta activa"
                    # Obtener número de contrato
                    sale = Sales.objects.filter(
                        property_sold=prop,
                        project=obj_project
                    ).exclude(
                        status__in=['Desistido', 'Anulado']
                    ).first()
                    if sale:
                        sale_contract = f"CTR-{sale.contract_number}"
                
                if change_needed:
                    changes.append({
                        'property_id': prop.pk,
                        'description': prop.description,
                        'block': prop.block,
                        'stage': prop.stage if hasattr(prop, 'stage') else '',
                        'current_state': prop.state,
                        'new_state': new_state,
                        'reason': reason,
                        'sale_contract': sale_contract
                    })
            
            return JsonResponse({
                'status': 'success',
                'changes': changes,
                'total': len(changes)
            })
        
        elif action == 'confirm':
            # Aplicar los cambios confirmados
            changes_data = json.loads(request.POST.get('changes', '[]'))
            applied_changes = []
            
            from django.db import transaction
            with transaction.atomic():
                for change in changes_data:
                    try:
                        prop = Properties.objects.select_for_update().get(pk=change['property_id'])
                        old_state = prop.state
                        prop.state = change['new_state']
                        prop.save()
                        
                        # Registrar en timeline
                        Timeline.objects.create(
                            user=request.user,
                            action=f"Revisión automática: {prop.description} cambió de {old_state} a {change['new_state']}",
                            project=obj_project,
                            aplication='sales'
                        )
                        
                        applied_changes.append({
                            'property': prop.description,
                            'old_state': old_state,
                            'new_state': change['new_state']
                        })
                    except Exception as e:
                        return JsonResponse({
                            'status': 'error',
                            'message': f'Error al actualizar {change["description"]}: {str(e)}'
                        })
            
            return JsonResponse({
                'status': 'success',
                'message': f'Se actualizaron {len(applied_changes)} inmuebles correctamente',
                'applied_changes': applied_changes
            })
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'})
        
urlpattern = [
    path('projectselection', project_selection),
    path('<project>/new_sale', new_sale),
    path('<project>/nonapprovedsales', non_approved_sales),
    path('<project>/toadjudicatesales', to_adjudicate_sales),
    path('<project>/adjudicatesales', adjudicate_sales),
    path('<project>/properties', properties_for_sales),
    path('<project>/graphs',graphs),
    path('<str:project>/files/<int:contract_number>/get/', get_sales_files, name='get_sales_files'),
    path('<str:project>/files/<int:contract_number>/add/', add_sales_file, name='add_sales_file'),
    path('<str:project>/files/<int:file_id>/delete/', delete_sales_file, name='delete_sales_file'),
    path('venta/<int:sale_id>/plan-pagos/', plan_pagos_detalle, name='plan_pagos_detalle'),
    path('cuota/<int:cuota_id>/editar/', editar_cuota, name='editar_cuota'),
    path('guardar_reestructuracion/', GuardarReestructuracionView.as_view(), name='guardar_reestructuracion'),
    path('<project>/reestructuraciones/', listado_reestructuraciones, name='listado_reestructuraciones'),
    path('<project>/fechas-entrega/', delivery_dates_list, name='delivery_dates_list'),
    path('<project>/planes-pago/', sales_plans_project_list, name='sales_plans_project_list'),
    path('plan-pago/<int:plan_id>/toggle/', sales_plan_toggle_status, name='sales_plan_toggle_status'),
    path('<project>/plan-pago/crear/', sales_plan_create_for_project, name='sales_plan_create_for_project'),
    path('generar_plantilla_excel/', generar_plantilla_excel, name='generar_plantilla_excel'),
    path('procesar_excel_plan_pagos/', procesar_excel_plan_pagos, name='procesar_excel_plan_pagos'),
] + [
    path('ajax/salesplansinfo', ajax_get_plans_info),
    path('ajax/<project>/comissions/<sale>', ajax_comissions),
    path('ajax/<project>/printsalesdocuments', ajax_print_documents),
    path('ajax/<project>/changesaleproperty', ajax_change_property),
    path('ajax/reestructuratesale/<sale>', ajax_reestructurate_payment),
    path('ajax/change_comissions/<sale>', ajax_change_comissions),
    path('ajax/desistsale/<sale>', ajax_desist_sale),
    path('ajax/<project>/changedates/<sale>',ajax_change_dates_adj),
    path('ajax/<project>/changeowners',ajax_change_owners),
    path('<project>/pdf/plan-pagos/<int:sale_id>/', generar_plan_pagos_pdf, name='generar_plan_pagos_pdf'),
    path('<project>/pdf/listado-recibos/<int:sale_id>/', generar_recibos_pdf, name='generar_recibos_pdf'),
    path('ajax/detalle_reestructuracion/<int:id>/', ajax_detalle_reestructuracion, name='detalle_reestructuracion'),
    path('ajax/aprobar_reestructuracion/<int:id>/', aprobar_reestructuracion, name='aprobar_reestructuracion'),
    path('ajax/rechazar_reestructuracion/<int:id>/', rechazar_reestructuracion, name='rechazar_reestructuracion'),
    path('ajax/<project>/delivery-dates/<int:sale_id>', ajax_delivery_dates, name='ajax_delivery_dates'),
    path('ajax/<str:project>/desvincular-recaudos/<int:sale_id>', desvincular_recaudos, name='desvincular_recaudos'),
    path('ajax/<str:project>/desaplicar-recibos/<int:sale_id>', desaplicar_recibos, name='desaplicar_recibos'),
    path('ajax/<str:project>/eliminar-detalles-orfanos/<int:sale_id>', eliminar_detalles_orfanos, name='eliminar_detalles_orfanos'),
    path('ajax/<str:project>/review-properties-status', review_properties_status, name='review_properties_status'),
    
]

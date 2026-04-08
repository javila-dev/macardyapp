import calendar
import re
import locale
import random
import string
import openpyxl
import json
from openpyxl.styles import Font, Alignment, borders, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from django.conf import settings
from django.contrib import messages
from django.db import IntegrityError
from django.db.models.query_utils import Q
from django.db.models import F, Value, When, Sum, Max, Count, Case
from django.db.models.functions import Concat
from django.http.response import FileResponse, HttpResponseRedirect, JsonResponse
from django.views.decorators.http import require_POST
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import path, reverse
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.contrib.auth.models import User
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.template.loader import render_to_string
from django.http import HttpResponse
from io import BytesIO

from finance.forms import SolicitudReciboForm, incomes_form, new_expense_form, new_sale_income_form
from finance.models import (PMT, CommentType, Collection_budget, Collection_budget_detail, Collection_feed, Collector_per_sale, Comissions_Payment, Commercial_budget, Commercial_budget_detail, Credit_info, get_commission_method, get_min_commission_value, Incomes,
                            Incomes_detail, Incomes_return, PMT_detail, Payment_methods, Sales_extra_info, SolicitudRecibo, cost_center, expenses_detail, payment_accounts)

from mcd_site.models import Counters, Parameters, Perfil, Projects, Timeline
from mcd_site.utils import JsonRender, parse_semantic_date, pdf_gen, project_permission, user_check_perms, user_permission,numbers_names, countries_data
from sales.forms import collectionfeed_Form, ComissionPositionForm
from sales.models import Assigned_comission, Comission_position, Paid_comissions, Payment_plans, Sales, Sales_history
from terceros.models import Sellers
from django.db import transaction

# Create your views here.
locale.setlocale(locale.LC_ALL,'es_CO.UTF-8')
#messages.success(request,'<div class="header">¡Lo hicimos!</div>Aprobaste el contrato')


@login_required
def project_selection(request):

    next_url = request.GET.get('next')
    menu = request.GET.get('menu')
    if next_url is None:
        messages.error(
            request, '<div class="header">Encontramos un error</div>No existe una url asociada a esta seleccion')
        HttpResponseRedirect('/landing')

    context = {
        'projects': Projects.objects.all(),
        'next': next_url,
        'menu': menu
    }

    return render(request, 'project_selection_finance.html', context)


@project_permission
@user_permission('recaudos ventas nuevas')
def new_sales_incomes(request, project):
    """
    VISTA DEPRECADA: Ahora los recibos de ventas nuevas se manejan
    a través del sistema de solicitudes de recibos.

    Se mantiene esta vista para evitar errores 404 en enlaces antiguos,
    pero redirige al usuario al sistema nuevo.
    """
    obj_project = Projects.objects.get(name=project)

    # ⚠️ Redirigir al sistema de solicitudes con mensaje informativo
    messages.info(
        request,
        '<div class="header"><i class="info circle icon"></i>Sistema actualizado</div>'
        'Los recibos para ventas nuevas ahora se manejan a través del sistema de <strong>Solicitudes de Recibos</strong>. '
        'Este método permite mejor trazabilidad y control. '
        '<br><br>Por favor, crea una solicitud desde el botón "Nueva Solicitud".'
    )
    return redirect('lista_solicitudes_recibo', project=project)

    # ⚠️ TODO EL CÓDIGO SIGUIENTE ESTÁ DEPRECADO PERO SE MANTIENE POR COMPATIBILIDAD
    # Si se necesita re-habilitar, simplemente comentar el redirect de arriba

    obj_incomes = Incomes.objects.filter(
        Q(sale__status='Pendiente') | Q(sale__status='Aprobado'),
        project=project,
    ).order_by('-receipt', '-add_date')
    context = {
        'project': obj_project,
        'receipts': obj_incomes,
        'form': new_sale_income_form(project=project)
    }

    if request.method == 'POST':
        sale = request.POST.get('sale')
        add_date = request.POST.get('add_date')
        dt_add_date = datetime.strptime(add_date, '%B %d, %Y')
        payment_date = request.POST.get('payment_date')
        dt_payment_date = datetime.strptime(payment_date, '%B %d, %Y')
        payment_method = request.POST.get('payment_method')
        description = request.POST.get('description')

        obj_sale = Sales.objects.get(pk=sale)

        receipt = Counters.objects.get(name='recibos', project=project)
        receipt_nro = receipt.value

        
        value_1 = request.POST.get('value_1').replace(',','')
        value_2 = request.POST.get('value_2')
        
        if value_2: 
            value_2 = request.POST.get('value_2').replace(',','')
        else:
            value_2 = 0 
        
        total_income = int(value_1) + int(value_2)
        
        payment_method_1 = request.POST.get('payment_method_1')
        payment_method_2 = request.POST.get('payment_method_2')
        
            
        pm1 = Payment_methods.objects.get(pk=payment_method_1)
        pm2 = None
        if payment_method_2:
            pm2 = Payment_methods.objects.get(pk=payment_method_2)
        
        income = Incomes.objects.create(
            project=obj_project, sale=obj_sale,
            receipt=receipt_nro, add_date=dt_add_date,
            payment_date=dt_payment_date, value=total_income,
            payment_method=pm1, description=description,
            user=request.user, 
            value1 = value_1, value2 = value_2,
            pm1 = pm1, pm2 = pm2
        )
        """ income = Incomes.objects.create(
            project=obj_project, sale=obj_sale,
            add_date=dt_add_date, payment_date=dt_payment_date,
            payment_method=obj_pm, value=value,
            description=description, user=request.user,
            receipt=receipt_nro
        ) """

        receipt.value += 1
        receipt.save()
        
        filename = f'Recibocaja{receipt_nro}_CTR{obj_sale.contract_number}_{project}.pdf'.replace(
            'ñ', 'n')

        divisor = 1
        if income.sale.second_owner.full_name() != " ":
            divisor += 1
        if income.sale.third_owner.full_name() != " ":
            divisor += 1

        divided_income = f'{int(income.value)/divisor:,.2f}'

        filename = f'Recibocaja{income.receipt}_CTR{income.sale.contract_number}_{project}.pdf'.replace(
            'ñ', 'n')

        pdf_file = pdf_gen(
            'pdf/income_recepit.html',
            {'income': income,'divided_income': divided_income},
            filename,
        )
        pdf_url = pdf_file.get('url')
        href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'
        messages.success(request,
                         '<div class="header">¡Lo hicimos!</div>El recibo generado lo puedes descargar '+href)

    return render(request, 'new_sale_incomes.html', context)


@project_permission
@user_permission('recaudos ventas adjudicadas')
def adjudicated_sales_incomes(request, project):
    obj_project = Projects.objects.get(name=project)
    
    context = {
        'project': obj_project,
        'form': incomes_form(project=project)
    }
    if request.is_ajax():
        if request.method == 'GET':
            sale = request.GET.get('sale')
            if sale:
                condonate_arrears = int(request.GET.get('condonate'))
                abono_capital = request.GET.get('abono_capital') == '1'
                tipo_abono = request.GET.get('tipo_abono')
                paid_day = parse_semantic_date(request.GET.get('paid_day'), 'date')
                total_income = request.GET.get('total_income')
                no_apply_data = {
                    'sale': sale,
                    'paid_day': paid_day,
                    'total_income': int(total_income)
                }
                applicated_pay = apply_income(
                    income=None,
                    condonate_arrears=condonate_arrears,
                    apply=False,
                    no_apply_data=no_apply_data,
                    abono_capital=abono_capital,
                    tipo_abono=tipo_abono
                )
            else:
                applicated_pay = []
            
            data = {
                'data': applicated_pay
            }
            return JsonResponse(data)

    if request.method == 'POST' and not request.is_ajax():
        sale = request.POST.get('sale')
        add_date = request.POST.get('add_date')
        dt_add_date = parse_semantic_date(add_date, 'date')
        payment_date = request.POST.get('payment_day')
        dt_payment_date = parse_semantic_date(payment_date, 'date')
        
        
        description = request.POST.get('description')
        arrears_condonate = request.POST.get('arrears_condonate')

        obj_sale = Sales.objects.get(pk=sale)
        obj_counter = Counters.objects.get(project=project, name='recibos')
        receipt_number = obj_counter.value
        
        value_1 = request.POST.get('value_1').replace(',','')
        value_2 = request.POST.get('value_2')
        
        if value_2: 
            value_2 = request.POST.get('value_2').replace(',','')
        else:
            value_2 = 0 
        
        total_income = int(value_1) + int(value_2)
        
        payment_method_1 = request.POST.get('payment_method_1')
        payment_method_2 = request.POST.get('payment_method_2')
        
            
        pm1 = Payment_methods.objects.get(pk=payment_method_1)
        pm2 = None
        if payment_method_2:
            pm2 = Payment_methods.objects.get(pk=payment_method_2)
        
        income = Incomes.objects.create(
            project=obj_project, sale=obj_sale,
            receipt=receipt_number, add_date=dt_add_date,
            payment_date=dt_payment_date, value=total_income,
            payment_method=pm1, description=description,
            user=request.user, 
            value1 = value_1, value2 = value_2,
            pm1 = pm1, pm2 = pm2
        )

        obj_counter.value += 1
        obj_counter.save()
        
        tipo_abono = request.POST.get('tipo_abono_capital')
        abono_capital = request.POST.get('capital_payment') == 'on'
        print('abono_capital:', abono_capital)
        apply_income(income, condonate_arrears=int(arrears_condonate), abono_capital=abono_capital, tipo_abono=tipo_abono)

        filename = f'Recibocaja{receipt_number}_CTR{obj_sale.contract_number}_{project}.pdf'.replace(
            'ñ', 'n')

        divisor = 1
        if income.sale.second_owner.full_name() != " ":
            divisor += 1
        if income.sale.third_owner.full_name() != " ":
            divisor += 1

        divided_income = f'{int(income.value)/divisor:,.2f}'

        pdf_file = pdf_gen(
            'pdf/income_recepit.html',
            {'income': income,'divided_income': divided_income},
            filename,
        )
        pdf_url = pdf_file.get('url')
        href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'
        messages.success(request,
                         '<div class="header">¡Lo hicimos!</div>El recibo generado lo puedes descargar '+href)
    
    return redirect('/finance/{}/solicitudes/'.format(project))
    return render(request, 'incomes.html', context)

@project_permission
@user_permission('lista de recaudos')
def incomes_list(request, project):
    
    if request.is_ajax():        
        if request.method == 'GET':
            
            todo = request.GET.get('todo')
            if todo == 'get-comments':
                
                receipt = request.GET.get('receipt')
                obj_receipt = Incomes.objects.get(receipt=receipt, project=project) 
                
                data_obs = {
                    'obs_1' :   obj_receipt.obs_1,
                    'obs_2' :   obj_receipt.obs_2
                }
                
                return JsonResponse(data_obs)
            elif todo == 'datatable':
                incomes_data = []
                
                date_from =  request.GET.get('from')
                date_to = request.GET.get('to')
                date_field = request.GET.get('date_field', 'add_date')
                if date_field not in ['add_date', 'payment_date']:
                    date_field = 'add_date'
                
                if date_from and date_to and (date_from != '' and date_to != ''):
                    date_from =  parse_semantic_date(request.GET.get('from'),'date')
                    date_to = parse_semantic_date(request.GET.get('to'), 'date')
                    filter_kwargs = {f'{date_field}__range': (date_from, date_to)}
                    order_field = f'-{date_field}'

                    obj_incomes = Incomes.objects.filter(
                        project=project, **filter_kwargs
                    ).order_by(order_field)
                    
                    incomes_data = JsonRender(
                        obj_incomes,
                        query_functions=['add_date_uk', 'payment_date_uk', 'fp']
                    ).render()
                    
                    what_to_show = request.GET.get('what_to_show')
                    
                    if what_to_show == 'incomes-and-devolutions' or  what_to_show == 'just-devolutions':
                        obj_returns = Incomes_return.objects.filter(date__range=(date_from, date_to),
                                                      sale__project = project)
                        if what_to_show == 'just-devolutions':
                            incomes_data= []
                            
                        for i in obj_returns:
                            incomes_data.append({
                                'add_date_uk': datetime.strftime(i.date,'%Y/%m/%d'),
                                'payment_date_uk': datetime.strftime(i.date,'%Y/%m/%d'),
                                'receipt': f'DEV-CTR{i.sale.contract_number}',
                                'sale':{
                                    'first_owner':{
                                        'first_name':i.sale.first_owner.first_name,
                                        'last_name':i.sale.first_owner.last_name
                                    },
                                    'contract_number':i.sale.contract_number,
                                    'property_sold':{
                                        'description':i.sale.property_sold.description
                                    },
                                    'status':i.sale.status
                                },
                                'fp': 'Devoluciones',
                                'user': {
                                    'username':i.user.username
                                },
                                'value':i.value
                            })
                            
                    
                data = {
                    'data': incomes_data
                }
                
                return JsonResponse(data)
            elif todo == 'report':
                data = {}
                
                date_from =  request.GET.get('from')
                date_to = request.GET.get('to')
                what_to_show = request.GET.get('what_to_show')
                date_field = request.GET.get('date_field', 'add_date')
                if date_field not in ['add_date', 'payment_date']:
                    date_field = 'add_date'
                
                if date_from and date_to and (date_from != '' and date_to != ''):
                    date_from =  parse_semantic_date(request.GET.get('from'),'date')
                    date_to = parse_semantic_date(request.GET.get('to'), 'date')
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = 'Ingresos'
                    
                    ws.append(["Recibo", "Dia", "Mes", "Año","Contrato", "Manzana", "Lote","Cliente",
                            "Forma pago 1","Pago 1","Forma pago 2","Pago 2","Total",
                            "Concepto","Cuenta/Banco","Observación"])
                    
                    filter_kwargs = {f'{date_field}__range': (date_from, date_to)}
                    day_field = f'{date_field}__day'
                    month_field = f'{date_field}__month'
                    year_field = f'{date_field}__year'

                    obj_incomes = Incomes.objects.filter(
                            project=project, **filter_kwargs
                        ).annotate(
                            client_full_name = Concat(
                                F('sale__first_owner__first_name'), 
                                Value(' '), 
                                F('sale__first_owner__last_name')),
                            value_1 = Case(
                                When(value1__isnull = True, then=F('value')),
                                When(value1__isnull = False, then=F('value1'))
                            )
                        ).order_by(date_field, 'receipt').values_list(
                            'receipt', day_field, month_field, year_field,
                            'sale__contract_number',
                            'sale__property_sold__block','sale__property_sold__location',
                            'client_full_name','payment_method__name','value_1','pm2__name','value2','value',
                            'description','obs_1','obs_2'
                        )
                        
                    obj_incomes = list(obj_incomes)
                    
                    if what_to_show == 'incomes-and-devolutions' or  what_to_show == 'just-devolutions':
                        obj_returns = Incomes_return.objects.filter(date__range=(date_from, date_to),
                                                      sale__project = project)
                        if what_to_show == 'just-devolutions':
                            obj_incomes= []
                            
                        for r in obj_returns:
                            obj_incomes.append((
                                f'DEV-CTR{r.sale.contract_number}',r.date.day,r.date.month,r.date.year,
                                r.sale.contract_number,r.sale.property_sold.block,
                                r.sale.property_sold.location, r.sale.first_owner.full_name(),
                                'Devoluciones',r.value,"","",r.value,
                                "DESISTIMIENTO DE CONTRATO","",""                            
                            ))
                    
                    for i in obj_incomes:
                        ws.append(i)
                    
                    ws.freeze_panes = "A2"
                    ws.sheet_view.zoomScale = 75
                    
                    for row in ['J','L','M']:
                        for cell in ws[row]:
                            cell.style = 'Comma'
                            
                    for row in ['B','C','D','E','F','G']:
                        for cell in ws[row]:
                            cell.alignment = Alignment(horizontal='center')
                    
                    for row in ws['A1:P1']:
                        for cell in row:
                            cell.font = Font(bold=True)
                    
                    ws.column_dimensions['A'].width = 7.8
                    ws.column_dimensions['H'].width = 43
                    ws.column_dimensions['I'].width = 16.5
                    ws.column_dimensions['J'].width = 16.5
                    ws.column_dimensions['K'].width = 16.5
                    ws.column_dimensions['L'].width = 16.5
                    ws.column_dimensions['M'].width = 18
                    ws.column_dimensions['N'].width = 53
                    ws.column_dimensions['O'].width = 22.5
                    ws.column_dimensions['P'].width = 22.5
                    
                    data_len = len(obj_incomes)
                    filters = ws.auto_filter
                    filters.ref = f"A1:P{data_len+1}"
                    
                    ws.auto_filter.add_sort_condition(f"C2:C{data_len+1}")
                    
                        
                    filename = f'Incomes_report_{date_from.date()}_a_{date_to.date()}.xlsx'
                        
                    wb.save(settings.MEDIA_ROOT / f"tmp/{filename}")
                    
                    data = {
                        'href':f'tmp/{filename}',
                    }
                
                
                return JsonResponse(data)
                
            
        elif request.method == 'POST':
            obs_1 = request.POST.get('obs-1')
            obs_2 = request.POST.get('obs-2')
            receipt = request.POST.get('receipt')
            
            obj_receipt = Incomes.objects.get(receipt=receipt, project=project) 
            obj_receipt.obs_1 = obs_1
            obj_receipt.obs_2 = obs_2
            obj_receipt.save()
            
            return JsonResponse({'response':'ok'})
        
    
    obj_project = Projects.objects.get(name=project)
    obj_incomes = Incomes.objects.filter(
        project=project
    ).order_by('-add_date')
    context = {
        'project': obj_project,
        'receipts': obj_incomes
    }

    return render(request, 'incomes_list.html', context)


@project_permission
@user_permission('liquidar anticipo comisiones')
def liquidate_comissions_advances(request, project):
    obj_project = Projects.objects.get(name=project)
    context = {
        'project': obj_project,
    }
    if request.method == 'GET' and request.GET:
        contract_number = request.GET.get('sale') 
        sale = Sales.objects.get(
            project=project, contract_number=contract_number) 
        context.update({
            'advance_detail': Comissions_Payment.objects.filter(sale=sale.pk),
            'selected': True,
            'sale': sale,
        })

    if request.method == 'POST':
        positions = request.POST.getlist('position_pk')
        comission = request.POST.getlist('comission')
        for i in range(0, len(positions)):
            paid = Assigned_comission.objects.get(pk=positions[i])
            Paid_comissions.objects.create(
                project=obj_project, assign_paid=paid,
                pay_date=date.today(), comission=comission[i].replace(',', ''),
                provision=0, net_payment=comission[i].replace(',', ''),
                type_of_payment='Anticipo', user=request.user
            )

        if len(positions) > 0:
            Sales_history.objects.create(
                user=request.user,
                action='Liquidó anticipo de comisiones',
                sale=paid.sale,
            )

        messages.success(
            request, '<div class="header">¡Lo hicimos!</div>Liquidaste el anticipo de comisiones.')

    obj_comissions = Sales_extra_info.objects.filter(
        project=project, status='Adjudicado'
    ).order_by('add_date')
    

    context.update({
        'advances': obj_comissions,
    })

    return render(request, 'liquidate_advances.html', context)


@project_permission
@user_permission('liquidar comisiones')
def liquidate_comissions(request, project):
    obj_project = Projects.objects.get(name=project)

    obj_comissions = Sales_extra_info.objects.filter(
        project=project, status='Adjudicado'
    )

    context = {
        'project': obj_project,
        'comissions': obj_comissions,
        'scale_data': [],
        'has_payable_scale': False,
    }

    if request.method == 'GET' and request.GET:
        contract_number = request.GET.get('contract_number')
        obj_sale = Sales.objects.get(
            project=project,
            contract_number=contract_number,
        )
        scale = Comissions_Payment.objects.filter(
            sale=obj_sale.pk
        )
        sale_extra = Sales_extra_info.objects.get(pk=obj_sale.pk)
        min_value = get_min_commission_value(obj_sale.project)
        method, advance_percent = get_commission_method(obj_sale.project)
        meets_minimum = (min_value <= 0) or (Decimal(str(sale_extra.total_payment())) >= min_value)
        scale_with_values = []
        has_payable = False
        for item in scale:
            calc = item.liquidate_comission()
            if calc.get('net_pay', 0) > 0:
                has_payable = True
            scale_with_values.append((item, calc))
        obj_payment = Paid_comissions.objects.filter(
            assign_paid__sale=obj_sale.pk)
        total_payment = obj_payment.aggregate(total=Sum('comission')
                                              ).get('total')
        if total_payment == None:
            total_payment = 0
        last_payment = obj_payment.aggregate(last=Max('pay_date')).get('last')
        if last_payment == None:
            last_payment = 'No hay pagos registrados'

        total_recaudo = sale_extra.total_payment()
        advance_ratio = Decimal('0')
        required_progress_amount = Decimal('0')
        if method == 1:
            required_progress_amount = Decimal(str(obj_sale.value)) * (Decimal(str(advance_percent)) / Decimal('100'))
            if required_progress_amount > 0:
                advance_ratio = Decimal(str(total_recaudo)) / required_progress_amount
        else:
            required_progress_amount = Decimal(str(obj_sale.value))
            if required_progress_amount > 0:
                advance_ratio = Decimal(str(total_recaudo)) / required_progress_amount
        if advance_ratio > 1:
            advance_ratio = Decimal('1')
        base_info = {
            'total_comission_value': sum(calc['comission_to_pay'] for _, calc in scale_with_values),
            'commission_base': obj_sale.comission_base,
            'method': 'Fracciones clásicas' if method == 0 else 'Mínimo de recaudo + % de avance',
            'min_value': min_value,
            'advance_percent': advance_percent,
            'meets_minimum': meets_minimum,
            'recaudo_total': total_recaudo,
            'avance_real': float(advance_ratio * 100),
            'avance_objetivo': float(required_progress_amount) if required_progress_amount else 0,
        }
        context.update({
            'selected': True,
            'sale': obj_sale,
            'scale_data': scale_with_values,
            'has_payable_scale': has_payable,
            'total_payment': total_payment,
            'last_pay': last_payment,
            'base_info': base_info,
        })

    if request.method == 'POST':
        positions = request.POST.getlist('position_pk')
        comissions = request.POST.getlist('comission')
        provisions = request.POST.getlist('provision')
        nets_pay = request.POST.getlist('net_pay')
        types_of_payment = request.POST.getlist('type')

        if not positions:
            messages.error(request, '<div class="header">Ups!</div>No seleccionaste ninguna comisión para liquidar.')
            return render(request, 'liquidate_comissions.html', context)

        for i in range(0, len(positions)):
            obj_assign_pos = Assigned_comission.objects.get(
                pk=positions[i]
            )

            comission = comissions[i].replace(',', '')
            provision = provisions[i].replace(',', '')
            net_pay = nets_pay[i].replace(',', '')
            type_of_payment = types_of_payment[i]
            Paid_comissions.objects.create(
                project=obj_project,
                assign_paid=obj_assign_pos,
                pay_date=date.today(),
                comission=comission,
                provision=provision,
                net_payment=net_pay,
                type_of_payment=type_of_payment,
                user=request.user
            )

        obj_sale = Assigned_comission.objects.get(
            pk=positions[0]
        ).sale
        payment_method = types_of_payment[0] if types_of_payment else 'Sin método'
        Sales_history.objects.create(
            sale=obj_sale,
            action='Liquidó comisiones con metodo '+payment_method,
            user=request.user
        )

        messages.success(
            request, '<div class="header">¡Lo hicimos!</div>Comision pagada')

    return render(request, 'liquidate_comissions.html', context)


@project_permission
@user_permission('ver lista comisiones')
def comissions_list(request, project):
    obj_project = Projects.objects.get(name=project)

    paid_comissions_qs = Paid_comissions.objects.filter(
        assign_paid__sale__project=project,
    )

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    start_dt = parse_date(start_date) if start_date else None
    end_dt = parse_date(end_date) if end_date else None

    if start_dt:
        paid_comissions_qs = paid_comissions_qs.filter(pay_date__gte=start_dt)
    if end_dt:
        paid_comissions_qs = paid_comissions_qs.filter(pay_date__lte=end_dt)

    paid_comissions = paid_comissions_qs.values('pay_date', 'user__username'
             ).annotate(nro_sales=Count('assign_paid__sale', distinct=True),
                        total_comission=Sum('comission'),
                        total_provission=Sum('provision'),
                        total_net=Sum('net_payment')).order_by('-pay_date', 'user__username')

    context = {
        'project': obj_project,
        'paid_comissions': paid_comissions,
        'pay_accounts': payment_accounts.objects.all(),
        'start_date': start_date or '',
        'end_date': end_date or '',

    }

    return render(request, 'comissions_list.html', context)


@project_permission
@user_permission('ver lista comisiones')
def export_paid_comissions(request, project):
    obj_project = Projects.objects.get(name=project)

    paid_comissions_qs = Paid_comissions.objects.filter(
        assign_paid__sale__project=project,
    ).select_related(
        'assign_paid__seller',
        'assign_paid__sale__first_owner',
        'assign_paid__position',
        'user'
    ).order_by('pay_date', 'assign_paid__seller__first_name', 'assign_paid__sale__contract_number')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    start_dt = parse_date(start_date) if start_date else None
    end_dt = parse_date(end_date) if end_date else None

    if start_dt:
        paid_comissions_qs = paid_comissions_qs.filter(pay_date__gte=start_dt)
    if end_dt:
        paid_comissions_qs = paid_comissions_qs.filter(pay_date__lte=end_dt)

    if not paid_comissions_qs.exists():
        messages.warning(request, 'No encontramos comisiones pagadas en el rango seleccionado.')
        return redirect(f'/finance/{project}/comissions/list')

    method_value, advance_percentage = get_commission_method(obj_project)
    method_label = 'Fracciones clásicas' if method_value == 0 else 'Mínimo de recaudo + % de avance'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Comisiones pagadas'

    headers = [
        'Fecha de pago', 'Gestor', 'Identificación', 'Contrato', 'Cliente', 'Cargo',
        'Tipo de pago', 'Método de liquidación', '% avance control', 'Valor comisión',
        'Retefuente', 'Valor neto', 'Usuario que pagó'
    ]

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)

    for row_idx, payment in enumerate(paid_comissions_qs, start=2):
        sale = payment.assign_paid.sale
        seller = payment.assign_paid.seller
        client = sale.first_owner.full_name() if sale and sale.first_owner else ''
        position_name = payment.assign_paid.position.name if payment.assign_paid.position else ''
        ws.cell(row=row_idx, column=1, value=payment.pay_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_idx, column=2, value=seller.full_name() if seller else '')
        ws.cell(row=row_idx, column=3, value=getattr(seller, 'pk', ''))
        ws.cell(row=row_idx, column=4, value=sale.contract_number if sale else '')
        ws.cell(row=row_idx, column=5, value=client)
        ws.cell(row=row_idx, column=6, value=position_name)
        ws.cell(row=row_idx, column=7, value=payment.type_of_payment)
        ws.cell(row=row_idx, column=8, value=method_label)
        ws.cell(row=row_idx, column=9, value=float(advance_percentage))
        comm_cell = ws.cell(row=row_idx, column=10, value=float(payment.comission))
        prov_cell = ws.cell(row=row_idx, column=11, value=float(payment.provision))
        net_cell = ws.cell(row=row_idx, column=12, value=float(payment.net_payment))
        ws.cell(row=row_idx, column=13, value=payment.user.get_full_name() or payment.user.username)

        for c in (comm_cell, prov_cell, net_cell):
            c.number_format = '#,##0'

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    start_label = start_dt.strftime('%Y%m%d') if start_dt else 'inicio'
    end_label = end_dt.strftime('%Y%m%d') if end_dt else 'hoy'
    filename = f'Comisiones_pagadas_{project}_{start_label}_{end_label}.xlsx'.replace('ñ', 'n')

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@project_permission
@user_permission('configurar cargos comisiones')
def comission_positions_parameters(request, project):
    obj_project = Projects.objects.get(name=project)
    form = ComissionPositionForm(project=obj_project)
    sellers_qs = Sellers.objects.filter(seller_state='Activo').order_by('first_name', 'last_name')
    form.fields['default'].queryset = sellers_qs
    group_choices = Comission_position._meta.get_field('group').choices
    min_param, _ = Parameters.objects.get_or_create(
        name='valor_minimo_comision',
        project=obj_project,
        defaults={'section': 'comisiones', 'state': True, 'value': 0}
    )
    method_param, _ = Parameters.objects.get_or_create(
        name='metodo_liquidacion_comision',
        project=obj_project,
        defaults={'section': 'comisiones', 'state': True, 'value': 0}
    )
    advance_param, _ = Parameters.objects.get_or_create(
        name='porcentaje_avance_comision',
        project=obj_project,
        defaults={'section': 'comisiones', 'state': True, 'value': 30}
    )
    return render(request, 'comission_parameters.html', {
        'project': obj_project,
        'form': form,
        'sellers': sellers_qs,
        'group_choices': group_choices,
        'min_commission_value': min_param.value or 0,
        'method_value': int(method_param.value or 0),
        'advance_percentage': advance_param.value or 30,
    })


@project_permission
@user_permission('configurar cargos comisiones')
def ajax_comission_positions(request, project):
    obj_project = Projects.objects.get(name=project)
    if request.method != 'GET':
        return JsonResponse({'ok': False, 'msg': 'Método no permitido'}, status=405)
    positions = Comission_position.objects.filter(project=obj_project).order_by('group', 'name')
    data = []
    for pos in positions:
        data.append({
            'id': pos.pk,
            'name': pos.name,
            'group': pos.group,
            'group_label': pos.get_group_display(),
            'rate': float(pos.rate),
            'rate_display': f'{pos.rate:.2f}',
            'advance_bonus': pos.advance_bonus or 0,
            'default_id': pos.default_id,
            'default_name': pos.default.full_name() if pos.default else '',
            'include_default': pos.include_default,
            'is_active': pos.is_active,
        })
    return JsonResponse({'positions': data})


@require_POST
@project_permission
@user_permission('configurar cargos comisiones')
def ajax_save_comission_position(request, project):
    obj_project = Projects.objects.get(name=project)
    pk = request.POST.get('pk')
    instance = None
    if pk:
        instance = get_object_or_404(Comission_position, pk=pk, project=obj_project)
    form = ComissionPositionForm(request.POST, instance=instance, project=obj_project)
    sellers_qs = Sellers.objects.filter(seller_state='Activo').order_by('first_name', 'last_name')
    form.fields['default'].queryset = sellers_qs
    if form.is_valid():
        obj = form.save(commit=False)
        if obj.project_id and obj.project_id != obj_project.pk:
            return JsonResponse({'ok': False, 'msg': 'No puedes mover un cargo a otro proyecto'}, status=400)
        obj.project = obj_project
        obj.save()
        action = 'Creó' if instance is None else 'Actualizó'
        Timeline.objects.create(
            user=request.user,
            action=f"{action} cargo de comisión {obj.name}",
            project=obj_project,
            aplication='Comisiones'
        )
        message = 'Cargo creado' if instance is None else 'Cargo actualizado'
        return JsonResponse({'ok': True, 'msg': message})
    errors = {field: [str(e) for e in errs] for field, errs in form.errors.items()}
    return JsonResponse({'ok': False, 'errors': errors}, status=400)


@require_POST
@project_permission
@user_permission('configurar cargos comisiones')
def ajax_toggle_comission_position(request, project):
    obj_project = Projects.objects.get(name=project)
    pk = request.POST.get('pk')
    if not pk:
        return JsonResponse({'ok': False, 'msg': 'ID requerido'}, status=400)
    obj = get_object_or_404(Comission_position, pk=pk, project=obj_project)
    obj.is_active = not obj.is_active
    obj.save(update_fields=['is_active'])
    action = 'Activó' if obj.is_active else 'Inactivó'
    Timeline.objects.create(
        user=request.user,
        action=f"{action} cargo de comisión {obj.name}",
        project=obj_project,
        aplication='Comisiones'
    )
    return JsonResponse({'ok': True, 'is_active': obj.is_active})


@require_POST
@project_permission
@user_permission('configurar cargos comisiones')
def ajax_set_min_comission_value(request, project):
    obj_project = Projects.objects.get(name=project)
    valor = request.POST.get('valor')
    try:
        valor_decimal = Decimal(valor)
        if valor_decimal < 0:
            raise ValueError('El valor debe ser mayor o igual a cero')
    except Exception:
        return JsonResponse({'ok': False, 'msg': 'Valor inválido'}, status=400)
    param, _ = Parameters.objects.get_or_create(
        name='valor_minimo_comision',
        project=obj_project,
        defaults={'section': 'comisiones', 'state': True, 'value': 0}
    )
    param.value = float(valor_decimal)
    param.save()
    Timeline.objects.create(
        user=request.user,
        action=f'Actualizó el valor mínimo de comisión a ${valor_decimal:,.0f}',
        project=obj_project,
        aplication='Comisiones'
    )
    return JsonResponse({'ok': True, 'value': param.value})


@require_POST
@project_permission
@user_permission('configurar cargos comisiones')
def ajax_set_liquidation_method(request, project):
    obj_project = Projects.objects.get(name=project)
    method = request.POST.get('method')
    advance = request.POST.get('advance')
    min_value = request.POST.get('min_value')
    try:
        method_val = int(method)
        if method_val not in (0, 1):
            raise ValueError
        advance_val = float(advance)
        if advance_val <= 0:
            raise ValueError
        min_val_decimal = Decimal(normalize_currency(min_value)) if method_val == 1 else Decimal('0')
        if method_val == 1 and min_val_decimal <= 0:
            raise ValueError
    except Exception:
        return JsonResponse({'ok': False, 'msg': 'Valores inválidos'}, status=400)

    Parameters.objects.update_or_create(
        name='metodo_liquidacion_comision',
        project=obj_project,
        defaults={'section': 'comisiones', 'state': True, 'value': method_val}
    )
    Parameters.objects.update_or_create(
        name='porcentaje_avance_comision',
        project=obj_project,
        defaults={'section': 'comisiones', 'state': True, 'value': advance_val}
    )
    Parameters.objects.update_or_create(
        name='valor_minimo_comision',
        project=obj_project,
        defaults={'section': 'comisiones', 'state': True, 'value': float(min_val_decimal)}
    )
    Timeline.objects.create(
        user=request.user,
        action=f'Cambió método de liquidación a {"mínimo de recaudo" if method_val else "fracciones clásicas"} (avance {advance_val}%)',
        project=obj_project,
        aplication='Comisiones'
    )
    return JsonResponse({'ok': True})


def reprint_receipt(request, project):
    if request.method == 'GET' and request.GET:
        if not user_check_perms(request, 'imprimir recaudo'):
            data = {
                'type': 'error',
                'title': 'Tenemos un problema',
                'msj': 'Tu usuario no tiene privilegios suficientes para hacer esto.'
            }
            return JsonResponse(data)

        receipt_number = request.GET.get('receipt_number')
        income = Incomes.objects.get(project=project, receipt=receipt_number)

        divisor = 1
        if income.sale.second_owner.full_name() != " ":
            divisor += 1
        if income.sale.third_owner.full_name() != " ":
            divisor += 1

        divided_income = f'{income.value/divisor:,.2f}'

        filename = f'Recibocaja{income.receipt}_CTR{income.sale.contract_number}_{project}.pdf'.replace(
            'ñ', 'n')

        pdf_file = pdf_gen(
            'pdf/income_recepit.html',
            {'income': income, 'divided_income': divided_income},
            filename,
        )
        if request.is_ajax():
            href = pdf_file.get('url')
            a = f'<a href="{href}" target="_blank"><strong>aqui</strong></a>'
            data = {
                'type': 'success',
                'title': 'Hecho',
                'msj': 'Puedes descargar el archivo '+a
            }
            return JsonResponse(data)

        f = open(str(pdf_file.get('root')), 'rb')

        return FileResponse(f, as_attachment=True, filename=filename)


@user_permission('ver presupuesto comercial')
def commercial_budget(request,project):
    obj_project = Projects.objects.get(pk=project)
    
    context = {
        'project':obj_project
    }
    
    if request.is_ajax():
        if request.method == 'GET':
            todo = request.GET.get('todo')
            
            if todo == 'datatable':
                approve = user_check_perms(request, 'aprobar presupuesto comercial')
                supervisor = user_check_perms(request, 'auditar presupuesto comercial')
                
                obj_budget = Commercial_budget.objects.filter(project=project).order_by('-add_date')
                
                if supervisor == False and approve == False:
                    obj_budget = obj_budget.filter(user_request=request.user)
                    
                data = {
                    'data': JsonRender(obj_budget, query_functions=('totalize','date_uk')).render()
                }
                
                return JsonResponse(data)
            
            elif todo == 'view-budget':
                budget = request.GET.get('budget')
                obj_budget_detail = Commercial_budget_detail.objects.filter(budget=budget)
                
                data = {
                    'data':JsonRender(obj_budget_detail).render()
                }
                
                return JsonResponse(data)
            
            elif todo == 'print-budget':
                budget_id = request.GET.get('budget')
                budget = Commercial_budget.objects.get(pk=budget_id)
                obj_budget_detail = Commercial_budget_detail.objects.filter(budget=budget_id)
                
                filename = f'Solicitud_Presupuesto_{obj_project.name_to_show}_{budget.add_date}.pdf'.replace(
                'ñ', 'n').replace(' ','_')

                pdf_file = pdf_gen(
                    'pdf/commercial_budget.html',
                    {'budget': budget,
                     'budget_detail':obj_budget_detail,
                     'project':obj_project,
                     'user':request.user,
                     'now':datetime.now()},
                    filename,
                )

                href = pdf_file.get('url')
                a = f'<a href="{href}" target="_blank"><strong>aqui</strong></a>'
                msj = f'Ya tenemos listo tu documento, puedes descargarlo '+a

                data = {
                    'type': 'success',
                    'title': '¡Lo hicimos!',
                    'msj': msj,
                }
                return JsonResponse(data)
        
        elif request.method == 'POST':
            todo = request.POST.get('todo')
            
            if todo == 'delete-budget':
                budget_id = request.POST.get('budget')
                user_check_perms(request,'eliminar presupuesto comercial',raise_exception=True)
                
                budget = Commercial_budget.objects.get(pk=budget_id)
                date_bg = budget.add_date
                budget.delete()
                
                action = f'Eliminó un presupuesto de {obj_project.name_to_show} del {date_bg}' 
                Timeline.objects.create(
                    user = request.user, action = action, project = obj_project,
                    aplication = 'Presupuesto comercial'
                )
                return JsonResponse({'success': True})
            
            elif todo == 'approve-budget':
                budget_id = request.POST.get('budget')
                user_check_perms(request,'aprobar presupuesto comercial',raise_exception=True)
                
                budget = Commercial_budget.objects.get(pk=budget_id)
                budget.status = 'Aprobado'
                budget.user_approve = request.user
                budget.approve_date = date.today()
                budget.save()
                
                msj='Se aprobó el presupuesto seleccionado'
                
                messages.success(
                    request, '<div class="header">¡Lo hicimos!</div>'+msj)
                
                return JsonResponse({'success': True},status=200)
            
            elif todo == 'edit-budget':
                budget_id = request.POST.get('budget')
                user_check_perms(request,'crear presupuesto comercial',raise_exception=True)
                
                descriptions = request.POST.getlist('detail_description')
                values = request.POST.getlist('detail_value')
                note = request.POST.get('observations')
                
                budget = Commercial_budget.objects.get(pk=budget_id)
                budget.notes = note
                budget.save()
                
                for k in Commercial_budget_detail.objects.filter(budget = budget):
                    k.delete()
                
                for i, j in zip(descriptions, values):
                    Commercial_budget_detail.objects.create(
                        budget = budget, value = j.replace(',',''),
                        concept = i
                    )
                
                action = f'Modificó el presupuesto de {obj_project.name_to_show} del {budget.add_date}' 
                Timeline.objects.create(
                    user = request.user, action = action, project = obj_project,
                    aplication = 'Presupuesto comercial'
                )
                
                msj='Se modificó el presupuesto seleccionado'
                
                messages.success(
                    request, '<div class="header">¡Lo hicimos!</div>'+msj)
                
                return JsonResponse({'success': True})
                
            
    else:
        if request.method == 'POST':
            
            user_check_perms(request,'crear presupuesto comercial',raise_exception=True)
            descriptions = request.POST.getlist('detail_description')
            values = request.POST.getlist('detail_value')
            note = request.POST.get('observations')
            
            budget = Commercial_budget.objects.create(
                user_request = request.user, notes = note,             
                project = obj_project   
            )
            
            for i, j in zip(descriptions, values):
                Commercial_budget_detail.objects.create(
                    budget = budget, value = j.replace(',',''),
                    concept = i
                )
                
            
    return render(request,'commercialbudget.html',context)


def incomes_actions(request, project):
    obj_project = Projects.objects.get(pk=project)
    if request.method == 'POST':
        income_receipt = request.POST.get('receipt_number')
        action = request.POST.get('action')

        if action == 'nullify':
            if not user_check_perms(request, 'anular recibo'):
                data = {
                    'type': 'error',
                    'title': 'Tenemos un problema',
                    'msj': 'No tienes los privilegios suficientes para hacer esto',
                }
                if request.is_ajax():
                    return JsonResponse(data)
                else:
                    raise PermissionDenied

            reload_page = request.POST.get('reload')
            obj_income = Incomes.objects.get(
                project=project,
                receipt=income_receipt
            )
            obj_income_detail = Incomes_detail.objects.filter(
                income=obj_income.pk
            )

            obj_nullify_method = Payment_methods.objects.get(
                name__icontains='Anulaciones'
            )

            for i in obj_income_detail:
                i.delete()

            obj_income.value = 0
            obj_income.description = 'Recibo Anulado'
            obj_income.payment_method = obj_nullify_method
            obj_income.save()

            Timeline.objects.create(
                user=request.user,
                action=f'Anuló el recibo Nº{income_receipt}',
                project=obj_project,
                aplication='finance'
            )

            filename = f'RC Anulado-{obj_income.receipt}_CTR{obj_income.sale.contract_number}_{project}.pdf'.replace(
                'ñ', 'n')

            pdf_file = pdf_gen(
                'pdf/income_recepit.html',
                {'income': obj_income},
                filename,
            )

            href = pdf_file.get('url')
            a = f'<a href="{href}" target="_blank"><strong>aqui</strong></a>'
            msj = f'Se anuló el recibo Nº{income_receipt}, puedes descargarlo '+a

            if reload_page:
                messages.success(
                    request, '<div class="header">¡Lo hicimos!</div>'+msj)

            if request.is_ajax():
                data = {
                    'type': 'success',
                    'title': '¡Lo hicimos!',
                    'msj': msj,
                }
                return JsonResponse(data)

            f = open(pdf_file.get('root'), 'rb')

            return FileResponse(f, as_attachment=True, filename=filename)

@project_permission
@user_permission('ver pmt')
def pmt(request, project):
    obj_project = Projects.objects.get(pk=project)

    pmt_sellers = Sellers.objects.filter(seller_state='Activo', pay_pmt=True,
                                         projects__name=project)

    pmt_value = Parameters.objects.get(name='valor pmt').value
    context = {
        'project': obj_project,
        'pmt_val': pmt_value,
        'active_pmt_sellers': pmt_sellers,
        'pay_accounts': payment_accounts.objects.all(),
    }

    if request.is_ajax():
        if request.method == 'GET':
            date = request.GET.get('date')
            data = {}
            if date:
                dt_date = datetime.strptime(date, '%d/%m/%Y')

                pmt_detail = PMT_detail.objects.filter(pmt__add_date=dt_date,
                                                       pmt__project=project,
                                                       value__gt=0)

                data = {
                    'data': JsonRender(pmt_detail, query_functions=['seller_fullname']).render()
                }
            return JsonResponse(data)

        if request.method == 'POST':
            action = request.POST.get('action')
            if action == 'approve':
                if not user_check_perms(request, 'aprobar pmt'):
                    return JsonResponse({}, status=401)

                date = request.POST.get('date')
                dt_date = datetime.strptime(date, '%d/%m/%Y')

                obj_pmt = PMT.objects.get(add_date=dt_date, project=project)
                obj_pmt.state = 'Aprobado'
                obj_pmt.user_approve = request.user
                obj_pmt.save()

                Timeline.objects.create(
                    user=request.user,
                    action=f'Aprobó el PMT del {date}',
                    project=obj_project,
                    aplication='finance'
                )

                messages.success(
                    request, '<div class="header">¡Lo hicimos!</div>Se aprobó el PMT seleccionado.')

            elif action == 'delete':
                if not user_check_perms(request, 'eliminar pmt'):
                    return JsonResponse({}, status=401)

                date = request.POST.get('date')
                dt_date = datetime.strptime(date, '%d/%m/%Y')

                obj_pmt = PMT.objects.get(add_date=dt_date, project=project)

                Timeline.objects.create(
                    user=request.user,
                    action=f'Eliminó el PMT del {date}',
                    project=obj_project,
                    aplication='finance'
                )

                obj_pmt.delete()

                messages.success(
                    request, '<div class="header">¡Lo hicimos!</div>Se eliminó el PMT seleccionado.')

            return JsonResponse({}, status=200)

    else:
        if request.method == 'POST':
            user_check_perms(request, 'crear pmt', raise_exception=True)
            sellers = request.POST.getlist('seller_pk')
            values = request.POST.getlist('pmt_value')
            obs = request.POST.get('observations')
            try:
                obj_pmt = PMT.objects.create(
                    project=obj_project,
                    observations=obs,
                    user=request.user,
                    state='Pendiente'
                )

                for i in range(0, len(sellers)):
                    value = int(values[i].replace(',', ''))
                    if value > 0:
                        obj_seller = Sellers.objects.get(pk=sellers[i])
                        PMT_detail.objects.create(
                            pmt=obj_pmt,
                            seller=obj_seller,
                            value=value
                        )

                messages.success(
                    request, '<div class="header">¡Lo hicimos!</div>Se creó el PMT sin problemas.')
            except IntegrityError:
                messages.error(
                    request, '<div class="header">Ups!</div>Ya existe un PMT liquidado para esta fecha y este proyecto, revisalo.')

    context.update({
        'pmts': PMT.objects.filter(project=project).order_by('-add_date'),
    })

    return render(request, 'pmt.html', context)


@project_permission
def collection_budget(request, project, portfolio_type=None):
        
    if not request.user.user_profile.has_permission(f'ver cartera {portfolio_type}'):
        raise PermissionDenied(
            'No tienes los permisos necesarios para ver esta cartera.')
    obj_project = Projects.objects.get(name=project)
    obj_budgets = Collection_budget.objects.filter(project=project)
    context = {
        'project': obj_project,
        'budgets': obj_budgets,
        'portfolio_type':portfolio_type,
        'form_collection_feed': collectionfeed_Form,
        'can_reassign' : request.user.user_profile.has_permission('reasignar gestor presupuesto'),
        'can_recalculate': request.user.user_profile.has_permission('recalcular presupuesto')
    }

    def create_budget_with_details(year:int, month:int):
        """Crea presupuesto y detalles para el periodo solicitado."""
        obj_project_local = Projects.objects.get(name=project)
        new_budget = Collection_budget.objects.create(
            project=obj_project_local,
            year=year,
            month=month,
            user=request.user
        )

        ventas = Sales_extra_info.objects.filter(status='Adjudicado', project=project)
        for venta in ventas:
            presupuesto = venta.budget(year, month)
            try:
                collector = venta.sale_collector.collector_user
            except Collector_per_sale.DoesNotExist:
                collector = request.user
            sale_portfolio_type = 'comercial' if venta.has_pending_ci_quota() else 'administrativa'
            if sale_portfolio_type == 'administrativa':
                prev = (
                    Collection_budget_detail.objects
                    .filter(
                        sale=venta,
                        budget__project=project
                    )
                    .exclude(
                        Q(budget__year=year, budget__month=month) |
                        Q(budget__year__gt=year) |
                        Q(budget__year=year, budget__month__gt=month)
                    )
                    .order_by('-budget__year', '-budget__month')
                    .first()
                )
                
                if prev and prev.portfolio_type == 'administrativa':
                    collector = prev.collector
                else:
                    if not obj_project_local.default_admin_collector:
                        # Buscar gestores administrativos
                        ids_adm = [p.usuario.pk for p in Perfil.objects.filter(
                            Q(rol__descripcion__icontains='gestor')&
                            Q(rol__descripcion__icontains='cartera')&
                            Q(rol__descripcion__icontains='administrativa'),
                            usuario__is_superuser=False,
                            usuario__is_active=True
                        ).distinct()]
                        
                        if not ids_adm:
                            # Si no hay administrativos, buscar comerciales
                            ids_com = [p.usuario.pk for p in Perfil.objects.filter(
                                Q(rol__descripcion__icontains='gestor')&
                                Q(rol__descripcion__icontains='cartera')&
                                Q(rol__descripcion__icontains='comercial'),
                                usuario__is_superuser=False,
                                usuario__is_active=True
                            ).distinct()]
                            
                            if not ids_com:
                                # Si no hay comerciales, usar superuser
                                ids_super = [u.pk for u in User.objects.filter(is_superuser=True, is_active=True)]
                                ids_adm = ids_super[:1] if ids_super else []
                            else:
                                ids_adm = ids_com[:1]
                        
                        if ids_adm:
                            obj_project_local.default_admin_collector_id = ids_adm[0]
                    collector = obj_project_local.default_admin_collector or request.user
            
            
            Collection_budget_detail.objects.create(
                budget=new_budget,
                sale=venta,
                portfolio_type=sale_portfolio_type,
                collector=collector,
                lt_30=presupuesto.get('lt_30'),
                lt_60=presupuesto.get('lt_60'),
                lt_90=presupuesto.get('lt_90'),
                lt_120=presupuesto.get('lt_120'),
                gt_120=presupuesto.get('gt_120')
            )
        return new_budget

    if request.is_ajax():
        if request.method == 'GET':
            
            today = date.today()
            month = int(request.GET.get('month',today.month))
            request_year = request.GET.get('year',today.year)
            year = int(request_year) if request_year != '' else today.year
            search_by = request.GET.get('search_by')
            to_search = request.GET.get('to_search')
            
            today = date.today()
            selected_date = date(year, month, 1)
            max_allowed_date = date(today.year, today.month, 1)

            if selected_date > max_allowed_date:
                return JsonResponse({'error': 'No se puede acceder a un periodo futuro.'}, status=400)

            # Validamos si ya existe
            if not Collection_budget.objects.filter(project=project, year=year, month=month).exists():
                with transaction.atomic():
                    create_budget_with_details(year, month)

            obj_budget = Collection_budget_detail.objects.filter(
                budget__month=month,
                budget__year=year,
                budget__project=project
            ).annotate(sale_name=Concat(F('sale__first_owner__first_name'), Value(' '),
                                        F('sale__first_owner__last_name')),
                       collector_name=Concat(F('collector__first_name'), Value(' '),
                                             F('collector__last_name'))
                       )
            
            
            if portfolio_type: obj_budget = obj_budget.filter(portfolio_type=portfolio_type)
            
            # Aplicar lógica de permisos para collector
            es_gestor_cartera = any(
                all(palabra in rol.descripcion.lower() for palabra in ['gestor', 'cartera'])
                for rol in request.user.user_profile.rol.all()
            )

            es_lider_cartera = any(
                all(palabra in rol.descripcion.lower() for palabra in ['lider', 'cartera'])
                for rol in request.user.user_profile.rol.all()
            )

            # Si es gestor sin ser líder, solo ve sus propias ventas
            if es_gestor_cartera and not es_lider_cartera:
                obj_budget = obj_budget.filter(collector=request.user)
            
            if search_by == 'sale':
                contract = to_search.split(' - ')[0]
                obj_budget = obj_budget.filter(sale__contract_number=contract,
                                               sale__project=project)
            elif search_by == 'collector':
                collector = to_search.split(' - ')[0]
                obj_budget = obj_budget.filter(collector=collector)

            entering = get_ci_status_notifications(project, year, month, obj_budget)
            collector_options = list(
                obj_budget.values(
                    'collector',
                    'collector__first_name',
                    'collector__last_name'
                ).distinct()
            )

            data = {
                'data': JsonRender(obj_budget,
                                   query_functions=('total', 'period_incomes'),
                                   annotates=('sale_name', 'collector_name')
                                   ).render(),
                'entering': entering,
                'collector_options': collector_options,
                
            }

            return JsonResponse(data)

        if request.method == 'POST':
            month = request.POST.get('month')
            year = request.POST.get('year')
            to_change = request.POST.get('to_change')
            to_search = request.POST.get('to_search')
            month_int = int(month)
            year_int = int(year)
            
            to_do = request.POST.get('to_do')
            obj_budget = Collection_budget_detail.objects.filter(
                budget__month=month,
                budget__year=year,
                budget__project=project
            )
            data = {}
            action = None
            if to_do == 'reasign':
                user_check_perms(
                    request, 'reasignar gestor presupuesto', raise_exception=True)
                new_user = request.POST.get('new-user-to-budget')
                if to_change == 'collector':
                    obj_budget = obj_budget.filter(collector=to_search, portfolio_type=portfolio_type)
                elif to_change == 'client':
                    obj_budget = obj_budget.filter(
                        sale__contract_number=to_search)

                for bg in obj_budget:
                    bg.collector = User.objects.get(pk=new_user)
                    bg.save()
                action = f'Cambió asesor asignado para venta(s) en el ppto del mes {month} año {year}'
                data = {
                    'type': 'success',
                    'title': '¡Tenemos lista tu solicitud!',
                    'msj': 'Se cambió el gestor asginado según tu solicitud.'
                }

            elif to_do == 'remove':
                user_check_perms(
                    request, 'eliminar ventas del presupuesto', raise_exception=True)
                if to_change == 'collector':
                    obj_budget = obj_budget.filter(collector=to_search)
                elif to_change == 'client':
                    obj_budget = obj_budget.filter(
                        sale__contract_number=to_search)

                for bg in obj_budget:
                    bg.delete()
                action = f'Eliminó venta(s) en el ppto del mes {month} año {year}'
                data = {
                    'type': 'success',
                    'title': '¡Tenemos lista tu solicitud!',
                    'msj': 'Se eliminó la(s) venta(s) del presupuesto actual.'
                }

            elif to_do == 'recalculate':
                user_check_perms(
                    request, 'recalcular presupuesto', raise_exception=True)
                if to_change == 'collector':
                    obj_budget = obj_budget.filter(collector=to_search)
                elif to_change == 'client':
                    obj_budget = obj_budget.filter(
                        sale__contract_number=to_search)

                for bg in obj_budget:
                    sale = Sales_extra_info.objects.get(pk=bg.sale.pk)
                    collection_budget = bg.budget
                    bg.delete()
                    budget = sale.budget(int(year), int(month))
                    Collection_budget_detail.objects.create(
                        budget=collection_budget,
                        sale=sale,
                        collector=budget.get('collector'),
                        lt_30=budget.get('lt_30'),
                        lt_60=budget.get('lt_60'),
                        lt_90=budget.get('lt_90'),
                        lt_120=budget.get('lt_120'),
                        gt_120=budget.get('gt_120')
                    )
                action = f'Recalculó el presupuesto de cobro en venta(s) para el ppto del mes {month} año {year}'
                data = {
                    'type': 'success',
                    'title': '¡Tenemos lista tu solicitud!',
                    'msj': 'Se recalculó el presupuesto de la venta seleccionada'
                }

            elif to_do == 'recalculate-all':
                user_check_perms(
                    request, 'recalcular presupuesto', raise_exception=True)
                selected_date = date(year_int, month_int, 1)
                max_allowed_date = date(date.today().year, date.today().month, 1)
                if selected_date > max_allowed_date:
                    return JsonResponse({
                        'type': 'error',
                        'title': 'Periodo no permitido',
                        'msj': 'No se puede recalcular un periodo futuro.'
                    }, status=400)
                try:
                    with transaction.atomic():
                        Collection_budget.objects.get(
                            month=month_int, year=year_int, project=project).delete()
                        create_budget_with_details(year_int, month_int)
                except Collection_budget.DoesNotExist:
                    return JsonResponse({
                        'type': 'error',
                        'title': 'No existe presupuesto',
                        'msj': 'No se encontró presupuesto para el periodo seleccionado.'
                    }, status=404)
                action = f'Recalculó todo el presupuesto de cartera para el ppto del mes {month} año {year}'
                data = {
                    'type': 'success',
                    'title': '¡Tenemos lista tu solicitud!',
                    'msj': 'Se recalculó todo el presupuesto del periodo y ya se actualizó la información.'
                }

            elif to_do == 'detele-budget-all':
                user_check_perms(
                    request, 'eliminar presupuesto completo', raise_exception=True)
                Collection_budget.objects.get(
                    month=month, year=year, project=project).delete()
                messages.success(
                    request, '<div class="header">¡Tenemos lista tu solicitud!</div>Se eliminó el presupuesto para el mes y año seleccionado')
                action = f'Eliminó el presupuesto de cartera para el ppto del mes {month} año {year}'

            if action:
                Timeline.objects.create(
                    user=request.user,
                    action=action,
                    project=obj_project,
                    aplication='collection'
                )
            return JsonResponse(data)

    else:
        if request.method == 'GET' and request.GET:
            month = request.GET.get('month')
            year = request.GET.get('year')
            obj_budget = Collection_budget_detail.objects.filter(
                budget__month=month,
                budget__year=year,
                budget__project=project
            )

            if not request.user.user_profile.has_rols(('Supervisor de cartera',)):
                obj_budget = obj_budget.filter(collector=request.user.pk)

            collectors = obj_budget.values('collector').order_by('collector'
                                                                 ).annotate(total=Count('collector'),
                                                                            collector_name=Concat(F('collector__first_name'),
                                                                                                  F('collector__last_name')))

            users_collectors = Perfil.objects.filter(
                rol__descripcion='Gestor de Cartera',
                usuario__is_active=True
            )

            context.update({
                'selected': True,
                'budget_sales': obj_budget,
                'month': month,
                'year': year,
                'collectors': collectors,
                'users_collectors': users_collectors
            })

        if request.method == 'POST':
            user_check_perms(
                request, 'agregar nuevo presupuesto', raise_exception=True)
            year = request.POST.get('year-to-add')
            month = request.POST.get('month-to-add')
            try:
                obj_collection = Collection_budget.objects.create(
                    project=obj_project,
                    month=month,
                    year=year,
                    user=request.user
                )
            except IntegrityError:
                messages.error(
                    request, f'<div class="header">¡Ups!</div>Ya existe un presupuesto cargado para el mes {month} del año {year}')
                return render(request, 'collection_budget.html', context)

            for sale in Sales_extra_info.objects.filter(status='Adjudicado', project=project):
                budget = sale.budget(int(year), int(month))
                Collection_budget_detail.objects.create(
                    budget=obj_collection,
                    sale=sale,
                    collector=budget.get('collector'),
                    lt_30=budget.get('lt_30'),
                    lt_60=budget.get('lt_60'),
                    lt_90=budget.get('lt_90'),
                    lt_120=budget.get('lt_120'),
                    gt_120=budget.get('gt_120')
                )

            messages.success(
                request, f'<div class="header">¡Lo hicimos!</div>Se ha cargado el presupuesto para el mes {month} del año {year}')
            context['budgets'] = Collection_budget.objects.filter(
                project=project)

    gc = Perfil.objects.filter(
                Q(rol__descripcion__icontains='gestor')&
                Q(rol__descripcion__icontains='cartera')|
                Q(rol__descripcion__icontains='lider'),
                usuario__is_active=True
                ).distinct()
                    
    
    obj_budgets = Collection_budget.objects.filter(project=project)
    
    context.update({
        'budgets': obj_budgets,
        "year": date.today().year,
        "month": date.today().month,
        "users_collectors" : gc,

    })

    return render(request, 'collection_budget.html', context)

@login_required
@project_permission
@user_permission('ver division de recaudos')
def incomes_division(request, project):
    obj_project = Projects.objects.get(name=project)

    context = {
        'project': obj_project,
        'lastcostcenter': cost_center.objects.filter(
            to_date__isnull=True, project=project,
        )
    }

    if request.is_ajax():
        if request.method == 'GET':
            month = request.GET.get('month-to-search')
            year = request.GET.get('year-to-search')
            month = int(month)
            year = int(year)
            month_data = calendar.monthrange(year, month)
            last_day = date(year, month, month_data[1])
            first_day = date(year, month, 1)
            cc_info = cost_center.objects.filter(
                from_date__gte=first_day, project=project,
            )
            data = {
                'exists': cc_info.exists(),
            }

            return JsonResponse(data)

    else:
        if request.method == 'GET' and request.GET:
            obj_incomes = Incomes.objects.filter(project=project)
            obj_expenses = expenses_detail.objects.filter(project=project)
            month = request.GET.get('month-to-search')
            year = request.GET.get('year-to-search')

            month = int(month)
            year = int(year)
            month_data = calendar.monthrange(year, month)
            last_day = date(year, month, month_data[1])
            first_day = date(year, month, 1)

            obj_incomes = obj_incomes.filter(
                add_date__gte=first_day,
                add_date__lte=last_day
            )
            obj_expenses = obj_expenses.filter(
                date__gte=first_day,
                date__lte=last_day
            )
            func = f'statics(month={int(month)},year={int(year)})'
            period = f'PERIODO: {year}-{month:02d}'

            obj_incomes = obj_incomes.aggregate(
                total=Sum('value')).get('total', 0)
            total_incomes = 0 if obj_incomes is None else obj_incomes

            obj_expenses = obj_expenses.aggregate(
                total=Sum('value')).get('total', 0)
            total_expenses = 0 if obj_expenses is None else obj_expenses

            cc_info = cost_center.objects.filter(
                Q(to_date__isnull=True) | Q(to_date__gte=last_day),
                from_date__lte=first_day, project=project,

            )

            try:
                rem = (total_incomes - total_expenses) * 100 / total_incomes
            except ZeroDivisionError:
                rem = (total_incomes - total_expenses) * 100

            context.update({
                'total_incomes': total_incomes,
                'total_expenses': total_expenses,
                'remaining': rem,
                'costcenter': JsonRender(cc_info, query_functions=[func, ]).render(),
                'period': period,
                'month': month,
                'year': year,
            })
        elif request.method == 'POST':
            month = request.POST.get('month')
            year = request.POST.get('year')
            cc_names = request.POST.getlist('cc_name')
            cc_perc = request.POST.getlist('cc_perc')
            month = int(month)
            year = int(year)
            month_data = calendar.monthrange(year, month)
            day_to_close = date(year, month, month_data[1])
            day_to_open = day_to_close + relativedelta(days=1)

            cc_to_close = cost_center.objects.filter(
                project = obj_project, to_date__isnull=True
            )
            for cc in cc_to_close:
                cc.to_date = day_to_close
                cc.save()

            for i in range (len(cc_names)):
                cost_center.objects.create(
                    name = cc_names[i],project = obj_project,
                    from_date = day_to_open, percentage = cc_perc[i]
                )

            messages.success(request,'<div class="header">¡Lo hicimos!</div>Se cerro la distribucion anterior.')


    return render(request, 'incomes_division.html', context)

@project_permission
@user_permission('ver gastos')
def expenses(request, project):
    obj_project = Projects.objects.get(name=project)

    context = {
        'project': obj_project,
        'form': new_expense_form(project=project)
    }

    if request.is_ajax():
        if request.method == 'GET':
            obj_expenses = expenses_detail.objects.filter(project=project)

            from_date = request.GET.get('from_date')
            to_date = request.GET.get('to_date')

            if from_date:
                from_dt = parse_semantic_date(from_date, 'date')
                to_dt = parse_semantic_date(to_date, 'date')
                obj_expenses = obj_expenses.filter(
                    date__gte=from_dt, date__lte=to_dt)

            data = {'data': JsonRender(obj_expenses).render()}

            return JsonResponse(data)

        if request.method == 'POST':
            to_do = request.POST.get('to_do')
            if to_do == 'remove':
                user_check_perms(request, 'eliminar gastos',
                                 raise_exception=True)
                id_expense = request.POST.get('id_expense')
                obj_expense = expenses_detail.objects.get(pk=id_expense)
                obj_expense.delete()

                data = {
                    'msj': 'La linea de gasto fue eliminada sin problemas.'
                }

                Timeline.objects.create(
                    action=f'Eliminó una linea de gasto',
                    aplication='finance',
                    user=request.user
                )

                return JsonResponse(data)
            if to_do == 'create':
                user_check_perms(request, 'registrar gastos',
                                 raise_exception=True)
                costcenter = request.POST.get('costcenter')
                description = request.POST.get('description')
                date = request.POST.get('date')
                value = request.POST.get('value')

                f_date = parse_semantic_date(date, 'date')
                f_value = value.replace(',', '')
                f_costcenter = cost_center.objects.get(pk=costcenter)

                if f_date.date() <  f_costcenter.from_date:
                    data = {
                        'cls': 'danger',
                        'msj': 'Estas intentando registrar un gasto en un periodo con centro de costos ya cerrado'
                    }

                    return JsonResponse(data)

                expenses_detail.objects.create(
                    project=obj_project,
                    date=f_date,
                    description=description,
                    value=f_value,
                    costcenter=f_costcenter,
                    user=request.user
                )

                data = {
                        'cls': 'success',
                    'msj': 'El gasto fue registrado sin problemas'
                }

                return JsonResponse(data)

    return render(request, 'expenses.html', context)

# ajax

from django.http import JsonResponse
from django.db.models import Sum
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import calendar

@login_required
def detalle_cliente(request, project):
    contrato = request.GET.get('contrato')
    if not contrato:
        return JsonResponse({'error': 'Contrato no especificado'}, status=400)

    try:
        venta = Sales.objects.get(contract_number=contrato, project__name=project)
    except Sales.DoesNotExist:
        return JsonResponse({'error': 'Venta no encontrada'}, status=404)

    hoy = datetime.today()
    historial = []

    # Últimos 6 meses de comportamiento
    for i in range(5, -1, -1):
        fecha_ref = hoy - relativedelta(months=i)
        y, m = fecha_ref.year, fecha_ref.month
        nombre_mes = fecha_ref.strftime('%b')

        # Busca el presupuesto de ese mes
        det = Collection_budget_detail.objects.filter(
            sale=venta,
            budget__year=y,
            budget__month=m
        ).first()

        cuota_mes = 0
        vencido = 0
        if det:
            cuota_mes = det.lt_30
            vencido = det.lt_60 + det.lt_90 + det.lt_120 + det.gt_120

        pagado = Incomes_detail.objects.filter(
            income__sale=venta,
            income__payment_date__year=y,
            income__payment_date__month=m
        ).aggregate(total=Sum(F('capital') + F('interest') + F('arrears') + F('others')))['total'] or 0

        historial.append({
            'mes': f"{nombre_mes} {y}",
            'cuota_mes': float(cuota_mes),
            'vencido': float(vencido),
            'pagado': float(pagado)
        })


    # Últimos 6 seguimientos
    seguimientos = Collection_feed.objects.filter(sale=venta).order_by('-add_date')[:6]
    seguimientos_json = [{
        'usuario': s.user.get_full_name(),
        'fecha': s.add_date.strftime('%d/%m/%Y'),
        'comentario': s.comment
    } for s in seguimientos]

    # Cuotas pactadas
    cuotas = Payment_plans.objects.filter(sale=venta).order_by('pay_date')

    # Días de mora
    cuota_vencida = next(
        (c for c in cuotas if c.pay_date < hoy.date() and c.paid() < c.total_payment()),
        None
    )

    dias_mora = (hoy.date() - cuota_vencida.pay_date).days if cuota_vencida else 0
    
    total_adeudado = 0
    for c in cuotas:
        if c.pay_date < hoy.date():  # Solo cuotas vencidas
            saldo = c.saldo()
            if saldo > 0:
                total_adeudado += saldo

    cuotas_mora = [
        c for c in cuotas
        if c.pay_date < hoy.date() and c.saldo() > 0
    ]
    num_cuotas_mora = len(cuotas_mora)
    
    ultimo_pago = Incomes_detail.objects.filter(
        income__sale=venta
    ).order_by('-income__payment_date').first()

    fecha_ultimo_pago = (
        ultimo_pago.income.payment_date.strftime('%d/%m/%Y')
        if ultimo_pago else '—'
    )


    # Días promedio de pago
    pagos = Incomes_detail.objects.filter(income__sale=venta).order_by('-income__payment_date')[:6]
    dias = sorted(set([p.income.payment_date.day for p in pagos]))

    return JsonResponse({
        'dias_mora': dias_mora,
        'dias_promedio_pago': dias,
        'historial': historial,
        'seguimientos': seguimientos_json,
        'total_adeudado': total_adeudado,
        'cuotas_mora': num_cuotas_mora,
        'fecha_ultimo_pago': fecha_ultimo_pago,
    })

def ajax_new_comment_feed(request, project):
    
    if request.is_ajax():
        if request.method == 'POST':
            sale = request.POST.get('sale')
            obj_sale = Sales.objects.get(pk=sale)
            comment_type_id = request.POST.get('comment_type')
            comment = request.POST.get('comment')
        
            # ✅ CORRECCIÓN: Obtener la instancia de CommentType
            try:
                comment_type = CommentType.objects.get(pk=comment_type_id)
            except CommentType.DoesNotExist:
                return JsonResponse({
                    'type': 'error',
                    'title': 'Error',
                    'msj': 'Tipo de comentario no válido',
                })

            comment_feed = Collection_feed.objects.create(
                sale=obj_sale,
                comment_type=comment_type,  # Usar la instancia en lugar del ID
                comment=comment,
                user=request.user
            )

            messages.success(
                request, '<div class="header">¡Lo hicimos!</div>Se adicionó un seguimiento')
            data = {
                'type': 'success',
                'title': '¡Lo hicimos!',
                'msj': 'Se adicionó un seguimiento',
            }

            return JsonResponse(data)

def ajax_comissions_actions(request, project):
    obj_project = Projects.objects.get(name=project)
    if request.is_ajax():
        if request.method == 'GET' and request.GET:
            aggrupation = request.GET.get('aggrupation')
            date = request.GET.get('date')
            user = request.GET.get('user')
            if date:
                date_dt = datetime.strptime(date, '%Y/%m/%d')
            else:
                date_dt = None
            if aggrupation == 'seller':
                obj_paid = Paid_comissions.objects.filter(
                    pay_date=date_dt, user__username=user
                )
                paid_detail = obj_paid.values(
                    'assign_paid__seller',
                ).order_by('assign_paid__seller').annotate(
                    seller_name=Concat(F('assign_paid__seller__first_name'), Value(
                        ' '), F('assign_paid__seller__last_name')),
                    total_comission=Sum('comission'),
                    total_provision=Sum('provision'),
                    total_net=Sum('net_payment')
                )
                data = {
                    'data': list(paid_detail)
                }
                return JsonResponse(data)

            elif aggrupation == 'sale':
                obj_paid = Paid_comissions.objects.filter(
                    pay_date=date_dt, user__username=user
                )
                paid_detail = obj_paid.values(
                    'assign_paid__sale'
                ).order_by('assign_paid__sale').annotate(
                    contract_number=F('assign_paid__sale__contract_number'),
                    first_owner=Concat(F('assign_paid__sale__first_owner__first_name'),
                                       Value(' '), F('assign_paid__sale__first_owner__last_name')),
                    method=F('type_of_payment'),
                    total_comission=Sum('comission'),
                    total_provision=Sum('provision'),
                    total_net=Sum('net_payment')
                )
                data = {
                    'data': list(paid_detail)
                }

                return JsonResponse(data)
            elif aggrupation == 'sale_by_seller':
                seller = request.GET.get('seller')
                obj_paid = Paid_comissions.objects.filter(
                    pay_date=date_dt, user__username=user, assign_paid__seller=seller,
                    assign_paid__sale__project=project
                )
                data = {
                    'data': JsonRender(obj_paid).render()
                }

                return JsonResponse(data)
            else:
                return JsonResponse(
                    {'data': 'You must send a request parameters'}
                )
        if request.method == 'POST':
            date = request.POST.get('date')
            username = request.POST.get('username')
            date_dt = datetime.strptime(date, '%Y/%m/%d')

            obj_paid = Paid_comissions.objects.filter(
                pay_date=date_dt, user__username=username
            )
            [pay.delete() for pay in obj_paid]
            Timeline.objects.create(
                user=request.user,
                action=f'Eliminó el pago de comision efectuado el {date} por {username}',
                project=obj_project,
                aplication='finance',
            )
            messages.success(
                request, '<div class="header">¡Lo hicimos!</div>Eliminaste el pago de la comision seleccionada')
            data = {}
            return JsonResponse(data)
    else:
        return JsonResponse(
            {'data': 'This view is ajax exclusive'}
        )

def ajax_print_comissions(request, project):
    obj_project = Projects.objects.get(name=project)
    if request.is_ajax():
        if request.method == 'GET' and request.GET:
            date = request.GET.get('date')
            user = request.GET.get('user')
            if date:
                date_dt = datetime.strptime(date, '%Y/%m/%d')
            else:
                date_dt = None
            
            full_comissions_paid = Paid_comissions.objects.filter(
                pay_date=date_dt, user__username=user
            ).order_by('assign_paid__sale__contract_number','assign_paid__seller')
                
            obj_paid = full_comissions_paid.exclude(type_of_payment='Anticipo')
            
            paid_detail = obj_paid.values(
                'assign_paid__seller', 'assign_paid__state'
            ).order_by('assign_paid__seller').annotate(
                seller_name=Concat(F('assign_paid__seller__first_name'), Value(
                    ' '), F('assign_paid__seller__last_name')),
                total_comission=Sum('comission'),
                total_provision=Sum('provision'),
                total_net=Sum('net_payment')
            )
            
            
            totals = obj_paid.aggregate(
                total_comission=Sum('comission'),
                total_provision=Sum('provision'),
                total_net=Sum('net_payment')
            )
            
            obj_ant = Paid_comissions.objects.filter(
                pay_date=date_dt, user__username=user,
                type_of_payment='Anticipo')
            
            
            paid_detail_ant = obj_ant.values(
                'assign_paid__seller', 'assign_paid__state'
            ).order_by('assign_paid__seller').annotate(
                seller_name=Concat(F('assign_paid__seller__first_name'), Value(
                    ' '), F('assign_paid__seller__last_name')),
                total_comission=Sum('comission'),
                total_provision=Sum('provision'),
                total_net=Sum('net_payment')
            )
                        
            totals_ant = obj_ant.aggregate(
                total_comission=Sum('comission'),
                total_provision=Sum('provision'),
                total_net=Sum('net_payment')
            )
            
            template = 'pdf/comissions_resume.html'
            context = {
                'full_comissions': full_comissions_paid,
                'date': date_dt,
                'user': user,
                'project': obj_project,
                'now': datetime.now(),
                'user': request.user,
                'comissions': paid_detail,
                'ant_detail': paid_detail_ant,
                'ant_exists':  obj_ant.exists(),
                'totals': totals,
                'totals_ant':totals_ant
            }
            filename = f'Liquidacion_comisiones_{obj_project.name_to_show}_{user}_{date_dt.date()}.pdf'.replace(
                'ñ', 'n')
            pdf = pdf_gen(template, context, filename)

            pdf_url = pdf.get('url')
            href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'msj': 'Puedes descargarlo haciendo click '+href
            }

            return JsonResponse(data)

def ajax_print_pmt(request, project):
    obj_project = Projects.objects.get(name=project)
    if request.is_ajax():
        if request.method == 'GET':
            todo = request.GET.get('todo')
            date = request.GET.get('date')
            date_dt = datetime.strptime(date, '%d/%m/%Y')

            obj_pmt = PMT.objects.get(
                add_date=date_dt, project=project
            )

            obj_pmts = PMT_detail.objects.filter(pmt=obj_pmt.pk)
            totals = obj_pmts.aggregate(
                total=Sum('value'),
            )
            
            if todo == 'print-pdf':
                template = 'pdf/pmt_resume.html'
                context = {
                    'pmt': obj_pmt,
                    'project': obj_project,
                    'now': datetime.now(),
                    'user': request.user,
                    'pmts': obj_pmts,
                    'totals': totals
                }
                filename = f'Liquidacion_PMT_{obj_project.name_to_show}_{date_dt.date()}.pdf'.replace(
                    'ñ', 'n')
                pdf = pdf_gen(template, context, filename)

                pdf_url = pdf.get('url')
                href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

                data = {
                    'type': 'success',
                    'title': '¡Tenemos listo el documento!',
                    'msj': 'Puedes descargarlo haciendo click '+href
                }

                return JsonResponse(data)
            
            elif todo == 'print-excel':
                wb = openpyxl.Workbook()
                ws = wb.active
                ws['D2'] = 'PAGO DE MOVILIDAD Y TELEFONIA'
                ws['D2'].font = Font(bold=True)
                ws['D3'] = 'NIT. 900.779.254 - 2'
                ws['D3'].font = Font(bold=True)
                ws['D4'] = 'SEDE BOGOTÁ'
                ws['D4'].font = Font(bold=True)
                ws['E4'] = obj_project.name_to_show
                ws['E4'].font = Font(bold=True)
                ws['D5'] = 'DIRECTOR: JORGE CANO'
                ws['D5'].font = Font(bold=True)
                ws['D6'] = f'HONORARIOS {obj_pmt.observations}'
                ws['D7'] = f'FECHA DE PAGO: {obj_pmt.add_date}'
                ws['D7'].font = Font(bold=True)
                
                ws._current_row = 9
                ws.append(["",'#','# IDENTIDAD','NOMBRE','CELULAR','DIRECCION','MUNICIPIO','VALOR','LETRAS',
                           'POR CONCEPTO DE','SE PAGÓ'])
                
                ws.merge_cells('K10:L10')
                ws['K11'] = 'SI'
                ws['L11'] = 'NO'
                
                ws.merge_cells('C2:C6')
                logo = Image(obj_project.logo)
                logo.width = 176
                logo.height = 150
                ws.add_image(logo,'C2')      
                
                cols = ['B','C','D','E','F','G','H','I','J']
                for i in cols:
                    ws.merge_cells(f'{i}10:{i}11')
                    ws[f'{i}10'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f'{i}10'].font = Font(bold=True)
                
                for row in ws['K10:L11']:
                    for cell in row:
                        cell.font = Font(bold=True) 
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                ws._current_row = 11
                
                i=1
                for seller in obj_pmts:
                    city = countries_data().city(seller.seller.city)
                    ws.append([
                       "", i , seller.seller.seller_document, seller.seller.full_name(),
                       seller.seller.phone, seller.seller.address, city, seller.value,
                       numbers_names(seller.value), 'PAGO DE MOVILIDAD Y TELEFONIA'
                    ])
                    i+=1
                j = i+11
                
                for row in ws[f'H12:H{j+12}']:
                    for cell in row:
                        cell.style = 'Comma' 
                        
                ws[f'G{j}'] = 'TOTAL'
                ws[f'G{j}'].font = Font(bold=True) 
                ws[f'G{j}'].alignment = Alignment(horizontal='right')
                ws[f'H{j}'] = f'=SUM(H12:H{j-1})'
                ws[f'H{j}'].font = Font(bold=True) 
                
                  
                
                thin_border = borders.Border(top=borders.Side(style='thin'),
                                             left = borders.Side(style='thin'),
                                             bottom=borders.Side(style='thin'),
                                             right=borders.Side(style='thin'))
                for row in ws[f'B10:L{j}']:
                    for cell in row:
                        cell.border = thin_border
                    
                ws[f'D{i+18}'].border = borders.Border(top=borders.Side(style='medium'))
                ws[f'D{i+18}'] = 'ELABORADO'
                ws[f'D{i+18}'].alignment = Alignment(horizontal='center')
                ws.merge_cells(f'D{i+18}:E{i+18}')
                
                ws[f'D{i+19}'] = obj_pmt.user.get_full_name().upper()
                ws[f'D{i+19}'].alignment = Alignment(horizontal='center')
                ws.merge_cells(f'D{i+19}:E{i+19}')
                
                ws[f'G{i+18}'].border = borders.Border(top=borders.Side(style='medium'))
                ws[f'G{i+18}'] = 'RECIBIDO'
                ws[f'G{i+18}'].alignment = Alignment(horizontal='center')
                ws.merge_cells(f'G{i+18}:I{i+18}')
                
                ws[f'G{i+19}'] = 'JORGE CANO'
                ws[f'G{i+19}'].alignment = Alignment(horizontal='center')
                ws.merge_cells(f'G{i+19}:I{i+19}')
                
                
                ws[f'E{i+23}'].border = borders.Border(top=borders.Side(style='medium'))
                ws[f'E{i+23}'] = 'AUTORIZADO'
                ws[f'E{i+23}'].alignment = Alignment(horizontal='center')
                ws.merge_cells(f'E{i+23}:G{i+23}')
                
                ws[f'E{i+24}'] = obj_pmt.user_approve.get_full_name().upper()
                ws[f'E{i+24}'].alignment = Alignment(horizontal='center')
                ws.merge_cells(f'E{i+24}:G{i+24}')
                
                ws.column_dimensions['A'].width = 3.64      
                ws.column_dimensions['B'].width = 3.64                            
                ws.column_dimensions['C'].width = 24.1                      
                ws.column_dimensions['D'].width = 33.55
                ws.column_dimensions['E'].width = 19.18                         
                ws.column_dimensions['F'].width = 27.7                      
                ws.column_dimensions['G'].width = 16.91
                ws.column_dimensions['H'].width = 17.18                         
                ws.column_dimensions['I'].width = 28.45                         
                ws.column_dimensions['J'].width = 29.3                      
                ws.column_dimensions['K'].width = 4.82
                ws.column_dimensions['L'].width = 4.82
                
                
                
                filename = f'tmp/Liquidacion_PMT_{obj_project.name_to_show}_{date_dt.date()}.xlsx'.replace(
                    'ñ', 'n')
                wb.save(settings.MEDIA_ROOT /filename )
                
                href = f'<a href="/media/{filename}" target="_blank"><strong>Aqui</strong></a>'

                data = {
                    'type': 'success',
                    'title': '¡Tenemos listo el documento!',
                    'msj': 'Puedes descargarlo haciendo click '+href
                }
                
                return JsonResponse(data)

            elif todo == 'print-document-support':
                pmt_seller = request.GET.get('detail_pk')
                obj_pmt_seller = PMT_detail.objects.get(pk=pmt_seller)
                template = 'pdf/pmt_doc_support.html'
                context = {
                    'pmt': obj_pmt,
                    'project': obj_project,
                    'now': datetime.now(),
                    'user': request.user,
                    'seller_pmt': obj_pmt_seller,
                }
                filename = f'Documento_soporte_PMT_{obj_project.name_to_show}_{date_dt.date()}.pdf'.replace(
                    'ñ', 'n')
                pdf = pdf_gen(template, context, filename)

                pdf_url = pdf.get('url')
                href = f'<a href="{pdf_url}" target="_blank"><strong>Aqui</strong></a>'

                data = {
                    'type': 'success',
                    'title': '¡Tenemos listo el documento!',
                    'msj': 'Puedes descargarlo haciendo click '+href
                }

                return JsonResponse(data)
            
def ajax_print_filetobank(request, project):
    obj_project = Projects.objects.get(pk=project)
    if request.is_ajax():
        if request.method == 'GET':
            type_of = request.GET.get('type_of')
            pay_account = request.GET.get('pay_account')

            obj_pay_account = payment_accounts.objects.get(pk=pay_account)
            proj_reference = obj_project.name_to_show.replace(" ", "").upper()

            base_file = settings.STATIC_ROOT/"files/InterfazPAB.xlsx"
            book = openpyxl.load_workbook(base_file)
            sheet = book.active
            sheet.cell(2, 1, obj_pay_account.nit_to_pay)
            sheet.cell(2, 2, 220)
            letter = random.choice(string.ascii_uppercase)
            number = random.randint(1, 9)
            sequence = f'{letter}{number}'
            sheet.cell(2, 3, "I")
            sheet.cell(2, 4, sequence)
            sheet.cell(2, 5, obj_pay_account.account_number)
            sheet.cell(2, 6, obj_pay_account.account_type)


            row = 4
            if type_of == 'pmt':
                
                sheet.cell(2, 7, 'PMT')
                pay_date = request.GET.get('date')
                date_dt = datetime.strptime(pay_date, '%d/%m/%Y')
                pay_detail = PMT_detail.objects.filter(
                    pmt__add_date=dt_date, pmt__project=project)
                proj_reference = obj_project.name_to_show.replace(
                    " ", "").upper()
                ref = f'PMT{proj_reference}'[:20]
                for line in pay_detail:
                    cc = line.seller.pk.split(' ')[-1]
                    
                    sheet.cell(row, 1, 1)



                    sheet.cell(row, 2, int(cc))
                    sheet.cell(row, 3, line.seller.full_name().upper()[:31])
                    if line.seller.account_type == "S":
                        trans_type = 37
                    elif line.seller.account_type == "D":
                        trans_type = 27
                    else:
                        trans_type = ''
                    sheet.cell(row, 4, trans_type)
                    sheet.cell(row, 5, line.seller.bank_entity.id_bank)
                    sheet.cell(row, 6, line.seller.bank_account_number)

                    sheet.cell(row, 9, ref)
                    sheet.cell(row, 11, line.value)
                    row += 1

                date_to_string = pay_date.replace('/', '-')
                doc_name = f'PAB_PMT_{proj_reference}_{date_to_string}.xlsx'

            elif type_of == 'comission':
                
                sheet.cell(2, 7, 'COMIS')
                comission_date = request.GET.get('comission_date')
                user_pay = request.GET.get('user_pay')
                date_dt = datetime.strptime(comission_date, '%Y/%m/%d')
                pay_detail = Paid_comissions.objects.filter(project=project, pay_date=date_dt,
                                                            user__username=user_pay)
                ref = f'COMIS{proj_reference}'[:20]

                for line in pay_detail:
                    cc = line.assign_paid.seller.pk.split(' ')[-1]
                    sheet.cell(row, 1, 1)
                    sheet.cell(row, 2, int(cc))
                    sheet.cell(
                        row, 3, line.assign_paid.seller.full_name().upper()[:31])
                    if line.assign_paid.seller.account_type == "S":
                        trans_type = 37
                    elif line.assign_paid.seller.account_type == "D":
                        trans_type = 27
                    else:
                        trans_type = ''
                    sheet.cell(row, 4, trans_type)
                    sheet.cell(row, 5, int(
                        line.assign_paid.seller.bank_entity.id_bank))
                    sheet.cell(row, 6, int(
                        line.assign_paid.seller.bank_account_number))

                    sheet.cell(row, 9, ref)
                    sheet.cell(row, 11, line.net_payment)
                    row += 1

                date_to_string = comission_date.replace('/', '-')
                doc_name = f'PAB_COMISIONES_{proj_reference}_{date_to_string}.xlsx'

            root = settings.MEDIA_ROOT/f'tmp/{doc_name}'
            book.save(root)

            href = f'<a href="/media/tmp/{doc_name}" target="_blank"><strong>aqui</strong></a>'
            href2 = f'<a href="https://www.satbancolombia.com/conversores#!/pab" target="_blank"><strong>aqui</strong></a>'
            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'msj': 'Puedes descargarlo haciendo click '+href+', y convertir tu archivo plano '+href2
            }

            return JsonResponse(data)

@login_required
def ajax_reapply_rc(request, project, sale_pk):
    if request.method == 'GET':
        try:
            with transaction.atomic():
                # Validar que la venta exista
                try:
                    obj_sale = Sales.objects.get(pk=sale_pk, project__name=project)
                except Sales.DoesNotExist:
                    return JsonResponse({
                        'response': 'error',
                        'message': f'Venta con ID {sale_pk} no encontrada en proyecto {project}'
                    })
                
                # Obtener recibos ordenados por fecha y número
                obj_incomes = Incomes.objects.filter(
                    sale=sale_pk
                ).order_by('add_date', 'receipt')
                
                if not obj_incomes.exists():
                    return JsonResponse({
                        'response': 'warning',
                        'message': 'No hay recibos para reaplicar en esta venta'
                    })
                
                # Eliminar aplicaciones actuales
                obj_incomes_details = Incomes_detail.objects.filter(
                    income__sale=sale_pk  # Corregir el filtro
                )
                
                deleted_count = obj_incomes_details.count()
                obj_incomes_details.delete()
                
                # Contadores para el reporte
                recibos_procesados = 0
                recibos_con_errores = 0
                errores_detalle = []
                
                # Reaplicar recibos uno por uno
                for income in obj_incomes:
                    try:
                        # Intentar aplicación normal
                        apply_income(income, condonate_arrears=100)
                        recibos_procesados += 1
                        
                    except ValueError as e:
                        error_msg = str(e)
                        
                        # Si es error de excedente, intentar con abono a capital
                        if ("Este pago aplica cuotas futuras" in error_msg or 
                            "El valor a pagar supera" in error_msg):
                            try:
                                apply_income(
                                    income, 
                                    condonate_arrears=100, 
                                    abono_capital=True, 
                                    tipo_abono="cuotas_futuras"
                                )
                                recibos_procesados += 1
                            except Exception as e2:
                                recibos_con_errores += 1
                                errores_detalle.append(f"Recibo {income.receipt}: {str(e2)}")
                        else:
                            # Error diferente, registrar y continuar
                            recibos_con_errores += 1
                            errores_detalle.append(f"Recibo {income.receipt}: {error_msg}")
                    
                    except Exception as e:
                        # Error no esperado
                        recibos_con_errores += 1
                        errores_detalle.append(f"Recibo {income.receipt}: Error inesperado - {str(e)}")
                
                # Registrar en historial solo si se procesó al menos un recibo
                if recibos_procesados > 0:
                    action_msg = f'Reaplicó {recibos_procesados} recibos de caja según plan de pagos vigente'
                    if recibos_con_errores > 0:
                        action_msg += f' ({recibos_con_errores} recibos con errores)'
                    
                    Sales_history.objects.create(
                        sale=obj_sale,
                        action=action_msg,
                        user=request.user
                    )
                
                # Preparar respuesta detallada
                if recibos_con_errores == 0:
                    messages.success(
                        request, 
                        f'<div class="header">Tarea completada</div>'
                        f'Se reaplicaron {recibos_procesados} recibos correctamente. '
                        f'Se eliminaron {deleted_count} aplicaciones anteriores.'
                    )
                    return JsonResponse({'response': 'ok'})
                else:
                    # Hubo errores pero también éxitos
                    if recibos_procesados > 0:
                        messages.warning(
                            request,
                            f'<div class="header">Tarea completada con advertencias</div>'
                            f'Se reaplicaron {recibos_procesados} recibos. '
                            f'{recibos_con_errores} recibos tuvieron errores.'
                        )
                        return JsonResponse({
                            'response': 'warning',
                            'message': f'Reaplicados: {recibos_procesados}, Errores: {recibos_con_errores}',
                            'errores': errores_detalle
                        })
                    else:
                        # Todos los recibos fallaron
                        messages.error(
                            request,
                            f'<div class="header">Error en la reaplicación</div>'
                            f'No se pudo reaplicar ningún recibo. Revisa el plan de pagos.'
                        )
                        return JsonResponse({
                            'response': 'error',
                            'message': 'Falló la reaplicación de todos los recibos',
                            'errores': errores_detalle
                        })
        
        except Exception as e:
            # Error crítico en la transacción
            return JsonResponse({
                'response': 'error', 
                'message': f'Error crítico en la reaplicación: {str(e)}'
            })
    
    return JsonResponse({'response': 'error', 'message': 'Método no permitido'})

def print_comissions_ds(request,project):
    if request.is_ajax():
        if request.method == 'GET':
            
            user_comission = request.GET.get('user')
            date_comission = request.GET.get('date').replace('/','-')
            seller = request.GET.get('seller')
            
            obj_project = Projects.objects.get(pk=project)
            obj_comission = Paid_comissions.objects.filter(user__username=user_comission,pay_date=date_comission,
                                                        project = project, assign_paid__seller = seller)
            seller = Sellers.objects.get(pk=seller)
            
            if project == 'altoscovenas':
                empresa = 'GRUPO GIBRALTAR S.A.S'
                nit_empresa = '900.779.254 - 2'
            else:
                empresa = 'CONSTRUCTORA MACARDY S.A.S'
                nit_empresa = '901.036.039 - 7'
            
            anticipos = obj_comission.filter(type_of_payment='Anticipo')
            comissions_list = Paid_comissions.objects.filter(
                user__username=user_comission,
                pay_date=date_comission,
                project=project,
                assign_paid__seller=seller
            ).exclude(type_of_payment='Anticipo')

            msj = 'Puedes descargar el(los) documentos a continuación:<ul>'

            if comissions_list.exists():
                wb = openpyxl.Workbook()
                ws = wb.active
                
                ws.column_dimensions['A'].width = 3.64      
                ws.column_dimensions['B'].width = 19                      
                ws.column_dimensions['C'].width = 28.4                    
                ws.column_dimensions['D'].width = 18.7
                ws.column_dimensions['E'].width = 19.1
                ws.column_dimensions['F'].width = 18.4             
                ws.column_dimensions['G'].width = 3.64   
                
                ws.merge_cells('B2:B7')
                logo = Image(obj_project.logo)
                logo.width = 176*2/3
                logo.height = 150*2/3
                ws.add_image(logo,'B2')   
                
                thin_border = borders.Border(top=borders.Side(style='thin'),
                                             left = borders.Side(style='thin'),
                                             bottom=borders.Side(style='thin'),
                                             right=borders.Side(style='thin'))
                
                
                ws['C1'] = 'Fecha de operación'
                ws['C1'].font = Font(bold=True)
                ws['D1'] = date_comission
                ws['C3'] = empresa
                ws['C3'].font = Font(bold=True)
                ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
                ws['C4'] = f'NIT. {nit_empresa}'
                ws['C4'].font = Font(bold=True)
                ws['C4'].alignment = Alignment(horizontal='center', vertical='center')
                ws['C6'] = 'DOCUMENTO SOPORTE'
                ws['C6'].font = Font(bold=True)
                ws['C6'].alignment = Alignment(horizontal='center', vertical='center')
                ws['C7'] = 'DEBE A'
                ws['C7'].font = Font(bold=True)
                ws['C7'].alignment = Alignment(horizontal='center', vertical='center')
                ws['E2'] = 'PROYECTO'
                ws['E2'].font = Font(bold=True)
                ws['F2'] = obj_project.name_to_show
                ws['E5'] = 'Número:______________________________'
                ws['E5'].font = Font(bold=True)
                ws['E6'] = obj_project.text_ds
                ws['E6'].font = Font(bold=True)
                ws['E6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                ws['B9'] = f'Vendedor o quien presta el servicio'
                ws['B9'].font = Font(bold=True)
                ws['B9'].alignment = Alignment(horizontal='center', vertical='center')
                ws['B9'].fill = PatternFill("solid",fgColor="00808080")
                ws['B9'].border = thin_border
                
                
                ws['F9'] = f'Nit/C.C.'
                ws['F9'].font = Font(bold=True)
                ws['F9'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F9'].fill = PatternFill("solid",fgColor="00808080")
                ws['F9'].border = thin_border
                
                ws['B10'] = seller.full_name()
                ws['B10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['B10'].border = thin_border
                
                ws['F10'] = seller.pk
                ws['F10'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F10'].border = thin_border
                
                ws['B11'] = f'Dirección'
                ws['B11'].font = Font(bold=True)
                ws['B11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['B11'].fill = PatternFill("solid",fgColor="00808080")
                ws['B11'].border = thin_border
                
                ws['D11'] = f'Municipio'
                ws['D11'].font = Font(bold=True)
                ws['D11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['D11'].fill = PatternFill("solid",fgColor="00808080")
                ws['D11'].border = thin_border
                
                ws['F11'] = f'Telefonos'
                ws['F11'].font = Font(bold=True)
                ws['F11'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F11'].fill = PatternFill("solid",fgColor="00808080")
                ws['B11'].border = thin_border
                
                ws['B12'] = seller.address
                ws['B12'].alignment = Alignment(horizontal='center', vertical='center')
                ws['B12'].border = thin_border
                
                cd = countries_data()
                ws['D12'] = f'{cd.country(seller.country)}, {cd.state(seller.state)}, {cd.city(seller.city)}'
                ws['D12'].alignment = Alignment(horizontal='center', vertical='center')
                ws['D12'].border = thin_border
                
                ws['F12'] = seller.phone
                ws['F12'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F12'].border = thin_border
                
                ws['B14'] = 'CONCEPTO'
                ws['B14'].font = Font(bold=True)
                ws['B14'].alignment = Alignment(horizontal='center', vertical='center')
                ws['B14'].border = thin_border
                ws['B14'].fill = PatternFill("solid",fgColor="00808080")
                
                ws['D14'] = 'VALOR BRUTO'
                ws['D14'].font = Font(bold=True)
                ws['D14'].alignment = Alignment(horizontal='center', vertical='center')
                ws['D14'].border = thin_border
                ws['D14'].fill = PatternFill("solid",fgColor="00808080")
                
                ws['E14'] = 'RETEFUENTE'
                ws['E14'].font = Font(bold=True)
                ws['E14'].alignment = Alignment(horizontal='center', vertical='center')
                ws['E14'].border = thin_border
                ws['E14'].fill = PatternFill("solid",fgColor="00808080")
                
                ws['F14'] = 'VALOR NETO'
                ws['F14'].font = Font(bold=True)
                ws['F14'].alignment = Alignment(horizontal='center', vertical='center')
                ws['F14'].border = thin_border
                ws['F14'].fill = PatternFill("solid",fgColor="00808080")
                comiss = comissions_list  
                i=15
                for ctr in comiss:                
                    ws[f'B{i}'] = f'PAGO COMISIONES CTR {ctr.assign_paid.sale.contract_number} - CARGO: {ctr.assign_paid.position.name}'
                    
                    
                    ws.merge_cells(f'B{i}:C{i}')
                    ws[f'D{i}'] = ctr.comission
                    ws[f'D{i}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    
                    ws[f'E{i}'] = ctr.provision
                    ws[f'E{i}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    ws[f'F{i}'] = ctr.net_payment
                    ws[f'F{i}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    i+=1
                
                
                for row in ws[f'D15:F{i}']:
                    for cell in row:
                        cell.style = 'Comma' 
                        
                for row in ws[f'B15:F{i}']:
                    for cell in row:
                        cell.border = thin_border
                
                j = i + 1
                
                for row in ws[f'F{j}:F{j+8}']:
                    for cell in row:
                        cell.style = 'Comma' 
                
                ws[f'E{j}'] = 'SUBTOTAL'
                ws[f'E{j}'].font = Font(bold=True)
                ws[f'E{j}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'E{j}'].border = thin_border
                ws[f'E{j}'].fill = PatternFill("solid",fgColor="00808080")
                
                ws[f'F{j}'] = f'=SUM(D15:D{i})'
                ws[f'F{j}'].font = Font(bold=True)
                ws[f'F{j}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'F{j}'].border = thin_border
                ws[f'F{j}'].fill = PatternFill("solid",fgColor="00808080")
                
                ws[f'E{j+2}'] = 'Retefuente'
                ws[f'E{j+2}'].font = Font(bold=True)
                ws[f'E{j+2}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'E{j+2}'].border = thin_border
                
                ws[f'F{j+2}'] = f'=SUM(E15:E{i})'
                ws[f'F{j+2}'].font = Font(bold=True)
                ws[f'F{j+2}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'F{j+2}'].border = thin_border
                
                ws[f'E{j+3}'] = 'ReteICA'
                ws[f'E{j+3}'].font = Font(bold=True)
                ws[f'E{j+3}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'E{j+3}'].border = thin_border
                ws[f'F{j+3}'].border = thin_border
                
                ws[f'E{j+4}'] = 'TOTAL'
                ws[f'E{j+4}'].font = Font(bold=True)
                ws[f'E{j+4}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'E{j+4}'].border = thin_border
                ws[f'E{j+4}'].fill = PatternFill("solid",fgColor="00808080")
                
                ws[f'F{j+4}'] = f'=F{j}-F{j+2}-F{j+3}'
                ws[f'F{j+4}'].font = Font(bold=True)
                ws[f'F{j+4}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'F{j+4}'].border = thin_border
                ws[f'F{j+4}'].fill = PatternFill("solid",fgColor="00808080")
                
                
                ws[f'E{j+5}'] = 'IVA implícito derivados del petróleo'
                ws[f'E{j+5}'].font = Font(bold=True)
                ws[f'E{j+5}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws[f'E{j+5}'].border = thin_border
                ws[f'F{j+5}'].border = thin_border
                
                ws.row_dimensions[j+5].height  = 32.5
                
                ws[f'B{j+8}'] = 'ACEPTO QUE NO SOY RESPONSABLE DE IVA'
                ws[f'B{j+8}'].font = Font(bold=True)
                ws[f'B{j+8}'].alignment = Alignment(horizontal='center', vertical='center')
                
                ws[f'B{j+8}'].border = borders.Border(top=borders.Side(style='medium'))
                
                
                ws[f'B{j+10}'] = '''Documento soporte de costos y gastos en operaciones con no obligados a expedir factura o documento equivalente'''
                ws[f'B{j+10}'].font = Font(bold=True)
                ws[f'B{j+10}'].alignment = Alignment(horizontal='center', vertical='center')
                
                ws[f'B{j+10}'].border = borders.Border(top=borders.Side(style='medium'))
                
                ws[f'B{j+11}'] = '''Articulo 1.6.1.4.12 Decreto Único reglamentario en materia tributaria 1625 de 2016 - Sustituido por el Decreto 358 de 2020'''
                ws[f'B{j+11}'].font = Font(bold=True)
                ws[f'B{j+11}'].alignment = Alignment(horizontal='center', vertical='center')
                
                ws.merge_cells('C3:D3') 
                ws.merge_cells('C4:D4') 
                ws.merge_cells('C6:D6')
                ws.merge_cells('C7:D7')  
                ws.merge_cells('E5:F5')    
                ws.merge_cells('E6:F7')
                
                ws.merge_cells('B9:E9') 
                ws.merge_cells('B10:E10')
                ws.merge_cells('B11:C11') 
                ws.merge_cells('B12:C12') 
                ws.merge_cells('D11:E11') 
                ws.merge_cells('D12:E12') 
                
                ws.merge_cells('B14:C14')
                
                ws.merge_cells(f'B{j+8}:C{j+8}')
                ws.merge_cells(f'B{j+10}:F{j+10}')
                ws.merge_cells(f'B{j+11}:F{j+11}')
                
                filename = f'tmp/Documento_soporte_comisiones_{seller.full_name()}.xlsx'.replace(
                    'ñ', 'n').replace(' ','_')
                wb.save(settings.MEDIA_ROOT /filename )
                
                href = f'<a href="/media/{filename}" target="_blank"><strong>Descargar</strong></a>'
                
                msj += '<li>Documento soporte comisiones: '+href + '</li>'

            if not anticipos.exists() and not comissions_list.exists():
                return JsonResponse({
                    'type': 'warning',
                    'title': 'Sin pagos para este filtro',
                    'msj': 'No encontramos comisiones pagadas (anticipos ni liquidaciones) con los criterios seleccionados.'
                })

            msj += '</ul>'
            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'msj': msj if comissions_list.exists() else 'Hay anticipos registrados, pero no liquidaciones para descargar.'
            }

            return JsonResponse(data)
            
from django.views.decorators.http import require_GET

@login_required
@require_GET
def get_cartera_proyecto_data(request, project_id, tipo_cartera=None):
    if not tipo_cartera:
        tipo_cartera = request.GET.get("tipo_cartera", "todos")

    ventas = Sales_extra_info.objects.filter(
        project_id=project_id,
        status='Adjudicado'
    ).select_related('first_owner')

    data = []
    for venta in ventas:
        cuotas = Credit_info.objects.filter(sale=venta)

        tiene_ci_pendiente = cuotas.filter(
            quota_type='CI',
            capital__gt=0
        ).exclude(
            incomes_detail__capital__gte=F('capital')
        ).exists()

        es_comercial = tiene_ci_pendiente and not venta.is_paid()
        es_financiera = not es_comercial

        if tipo_cartera == "comercial" and not es_comercial:
            continue
        if tipo_cartera == "financiera" and not es_financiera:
            continue

        total_pagado = venta.total_payment()
        saldo = venta.remain_value()
        mora = Incomes_detail.objects.filter(quota__sale=venta).aggregate(
            total_mora=Sum('arrears')
        ).get('total_mora') or 0

        try:
            collector = venta.sale_collector.collector_user.get_full_name()
        except:
            collector = '—'

        data.append({
            'contrato': f'CTR{venta.contract_number}',
            'cliente': venta.first_owner.full_name(),
            'tipo_cartera': "Comercial" if es_comercial else "Financiera",
            'pagado': f'${total_pagado:,.0f}',
            'saldo': f'${saldo:,.0f}',
            'mora': f'${mora:,.0f}',
            'gestor': collector,
            'ver': f'<a href="/cartera/comportamiento/{venta.id_sale}/" class="ui mini button">Ver</a>'
        })

    return JsonResponse({'data': data})

@login_required
@project_permission
@user_permission('ver solicitudes de recibos')
def lista_solicitudes_recibo(request, project):
    """Vista para listar solicitudes de recibo"""
    # Obtener el objeto proyecto
    obj_project = Projects.objects.get(name=project)
    
    # Filtros de fecha
    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')
    
    if not fecha_desde or not fecha_hasta:
        # Valores por defecto si no se especifican
        fecha_hasta = date.today()
        fecha_desde = fecha_hasta - timedelta(days=30)
    else:
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    
    # Obtener solicitudes ordenadas por fecha de creación descendente (más recientes primero)
    solicitudes = SolicitudRecibo.objects.filter(
        sale__project__name=project,
        add_date__gte=fecha_desde,
        add_date__lte=fecha_hasta
    ).select_related(
        'sale', 'sale__first_owner', 'creado_por', 'revisado_por', 
        'condonacion_autorizada_por', 'recibo_generado'
    ).order_by('-add_date', '-id')
    
    context = {
        'project': obj_project,  # Agregar el objeto proyecto al contexto
        'solicitudes': solicitudes,
        'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
        'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d'),
    }
    
    return render(request, 'solicitudes_lista.html', context)

@login_required
def crear_solicitud_recibo(request, project):
    proyecto = get_object_or_404(Projects, name=project)
    is_modal = (
        request.method == 'GET' and request.GET.get('modal') == '1'
    ) or (
        request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest'
    )
    
    action_url = reverse('crear_solicitud_recibo', kwargs={'project': project})
    if request.method == 'POST':
        form = SolicitudReciboForm(request.POST, request.FILES, project=proyecto)
        form.instance.project = proyecto

        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.project = proyecto
            solicitud.creado_por = request.user
            solicitud.estado = 'pendiente'
            solicitud.save()

            if is_modal:
                return JsonResponse({'result': 'ok'})

            messages.success(request, 'Solicitud creada correctamente.')
            return redirect('lista_solicitudes_recibo', project=project)
        else:
            return render(request, 'solicitud_form.html', {
                'form': form,
                'project': project,
                'action': action_url
            })
    else:
        form = SolicitudReciboForm(project=proyecto)

    template = 'solicitud_form.html'
    
    return render(request, template, {
        'form': form,
        'project': project,
        'action': action_url
    })

@login_required
def editar_solicitud_recibo(request, project, pk):
    proyecto = get_object_or_404(Projects, name=project)
    solicitud = get_object_or_404(SolicitudRecibo, pk=pk, project=proyecto)

    is_modal = (
        request.method == 'GET' and request.GET.get('modal') == '1'
    ) or (
        request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest'
    )

    if solicitud.estado != 'pendiente':
        if is_modal:
            return JsonResponse({'result': 'forbidden'})
        messages.error(request, 'Solo se pueden editar solicitudes pendientes.')
        return redirect('lista_solicitudes_recibo', project=project)

    if solicitud.condonacion_autorizada:
        form = SolicitudReciboForm(request.POST or None, request.FILES or None, instance=solicitud, project=proyecto)
        if request.method == 'POST':
            data = request.POST.copy()
            data['arrears_condonate'] = solicitud.arrears_condonate
            form.data = data
        form.fields['arrears_condonate'].disabled = True
        
        if request.method == 'POST' and form.is_valid():
            solicitud_editada = form.save(commit=False)
            solicitud_editada.condonacion_autorizada = True
            solicitud_editada.condonacion_autorizada_por = solicitud.condonacion_autorizada_por
            solicitud_editada.arrears_condonate = solicitud.arrears_condonate
            solicitud_editada.save()
    else:
        form = SolicitudReciboForm(request.POST or None, request.FILES or None, instance=solicitud, project=proyecto)
    
    action_url = reverse('editar_solicitud_recibo', kwargs={'project': project, 'pk': pk})
    template = 'solicitud_form.html'

    if request.method == 'POST':
        form.instance.project = proyecto
        if form.is_valid():
            form.save()
            if is_modal:
                return JsonResponse({'result': 'ok'})
            messages.success(request, 'Solicitud actualizada.')
            return redirect('lista_solicitudes_recibo', project=project)

    return render(request, template, {
        'form': form,
        'project': project,
        'action': action_url
    })

@login_required
def eliminar_solicitud_recibo(request, project, pk):
    proyecto = get_object_or_404(Projects, name=project)
    solicitud = get_object_or_404(SolicitudRecibo, pk=pk, project=proyecto)

    if solicitud.estado != 'pendiente':
        messages.error(request, 'Solo puedes eliminar solicitudes pendientes.')
    else:
        solicitud.delete()
        messages.success(request, 'Solicitud eliminada correctamente.')

    return redirect('lista_solicitudes_recibo', project=project)

def _validar_solicitud_venta_nueva(request, obj_project, solicitud, solicitud_pk):
    """
    Valida solicitudes de recibo para ventas NO adjudicadas (Pendiente/Aprobado).
    Reutiliza la lógica de new_sales_incomes() con las ventajas del sistema de solicitudes:
    - Trazabilidad (solicitud → recibo)
    - Archivos de soporte
    - Segregación de funciones (quien solicita ≠ quien valida)
    """

    # Validaciones de estado de solicitud
    if solicitud.estado == 'aprobado':
        messages.error(request, "La solicitud ya está aprobada.")
        return redirect('lista_solicitudes_recibo', project=obj_project.name)

    # Preparar datos iniciales para el formulario (formato de new_sale_income_form)
    # IMPORTANTE: Usar locale inglés temporalmente para formatear fechas
    # porque el formulario/calendario JavaScript espera formato en inglés
    import locale as locale_module

    current_locale = locale_module.getlocale(locale_module.LC_TIME)
    try:
        locale_module.setlocale(locale_module.LC_TIME, 'en_US.UTF-8')
    except:
        try:
            locale_module.setlocale(locale_module.LC_TIME, 'C')  # Fallback a locale C (inglés básico)
        except:
            pass  # Si falla, continuar con el locale actual

    initial_data = {
        'sale': solicitud.sale.pk,
        'add_date': solicitud.add_date.strftime('%B %d, %Y') if solicitud.add_date else datetime.today().strftime('%B %d, %Y'),
        'payment_date': solicitud.payment_date.strftime('%B %d, %Y') if solicitud.payment_date else '',
        'payment_method_1': solicitud.pm1.pk if solicitud.pm1 else None,
        'value_1': solicitud.value1 or '',
        'payment_method_2': solicitud.pm2.pk if solicitud.pm2 else None,
        'value_2': solicitud.value2 or '',
        'description': solicitud.description or '',
        'receipt_number': '',  # Se auto-genera
    }

    # Restaurar locale original
    try:
        locale_module.setlocale(locale_module.LC_TIME, current_locale)
    except:
        pass

    if request.method == 'POST':
        try:
            # Validar estado
            solicitud.refresh_from_db()
            if solicitud.estado != 'pendiente':
                messages.error(request, "Esta solicitud ya fue procesada anteriormente.")
                return redirect('lista_solicitudes_recibo', project=obj_project.name)

            with transaction.atomic():
                # Bloquear solicitud
                solicitud = SolicitudRecibo.objects.select_for_update().get(pk=solicitud_pk)

                if solicitud.estado != 'pendiente':
                    raise ValueError("Solicitud ya procesada por otro usuario.")

                solicitud.estado = 'procesando'
                solicitud.save()

                # Obtener y bloquear contador
                obj_counter = Counters.objects.select_for_update().get(
                    project=obj_project,
                    name='recibos'
                )
                numero_recibo = obj_counter.value

                # Parsear fechas (igual que new_sales_incomes líneas 92-94)
                add_date = request.POST.get('add_date')
                dt_add_date = datetime.strptime(add_date, '%B %d, %Y')
                payment_date = request.POST.get('payment_date')
                dt_payment_date = datetime.strptime(payment_date, '%B %d, %Y')
                description = request.POST.get('description')

                # Preparar valores (líneas 104-112 de new_sales_incomes)
                value_1 = request.POST.get('value_1').replace(',', '')
                value_2 = request.POST.get('value_2')

                if value_2:
                    value_2 = value_2.replace(',', '')
                else:
                    value_2 = 0

                total_income = int(value_1) + int(value_2)

                # Métodos de pago (líneas 114-121)
                payment_method_1 = request.POST.get('payment_method_1')
                pm1 = Payment_methods.objects.get(pk=payment_method_1)
                pm2 = None
                payment_method_2 = request.POST.get('payment_method_2')
                if payment_method_2:
                    pm2 = Payment_methods.objects.get(pk=payment_method_2)

                # ⚠️ SIN apply_income() - Crear Income directo (líneas 123-131)
                income = Incomes.objects.create(
                    project=obj_project,
                    sale=solicitud.sale,
                    receipt=numero_recibo,
                    add_date=dt_add_date,
                    payment_date=dt_payment_date,
                    value=total_income,
                    payment_method=pm1,
                    description=description,
                    user=request.user,
                    value1=value_1,
                    value2=value_2,
                    pm1=pm1,
                    pm2=pm2
                )

                # Incrementar contador (líneas 140-141)
                obj_counter.value += 1
                obj_counter.save()

                # ✅ VINCULAR recibo con solicitud (ventaja del sistema de solicitudes)
                solicitud.recibo_generado = income
                solicitud.estado = 'aprobado'
                solicitud.revisado_por = request.user
                solicitud.confirmado_en = timezone.now()
                solicitud.save()

                # Timeline para auditoría
                Timeline.objects.create(
                    user=request.user,
                    action=f'Validó solicitud #{solicitud_pk} (venta {solicitud.sale.get_status_display()}) - Recibo #{numero_recibo}',
                    project=obj_project,
                    aplication='finance'
                )

                # Generar PDF (líneas 143-165)
                divisor = 1
                if income.sale.second_owner.full_name() != " ":
                    divisor += 1
                if income.sale.third_owner.full_name() != " ":
                    divisor += 1

                divided_income = f'{int(income.value)/divisor:,.2f}'

                filename = f'Recibocaja{income.receipt}_CTR{income.sale.contract_number}_{obj_project.name}.pdf'.replace('ñ', 'n')

                pdf_file = pdf_gen(
                    'pdf/income_recepit.html',
                    {'income': income, 'divided_income': divided_income},
                    filename,
                )
                pdf_url = pdf_file.get('url')
                href = f'<a href="{pdf_url}" target="_blank"><strong>Aquí</strong></a>'

                messages.success(
                    request,
                    f'<div class="header">¡Solicitud validada!</div>Recibo #{numero_recibo} generado correctamente. Descárgalo {href}'
                )
                return redirect('lista_solicitudes_recibo', project=obj_project.name)

        except ValueError as ve:
            messages.error(request, str(ve))
            return redirect('lista_solicitudes_recibo', project=obj_project.name)

        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(
                f"Error validando solicitud venta nueva {solicitud_pk}: {str(e)}",
                exc_info=True,
                extra={
                    'solicitud_pk': solicitud_pk,
                    'user': request.user.username,
                    'sale': solicitud.sale.pk if solicitud.sale else None,
                }
            )
            messages.error(request, f'Error al validar la solicitud: {e}')
            return redirect('lista_solicitudes_recibo', project=obj_project.name)

    else:  # GET
        form = new_sale_income_form(initial=initial_data, project=obj_project)

    context = {
        'project': obj_project,
        'form': form,
        'solicitud': solicitud,
        'validacion_solicitud': True,
        'es_venta_nueva': True,  # Flag para el template
    }

    return render(request, 'new_sale_incomes_validation.html', context)

@login_required
@project_permission
@user_permission('validar solicitud recibo')
def validar_solicitud_recibo(request, project, pk):
    obj_project = Projects.objects.get(name=project)
    solicitud = get_object_or_404(SolicitudRecibo, pk=pk)

    # 🔑 BIFURCACIÓN: Detectar tipo de venta según su status
    if solicitud.sale.status in ['Pendiente', 'Aprobado']:
        # ✅ Flujo para ventas NO adjudicadas (usa lógica de new_sales_incomes)
        return _validar_solicitud_venta_nueva(request, obj_project, solicitud, pk)

    # ⚠️ TODO LO SIGUIENTE ES EL FLUJO ORIGINAL PARA VENTAS ADJUDICADAS (NO TOCAR)
    # Validación de reestructuración pendiente
    from sales.models import PaymentPlanRestructuring
    tiene_reestructuracion_pendiente = PaymentPlanRestructuring.objects.filter(
        sale=solicitud.sale, status='Pendiente'
    ).exists()
    if tiene_reestructuracion_pendiente:
        messages.error(
            request,
            "No se puede validar la solicitud porque la venta tiene una reestructuración pendiente."
        )
        return redirect('lista_solicitudes_recibo', project=project)

    # Validaciones de estado
    if solicitud.arrears_condonate > 0 and not solicitud.condonacion_autorizada:
        messages.error(request, "La condonación debe ser autorizada antes de validar el recaudo.")
        return redirect('lista_solicitudes_recibo', project=project)
    elif solicitud.estado == 'aprobado':
        messages.error(request, "La solicitud ya está aprobada.")
        return redirect('lista_solicitudes_recibo', project=project)

    # Datos iniciales para el formulario
    solicitud_add_date = solicitud.add_date
    solicitud_add_date_str = solicitud_add_date.strftime('%Y-%m-%d') if solicitud_add_date else ''

    initial_data = {
        'sale': solicitud.sale.pk if solicitud.sale else None,
        'add_date': solicitud_add_date_str,
        'payment_day': solicitud.payment_date.strftime('%Y-%m-%d') if solicitud.payment_date else '',
        'payment_method_1': solicitud.pm1.pk if solicitud.pm1 else None,
        'value_1': solicitud.value1 or '',
        'payment_method_2': solicitud.pm2.pk if solicitud.pm2 else None,
        'value_2': solicitud.value2 or '',
        'arrears_condonate': solicitud.arrears_condonate,
        'description': solicitud.description or '',
        'capital_payment': solicitud.capital_payment,
        'tipo_abono_capital': solicitud.tipo_abono_capital or '',
    }
    
    if request.method == 'POST':
        # ✅ CAMBIO 1: Verificar estado antes de procesar
        solicitud.refresh_from_db()
        if solicitud.estado != 'pendiente':
            messages.error(request, "Esta solicitud ya fue procesada anteriormente.")
            return redirect('lista_solicitudes_recibo', project=project)
        
        data = request.POST.copy()
        data['add_date'] = solicitud_add_date

        # Parsear fechas
        try:
            payment_day = parse_semantic_date(data.get('payment_day', ''), 'date')

            if not payment_day:
                payment_day = solicitud.payment_date

            data['payment_day'] = payment_day

        except Exception:
            data['payment_day'] = solicitud.payment_date

        form = incomes_form(data, project=project, lock_add_date=True)
        
        if form.is_valid():
            # ✅ CAMBIO 2: Preparar y validar ANTES de la transacción
            try:
                # Preparar valores
                value_1 = int(data.get('value_1', '0').replace(',', ''))
                value_2 = int(data.get('value_2', '0').replace(',', '')) if data.get('value_2') else 0
                total_income = value_1 + value_2
                
                tipo_abono = data.get('tipo_abono_capital')
                abono_capital = data.get('capital_payment') == 'on'
                arrears_condonate = int(data.get('arrears_condonate', 0))
                
                # ✅ CAMBIO 3: VALIDACIÓN PREVIA (simulación sin crear nada)
                data_simulacion = {
                    'sale': solicitud.sale.pk,
                    'total_income': total_income,
                    'paid_day': data['payment_day'],
                    'rate': solicitud.sale.sale_plan.rate,
                }
                
                resultado_simulacion = apply_income(
                    None,  # No pasamos income real
                    condonate_arrears=arrears_condonate,
                    apply=False,  # ← MODO SIMULACIÓN
                    no_apply_data=data_simulacion,
                    abono_capital=abono_capital,
                    tipo_abono=tipo_abono
                )
                
                # Verificar si la simulación detectó errores
                if isinstance(resultado_simulacion, dict) and 'error' in resultado_simulacion:
                    messages.error(request, resultado_simulacion['error'])
                    return redirect('lista_solicitudes_recibo', project=project)
                
                # ✅ CAMBIO 4: UNA SOLA TRANSACCIÓN ATÓMICA
                # Todo o nada: si algo falla, se revierte TODO
                with transaction.atomic():
                    # ✅ CAMBIO 5: Bloqueo pesimista desde el inicio
                    # select_for_update() bloquea el registro para evitar race conditions
                    solicitud = SolicitudRecibo.objects.select_for_update().get(pk=pk)
                    
                    # Verificar estado dentro de la transacción
                    if solicitud.estado != 'pendiente':
                        raise ValueError("Esta solicitud ya fue procesada por otro usuario.")
                    
                    # Cambiar estado a procesando inmediatamente
                    solicitud.estado = 'procesando'
                    solicitud.save()
                    
                    # ✅ CAMBIO 6: Bloquear contador con select_for_update()
                    # Evita que dos usuarios obtengan el mismo número de recibo
                    obj_counter = Counters.objects.select_for_update().get(
                        project=obj_project, 
                        name='recibos'
                    )
                    numero_recibo = obj_counter.value
                    
                    # Obtener métodos de pago
                    pm1 = Payment_methods.objects.get(pk=data.get('payment_method_1'))
                    pm2 = None
                    if data.get('payment_method_2'):
                        pm2 = Payment_methods.objects.get(pk=data.get('payment_method_2'))
                    
                    # ✅ CAMBIO 7: Crear income
                    # Ahora sabemos que no fallará porque lo validamos antes
                    income = Incomes.objects.create(
                        project=obj_project,
                        sale=solicitud.sale,
                        receipt=numero_recibo,
                        add_date=data['add_date'],
                        payment_date=data['payment_day'],
                        value=total_income,
                        payment_method=pm1,
                        description=data.get('description', ''),
                        user=request.user,
                        value1=value_1,
                        value2=value_2,
                        pm1=pm1,
                        pm2=pm2
                    )
                    
                    # ✅ CAMBIO 8: Aplicar income (SIN @transaction.atomic en la función)
                    # Se ejecuta dentro de esta transacción
                    apply_income(
                        income, 
                        condonate_arrears=arrears_condonate, 
                        abono_capital=abono_capital, 
                        tipo_abono=tipo_abono
                    )
                    
                    # ✅ CAMBIO 9: Incrementar contador SOLO si todo fue exitoso
                    # Antes estaba aquí, pero es mejor al final
                    obj_counter.value += 1
                    obj_counter.save()
                    
                    # ✅ CAMBIO 10: Actualizar solicitud a aprobado
                    # Refrescar por si acaso hubo cambios
                    solicitud.refresh_from_db()
                    solicitud.recibo_generado = income
                    solicitud.estado = 'aprobado'
                    solicitud.revisado_por = request.user
                    solicitud.confirmado_en = timezone.now()
                    solicitud.save()
                    
                    # ✅ CAMBIO 11: Timeline de auditoría
                    Timeline.objects.create(
                        user=request.user,
                        action=f'Validó solicitud de recibo #{solicitud.pk} - Recibo #{numero_recibo}',
                        project=obj_project,
                        aplication='finance'
                    )
                
                # ✅ Si llegamos aquí, el commit fue exitoso
                messages.success(
                    request, 
                    f'Solicitud validada correctamente. Recibo #{numero_recibo} generado.'
                )
                return redirect('lista_solicitudes_recibo', project=project)
                
            except ValueError as ve:
                # ✅ CAMBIO 12: Manejo específico de errores de validación
                messages.error(request, str(ve))
                return redirect('lista_solicitudes_recibo', project=project)
                
            except Exception as e:
                # ✅ CAMBIO 13: Logging detallado para debugging
                import logging
                logger = logging.getLogger(__name__)
                logger.error(
                    f"Error validando solicitud {pk}: {str(e)}", 
                    exc_info=True,  # Incluye el stack trace completo
                    extra={
                        'solicitud_pk': pk,
                        'user': request.user.username,
                        'sale': solicitud.sale.pk if solicitud.sale else None,
                        'total_income': total_income,
                    }
                )
                
                # Si hay error, transaction.atomic hace rollback automático
                # TODO se revierte: income, contador, estado de solicitud
                messages.error(request, f'Error al validar la solicitud: {e}')
                return redirect('lista_solicitudes_recibo', project=project)
        else:
            # Errores en el formulario
            messages.error(request, 'Hay errores en el formulario. Verifica los datos.')
            return redirect('lista_solicitudes_recibo', project=project)
    else:
        # GET request - mostrar formulario
        form = incomes_form(initial=initial_data, project=project, lock_add_date=True)
        
    context = {
        'project': obj_project,
        'form': form,
        'solicitud': solicitud,
        'validacion_solicitud': True,
    }
    return render(request, 'incomes.html', context)

@login_required
@user_permission('autorizar condonacion de mora')
def autorizar_condonacion(request, project, pk):
    sol = get_object_or_404(SolicitudRecibo, pk=pk, project=project)
    perfil = request.user.user_profile
    sol.condonacion_autorizada = True
    sol.condonacion_autorizada_por = request.user
    sol.save()
    Timeline.objects.create(
        user=request.user,
        action=f'Autorizó condonación de mora en solicitud #{sol.pk}',
        project=sol.project,
        aplication='finance'
    )
    messages.success(request, "Condonación autorizada.")
    return redirect('lista_solicitudes_recibo', project=project)

@login_required
@user_permission('modificar parametros cartera')
def parametros(request, project):
    
    obj_p = Projects.objects.get(name=project)
    tasa = Parameters.objects.get(name='tasa de mora mv').value
    ids_com = [p.usuario.pk for p in Perfil.objects.filter(
        Q(rol__descripcion__icontains='gestor')&
        Q(rol__descripcion__icontains='cartera')&
        Q(rol__descripcion__icontains='comercial'),
        usuario__is_superuser=False,
        usuario__is_active=True
    ).distinct()]

    ids_adm = [p.usuario.pk for p in Perfil.objects.filter(
        Q(rol__descripcion__icontains='gestor')&
        Q(rol__descripcion__icontains='cartera')&
        Q(rol__descripcion__icontains='administrativa'),
        usuario__is_superuser=False,
        usuario__is_active=True
    ).distinct()]

    for uid in ids_com:
        ComisionGestorCartera.objects.get_or_create(usuario_id=uid, tipo_cartera='comercial', defaults={'porcentaje_comision': 0})
    for uid in ids_adm:
        ComisionGestorCartera.objects.get_or_create(usuario_id=uid, tipo_cartera='administrativa', defaults={'porcentaje_comision': 0})

    gestores_comercial = ComisionGestorCartera.objects.filter(usuario_id__in=ids_com, tipo_cartera='comercial')
    gestores_administrativa = ComisionGestorCartera.objects.filter(usuario_id__in=ids_adm, tipo_cartera='administrativa')

    param, _ = Parameters.objects.get_or_create(
        name='rangos_comision',
        project=obj_p,
        defaults={'section': 'comisiones', 'state': True, 'json': [
            {"min": 10000000, "max": 25000000, "porc": 2.0},
            {"min": 25000001, "max": 50000000, "porc": 1.75},
            {"min": 50000001, "max": None, "porc": 1.5}
        ]}
    )

    return render(request, 'parameters.html', {
        'project': obj_p,
        'tasa_mora': tasa,
        'gestores_comercial': gestores_comercial,
        'gestores_administrativa': gestores_administrativa,
        'rangos_comision_json': json.dumps(param.json or []),
    })

@login_required
@project_permission
@user_permission('ver comisiones de cartera')
def comisiones_cartera(request, project):
    obj_p = Projects.objects.get(name=project)
    hoy = timezone.now()
    anio = int(request.GET.get('anio', hoy.year))
    mes = int(request.GET.get('mes', hoy.month))
    anios = range(hoy.year-5, hoy.year+1)
    meses = [(i, timezone.datetime(2000, i, 1).strftime('%B').capitalize()) for i in range(1,13)]
    
    # Verificar si es gestor o líder
    es_gestor_cartera = any(
        all(palabra in rol.descripcion.lower() for palabra in ['gestor', 'cartera'])
        for rol in request.user.user_profile.rol.all()
    )
    
    es_lider_cartera = any(
        all(palabra in rol.descripcion.lower() for palabra in ['lider', 'cartera'])
        for rol in request.user.user_profile.rol.all()
    )
    
    gestores = ComisionGestorCartera.objects.filter(usuario__is_active=True)
    
    # Si es gestor pero no líder, solo ver sus propios datos
    if es_gestor_cartera and not es_lider_cartera:
        gestores = gestores.filter(usuario=request.user)
    
    gestores = gestores.distinct('usuario')
    
    return render(request, 'comisiones_cartera.html', {
        'project': obj_p,
        'anio': anio,
        'mes': mes,
        'anios': anios,
        'meses': meses,
        'gestores': gestores,
        'es_lider': es_lider_cartera,
        'es_gestor': es_gestor_cartera and not es_lider_cartera
    })
    
@login_required
def ajax_comisiones_cartera(request, project):
    anio = int(request.GET.get('anio'))
    mes = int(request.GET.get('mes'))
    
    # Verificar si es gestor o líder
    es_gestor_cartera = any(
        all(palabra in rol.descripcion.lower() for palabra in ['gestor', 'cartera'])
        for rol in request.user.user_profile.rol.all()
    )
    
    es_lider_cartera = any(
        all(palabra in rol.descripcion.lower() for palabra in ['lider', 'cartera'])
        for rol in request.user.user_profile.rol.all()
    )
    
    # Si es gestor pero no líder, forzar filtro por usuario actual
    if es_gestor_cartera and not es_lider_cartera:
        request.GET = request.GET.copy()
        request.GET['gestor'] = request.user.id
    
    gestores = get_gestores_comisiones(project, anio, mes, req=request)
    
    html = render_to_string('tabla_comisiones_cartera.html', {
        'gestores': gestores,
        'es_lider': es_lider_cartera,
        'es_gestor': es_gestor_cartera and not es_lider_cartera
    })
    return HttpResponse(html)


@login_required
def ajax_abonos_capital(request, project):
    from finance.models import AbonoCapital
    anio = int(request.GET.get('anio'))
    mes = int(request.GET.get('mes'))
    gestor_id = request.GET.get('gestor')
    tipo_cartera = request.GET.get('tipo_cartera')

    # Verificar si es gestor o líder
    es_gestor_cartera = any(
        all(palabra in rol.descripcion.lower() for palabra in ['gestor', 'cartera'])
        for rol in request.user.user_profile.rol.all()
    )
    es_lider_cartera = any(
        all(palabra in rol.descripcion.lower() for palabra in ['lider', 'cartera'])
        for rol in request.user.user_profile.rol.all()
    )

    # Si es gestor pero no líder, forzar filtro por usuario actual
    if es_gestor_cartera and not es_lider_cartera:
        gestor_id = request.user.id

    abonos = AbonoCapital.objects.filter(
        income__payment_date__year=anio,
        income__payment_date__month=mes,
        sale__project__name=project
    ).select_related('income', 'sale')

    if gestor_id:
        abonos = [a for a in abonos if str(a.sale.collection_budget_detail_set.filter(
            budget__year=a.income.payment_date.year,
            budget__month=a.income.payment_date.month
        ).first().collector_id if a.sale.collection_budget_detail_set.filter(
            budget__year=a.income.payment_date.year,
            budget__month=a.income.payment_date.month
        ).exists() else getattr(getattr(a.sale, "sale_collector", None), "collector_user_id", None)) == str(gestor_id)]

    if tipo_cartera:
        abonos = [a for a in abonos if any(
            det.portfolio_type == tipo_cartera and
            det.budget.year == a.income.payment_date.year and
            det.budget.month == a.income.payment_date.month
            for det in a.sale.collection_budget_detail_set.all()
        )]

    html = render_to_string('tabla_abonos_capital.html', {
        'abonos': abonos,
        'es_lider': es_lider_cartera,
        'es_gestor': es_gestor_cartera and not es_lider_cartera
    })
    return HttpResponse(html)

@require_POST
@login_required
def ajax_set_tasa_mora(request, project):
    perfil = request.user.user_profile
    if not perfil.has_permission('modificar parametros cartera'):
        return JsonResponse({'status':'error','msg':'No tienes permiso para modificar la tasa.'})
    tasa = request.POST.get('tasa')
    try:
        param = Parameters.objects.get(name='tasa de mora mv')
        param.value = tasa
        param.save()
        Timeline.objects.create(
            user=request.user,
            action=f'Cambió la tasa de mora a {tasa}%',
            project=param.project if hasattr(param, 'project') else None,
            aplication='finance'
        )
        return JsonResponse({'status':'ok','tasa':tasa})
    except Exception as e:
        return JsonResponse({'status':'error','msg':str(e)})

from finance.models import ComisionGestorCartera, AbonoCapital

@require_POST
@login_required
def ajax_set_comision_gestor(request, project):
    perfil = request.user.user_profile
    if not (perfil.has_permission('ver cartera comercial') or perfil.has_permission('ver cartera administrativa')):
        return JsonResponse({'status':'error','msg':'Sin permiso'})
    id = request.POST.get('id')
    valor = request.POST.get('valor')
    try:
        obj = ComisionGestorCartera.objects.get(id=id)
        obj.porcentaje_comision = valor
        obj.save()
        Timeline.objects.create(
            user=request.user,
            action=f'Cambió comisión gestor {obj.usuario} {obj.tipo_cartera} a {valor}%',
            project=Projects.objects.get(name=project),
            aplication='finance'
        )
        return JsonResponse({'status':'ok'})
    except Exception as e:
        return JsonResponse({'status':'error','msg':str(e)})


@login_required
@project_permission
def preview_reversion_abono(request, project, abono_id):
    """
    Endpoint para obtener preview de la reversión de un abono a capital
    Retorna información del backup vs estado actual
    """
    from sales.models import backup_payment_plans

    try:
        # Obtener el abono
        abono = get_object_or_404(AbonoCapital, id=abono_id, sale__project__name=project)
        sale = abono.sale

        # Validar que no tenga comisión pagada
        if abono.comision_pagado_a is not None or abono.valor_comision:
            nombre_asesor = abono.comision_pagado_a.get_full_name() if abono.comision_pagado_a else "un asesor"
            return JsonResponse({
                'can_revert': False,
                'reason': f'No se puede revertir este abono porque ya tiene una comisión catch-out pagada a {nombre_asesor}. '
                          f'Debe revertir primero la comisión antes de poder revertir el abono.',
                'error_type': 'comision_pagada'
            })

        # Buscar backup inmediatamente anterior al abono
        backup_anterior = backup_payment_plans.objects.filter(
            sale=sale,
            backup_date__lt=abono.fecha
        ).values('backup_date').distinct().order_by('-backup_date').first()

        if not backup_anterior:
            return JsonResponse({
                'can_revert': False,
                'reason': 'No se encontró backup anterior a este abono',
                'error_type': 'no_backup'
            })

        backup_fecha = backup_anterior['backup_date']

        # Obtener cuotas del backup
        cuotas_backup = backup_payment_plans.objects.filter(
            sale=sale,
            backup_date=backup_fecha
        )

        # Calcular info del backup
        backup_info = {
            'fecha': backup_fecha.strftime('%d/%m/%Y %H:%M'),
            'total_cuotas': cuotas_backup.count(),
            'cuotas_pendientes': cuotas_backup.filter(
                id_payment__isnull=True
            ).count() if cuotas_backup.filter(id_payment__isnull=True).exists() else cuotas_backup.count(),
            'saldo': float(sum(c.capital + c.interest for c in cuotas_backup)),
            'tipos': {}
        }

        # Contar por tipo
        for tipo in cuotas_backup.values('quota_type').distinct():
            tipo_name = tipo['quota_type']
            count = cuotas_backup.filter(quota_type=tipo_name).count()
            backup_info['tipos'][tipo_name] = count

        # Obtener plan actual
        cuotas_actual = Payment_plans.objects.filter(sale=sale)

        # Calcular info actual
        actual_info = {
            'total_cuotas': cuotas_actual.count(),
            'cuotas_pendientes': cuotas_actual.filter(
                id_payment__isnull=True
            ).count() if cuotas_actual.filter(id_payment__isnull=True).exists() else cuotas_actual.count(),
            'saldo': float(sum(c.capital + c.interest for c in cuotas_actual)),
            'tipos': {}
        }

        # Contar por tipo
        for tipo in cuotas_actual.values('quota_type').distinct():
            tipo_name = tipo['quota_type']
            count = cuotas_actual.filter(quota_type=tipo_name).count()
            actual_info['tipos'][tipo_name] = count

        # Buscar abonos posteriores
        abonos_posteriores = AbonoCapital.objects.filter(
            sale=sale,
            fecha__gt=abono.fecha
        ).count()

        # Buscar recibos posteriores (excluyendo los que son abonos)
        recibos_posteriores_ids = AbonoCapital.objects.filter(
            sale=sale,
            fecha__gte=abono.fecha
        ).values_list('income_id', flat=True)

        recibos_posteriores = Incomes.objects.filter(
            sale=sale,
            payment_date__gt=abono.fecha
        ).exclude(
            id__in=recibos_posteriores_ids
        ).count()

        return JsonResponse({
            'can_revert': True,
            'abono': {
                'id': abono.id,
                'fecha': abono.fecha.strftime('%d/%m/%Y'),
                'recibo': abono.income.receipt,
                'valor': float(abono.capital_aplicado),
                'tipo': abono.get_tipo_display(),
                'cuotas_afectadas': abono.cuotas_afectadas,
                'cliente': sale.first_owner.full_name(),
                'contrato': sale.contract_number
            },
            'backup': backup_info,
            'actual': actual_info,
            'recibos_posteriores': recibos_posteriores,
            'abonos_posteriores': abonos_posteriores
        })

    except Exception as e:
        return JsonResponse({
            'can_revert': False,
            'reason': f'Error al obtener información: {str(e)}',
            'error_type': 'exception'
        }, status=500)


@login_required
@project_permission
@require_POST
def revertir_abono_capital(request, project, abono_id):
    """
    Endpoint para ejecutar la reversión de un abono a capital
    """
    from sales.models import backup_payment_plans

    try:
        # Obtener motivo
        motivo = request.POST.get('motivo', '').strip()
        if not motivo:
            return JsonResponse({
                'status': 'error',
                'message': 'El motivo es obligatorio'
            })

        # Obtener el abono
        abono = get_object_or_404(AbonoCapital, id=abono_id, sale__project__name=project)
        sale = abono.sale

        # Validar que no tenga comisión pagada
        if abono.comision_pagado_a is not None or abono.valor_comision:
            nombre_asesor = abono.comision_pagado_a.get_full_name() if abono.comision_pagado_a else "un asesor"
            return JsonResponse({
                'status': 'error',
                'message': f'No se puede revertir este abono porque ya tiene una comisión catch-out pagada a {nombre_asesor}. '
                          f'Debe revertir primero la comisión antes de poder revertir el abono.'
            })

        # Buscar backup anterior
        backup_anterior = backup_payment_plans.objects.filter(
            sale=sale,
            backup_date__lt=abono.fecha
        ).values('backup_date').distinct().order_by('-backup_date').first()

        if not backup_anterior:
            return JsonResponse({
                'status': 'error',
                'message': 'No se encontró backup anterior para restaurar'
            })

        backup_fecha = backup_anterior['backup_date']

        # Iniciar transacción atómica
        with transaction.atomic():
            # 1. Guardar info del abono para el log
            abono_info = f"Abono {abono.id}: {abono.income.receipt} - ${abono.capital_aplicado} ({abono.get_tipo_display()})"

            # 2. Obtener abonos y recibos posteriores ANTES de eliminar
            abonos_posteriores = list(AbonoCapital.objects.filter(
                sale=sale,
                fecha__gt=abono.fecha
            ).order_by('fecha').values(
                'income_id', 'tipo', 'capital_aplicado'
            ))

            recibos_posteriores_ids = AbonoCapital.objects.filter(
                sale=sale,
                fecha__gte=abono.fecha
            ).values_list('income_id', flat=True)

            recibos_posteriores = list(Incomes.objects.filter(
                sale=sale,
                payment_date__gt=abono.fecha
            ).exclude(
                id__in=recibos_posteriores_ids
            ).order_by('add_date', 'receipt').values_list('id', flat=True))

            # 3. Eliminar TODO el plan de pagos actual
            Payment_plans.objects.filter(sale=sale).delete()

            # 4. Eliminar todos los Incomes_detail
            Incomes_detail.objects.filter(income__sale=sale).delete()

            # 5. Eliminar el AbonoCapital
            abono.delete()

            # 6. Restaurar plan desde backup
            cuotas_backup = backup_payment_plans.objects.filter(
                sale=sale,
                backup_date=backup_fecha
            )

            import re
            for cuota_backup in cuotas_backup:
                # Recalcular id_quota para cuotas ABCAP si usan sale.pk en lugar de contract_number
                id_quota = cuota_backup.id_quota
                if cuota_backup.quota_type == 'ABCAP':
                    match = re.match(r'ABCAP(\d+)CTR(\d+)', id_quota)
                    if match:
                        numero_secuencia = match.group(1)
                        numero_en_codigo = int(match.group(2))
                        # Si el código usa pk en lugar de contract_number, corregirlo
                        if numero_en_codigo != sale.contract_number:
                            id_quota = f"ABCAP{numero_secuencia}CTR{sale.contract_number}"

                Payment_plans.objects.create(
                    id_payment=cuota_backup.id_payment,
                    id_quota=id_quota,
                    quota_type=cuota_backup.quota_type,
                    sale=cuota_backup.sale,
                    pay_date=cuota_backup.pay_date,
                    capital=cuota_backup.capital,
                    interest=cuota_backup.interest,
                    others=cuota_backup.others,
                    project=cuota_backup.project
                )

            # 7. Re-aplicar abonos posteriores
            for abono_post in abonos_posteriores:
                income = Incomes.objects.get(id=abono_post['income_id'])
                try:
                    apply_income(
                        income=income,
                        condonate_arrears=100,
                        abono_capital=True,
                        tipo_abono=abono_post['tipo']
                    )
                except Exception as e:
                    # Si falla, intentar con cuotas_futuras
                    apply_income(
                        income=income,
                        condonate_arrears=100,
                        abono_capital=True,
                        tipo_abono='cuotas_futuras'
                    )

            # 8. Re-aplicar recibos posteriores
            for recibo_id in recibos_posteriores:
                income = Incomes.objects.get(id=recibo_id)
                try:
                    apply_income(income=income, condonate_arrears=100)
                except ValueError as e:
                    # Si no cabe, intentar como abono a capital
                    if "supera" in str(e).lower() or "futuras" in str(e).lower():
                        apply_income(
                            income=income,
                            condonate_arrears=100,
                            abono_capital=True,
                            tipo_abono='cuotas_futuras'
                        )
                    else:
                        raise

            # 9. Registrar en historial
            Sales_history.objects.create(
                sale=sale,
                action=f'Revirtió abono a capital: {abono_info}. Motivo: {motivo}',
                user=request.user
            )

            Timeline.objects.create(
                user=request.user,
                action=f'Revirtió abono a capital en CTR{sale.contract_number}. Motivo: {motivo}',
                project=sale.project,
                aplication='finance'
            )

            return JsonResponse({
                'status': 'ok',
                'message': 'Abono revertido correctamente',
                'abonos_reaplicados': len(abonos_posteriores),
                'recibos_reaplicados': len(recibos_posteriores)
            })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Error al revertir: {str(e)}'
        }, status=500)

@login_required
@project_permission
@user_permission('ver abonos a capital')
def abonos_capital_periodo(request, project):
    obj_project = get_object_or_404(Projects, name=project)
    hoy = timezone.now()
    anio = int(request.GET.get('anio', hoy.year))
    mes = int(request.GET.get('mes', hoy.month))
    anios = range(hoy.year-5, hoy.year+1)
    meses = [(i, timezone.datetime(2000, i, 1).strftime('%B').capitalize()) for i in range(1,13)]
    gestores = ComisionGestorCartera.objects.filter(usuario__is_active=True).distinct('usuario')
    return render(request, 'abonos_a_capital.html', {
        'project': obj_project,
        'anio': anio,
        'mes': mes,
        'anios': anios,
        'meses': meses,
        'gestores': gestores,
    })

@require_POST
@login_required
def pagar_comision_abono(request, project):
    abono_id = request.POST.get('abono_id')
    pagado_a = request.POST.get('pagado_a')
    valor = request.POST.get('valor_comision')
    fecha_pago = request.POST.get('fecha_pago')
    porc = request.POST.get('porcentaje_comision')
    try:
        abono = AbonoCapital.objects.get(id=abono_id)
        user = User.objects.get(id=pagado_a, is_active=True)
        abono.valor_comision = valor.replace(',', '')
        abono.porcentaje_comision = porc or None
        abono.comision_pagado_a = user
        abono.comision_pagado_el = fecha_pago or timezone.now()
        abono.comision_pagado_por = request.user
        abono.save()
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'msg': str(e)})

@require_POST
@login_required
def set_rangos_comision(request, project):
    perfil = request.user.user_profile
    if not perfil.has_permission('modificar parametros cartera'):
        return JsonResponse({'ok': False, 'msg': 'Sin permiso'})
    rangos = request.POST.get('rangos')
    try:
        param = Parameters.objects.get(name='rangos_comision', project__name=project)
        param.json = json.loads(rangos)
        param.save()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'msg': str(e)})

@login_required
def get_rangos_comision(request, project):
    param = Parameters.objects.get(name='rangos_comision', project__name=project)
    return JsonResponse(param.json or [], safe=False)


@require_POST
@login_required
def set_default_admin_collector(request, project):
    try:
        obj = Projects.objects.get(name=project)
        uid = request.POST.get('id')
        if not uid:
            return JsonResponse({'ok': False, 'msg': 'ID requerido'})
        obj.default_admin_collector_id = uid
        obj.save()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'msg': str(e)})
    
@login_required
def ajax_print_comisiones_cartera(request, project):
    if request.is_ajax():
        if request.method == 'GET':
            anio = int(request.GET.get('anio'))
            mes = int(request.GET.get('mes'))
            gestor_id = request.GET.get('gestor')
            tipo_cartera = request.GET.get('tipo_cartera')
            
            obj_project = Projects.objects.get(name=project)
            
            # Verificar permisos de rol igual que en ajax_comisiones_cartera
            es_gestor_cartera = any(
                all(palabra in rol.descripcion.lower() for palabra in ['gestor', 'cartera'])
                for rol in request.user.user_profile.rol.all()
            )
            
            es_lider_cartera = any(
                all(palabra in rol.descripcion.lower() for palabra in ['lider', 'cartera'])
                for rol in request.user.user_profile.rol.all()
            )
            
            # Si es gestor pero no líder, forzar filtro por usuario actual
            if es_gestor_cartera and not es_lider_cartera:
                gestor_id = request.user.id
            
            # Crear una copia del request.GET con los filtros
            filtros_get = request.GET.copy()
            if es_gestor_cartera and not es_lider_cartera:
                filtros_get['gestor'] = request.user.id
            
            # Crear objeto simulado para pasar a get_gestores_comisiones
            class MockRequest:
                def __init__(self, get_data):
                    self.GET = get_data
            
            mock_req = MockRequest(filtros_get)
            gestores = get_gestores_comisiones(project, anio, mes, req=mock_req)
            
            # Calcular totales
            total_presupuesto = 0
            total_cobrado = 0
            total_abonos = 0
            total_comision = 0
            
            gestores_data = []
            for user_id, tipos in gestores.items():
                for tipo_data in tipos:
                    gestores_data.append(tipo_data)
                    total_presupuesto += tipo_data['presupuesto']
                    total_cobrado += tipo_data['cobrado']
                    total_abonos += tipo_data['abonos_capital']
                    total_comision += tipo_data['comision_total']
            
            totales = {
                'total_presupuesto': total_presupuesto,
                'total_cobrado': total_cobrado,
                'total_abonos': total_abonos,
                'total_comision': total_comision
            }
            
            # Preparar contexto
            mes_nombre = calendar.month_name[mes].capitalize()
            filtros_aplicados = []
            if gestor_id:
                try:
                    usuario = User.objects.get(id=gestor_id)
                    filtros_aplicados.append(f"Gestor: {usuario.get_full_name()}")
                except:
                    pass
            if tipo_cartera:
                filtros_aplicados.append(f"Tipo: {tipo_cartera.capitalize()}")
            
            template = 'pdf/comisiones_cartera_resume.html'
            context = {
                'gestores': gestores_data,
                'totales': totales,
                'anio': anio,
                'mes': mes,
                'mes_nombre': mes_nombre,
                'project': obj_project,
                'now': datetime.now(),
                'user': request.user,
                'filtros_aplicados': filtros_aplicados,
                'es_lider': es_lider_cartera,
                'es_gestor': es_gestor_cartera and not es_lider_cartera
            }
            
            filename = f'Comisiones_cartera_{obj_project.name_to_show}_{mes_nombre}_{anio}.pdf'.replace('ñ', 'n').replace(' ', '_')
            pdf = pdf_gen(template, context, filename)
            
            pdf_url = pdf.get('url')
            
            data = {
                'type': 'success',
                'title': '¡Tenemos listo el documento!',
                'pdf_url': pdf_url,
                'filename': filename
            }
            
            return JsonResponse(data)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required
@project_permission
def export_abono_capital_receipt(request, project, abono_id):
    """Exporta recibo de abono a capital en PDF"""
    obj_project = Projects.objects.get(name=project)
    
    try:
        abono = get_object_or_404(AbonoCapital, pk=abono_id, sale__project=obj_project)
        
        context = {
            'abono': abono,
            'project': obj_project,
            'income': abono.income,
            'sale': abono.sale,
            'cliente': abono.sale.first_owner,
            'gestor': abono.gestor(),
            'now': timezone.now(),
            'user': request.user,
        }
        
        filename = f'Recibo_AbonoCapital_{abono.id}_CTR{abono.sale.contract_number}_{obj_project.name}.pdf'.replace('ñ', 'n')
        
        pdf_file = pdf_gen(
            'pdf/abono_capital_receipt.html',
            context,
            filename
        )
        
        if request.is_ajax():
            href = pdf_file.get('url')
            data = {
                'type': 'success',
                'title': '¡Recibo generado!',
                'msj': f'Puedes descargar el recibo <a href="{href}" target="_blank"><strong>aquí</strong></a>'
            }
            return JsonResponse(data)
        
        f = open(pdf_file.get('root'), 'rb')
        return FileResponse(f, as_attachment=True, filename=filename)
        
    except Exception as e:
        if request.is_ajax():
            return JsonResponse({
                'type': 'error',
                'title': 'Error',
                'msj': f'Error al generar el recibo: {str(e)}'
            })
        messages.error(request, f'Error al generar el recibo: {str(e)}')
        return redirect('abonos_capital_periodo', project=project)

@login_required
def ajax_calcular_mora_actual(request, project):
    """Vista AJAX para calcular la mora actual de un cliente"""
    if not request.is_ajax():
        return JsonResponse({'error': 'Solo AJAX'}, status=400)
    
    sale_id = request.GET.get('sale_id')
    fecha_pago = request.GET.get('fecha_pago')
    
    if not sale_id:
        return JsonResponse({'mora_actual': 0, 'mora_formateada': '$0'})
    
    try:
        sale = Sales.objects.get(pk=sale_id, project__name=project)
        
        # Usar fecha de pago si se proporciona, sino usar hoy
        if fecha_pago:
            try:
                fecha = datetime.strptime(fecha_pago, '%Y-%m-%d').date()
            except ValueError:
                fecha = date.today()
        else:
            fecha = date.today()
        
        mora_total = 0
        cuotas_con_mora = []
        
        cuotas = Credit_info.objects.filter(sale=sale).order_by('pay_date')
        
        for cuota in cuotas:
            if cuota.pay_date <= fecha:
                pending = cuota.quota_pending()
                total_pending = pending.get('total_pending', 0)
                
                if total_pending > 0:
                    arrears_info = cuota.arrears_info(paid_day=fecha)
                    valor_mora = arrears_info.get('r_value', 0)
                    dias_mora = arrears_info.get('days', 0)
                    
                    if valor_mora > 0:
                        mora_total += valor_mora
                        cuotas_con_mora.append({
                            'cuota': cuota.id_quota,
                            'fecha_vencimiento': cuota.pay_date.strftime('%d/%m/%Y'),
                            'dias_mora': dias_mora,
                            'valor_mora': valor_mora,
                            'valor_mora_formateado': f'${valor_mora:,.0f}'
                        })
        
        return JsonResponse({
            'mora_actual': mora_total,
            'mora_formateada': f'${mora_total:,.0f}',
            'cuotas_con_mora': cuotas_con_mora,
            'total_cuotas_mora': len(cuotas_con_mora)
        })
        
    except Sales.DoesNotExist:
        return JsonResponse({'error': 'Venta no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@project_permission
def ajax_calcular_mora_solicitud(request, project):
    """Vista AJAX para calcular la mora de una solicitud específica"""
    solicitud_id = request.GET.get('solicitud_id')
    
    if not solicitud_id:
        return JsonResponse({'error': 'ID de solicitud requerido'}, status=400)
    
    try:
        solicitud = SolicitudRecibo.objects.get(
            pk=solicitud_id, 
            sale__project__name=project
        )
        
        mora_total = 0
        cuotas = Credit_info.objects.filter(sale=solicitud.sale)
        
        for cuota in cuotas:
            if cuota.pay_date <= solicitud.payment_date:
                pending = cuota.quota_pending()
                if pending.get('total_pending', 0) > 0:
                    arrears_info = cuota.arrears_info(paid_day=solicitud.payment_date)
                    mora_total += arrears_info.get('r_value', 0)
        
        # Calcular valor de condonación
        valor_condonacion = 0
        if solicitud.arrears_condonate > 0:
            valor_condonacion = mora_total * (solicitud.arrears_condonate / 100)
        
        return JsonResponse({
            'mora_actual': mora_total,
            'mora_formateada': f'${mora_total:,.0f}',
            'valor_condonacion': valor_condonacion,
            'valor_condonacion_formateado': f'${valor_condonacion:,.0f}',
        })
        
    except SolicitudRecibo.DoesNotExist:
        return JsonResponse({'error': 'Solicitud no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@project_permission
def ajax_calcular_mora_actual(request, project):
    """Vista AJAX para calcular la mora actual de un cliente"""
    sale_id = request.GET.get('sale_id')
    fecha_pago = request.GET.get('fecha_pago')
    
    if not sale_id:
        return JsonResponse({'mora_actual': 0, 'mora_formateada': '$0'})
    
    try:
        sale = Sales.objects.get(pk=sale_id, project__name=project)
        
        # Usar fecha de pago si se proporciona, sino usar hoy
        if fecha_pago:
            try:
                fecha = datetime.strptime(fecha_pago, '%Y-%m-%d').date()
            except ValueError:
                fecha = date.today()
        else:
            fecha = date.today()
        
        mora_total = 0
        cuotas_con_mora = []
        
        cuotas = Credit_info.objects.filter(sale=sale).order_by('pay_date')
        
        for cuota in cuotas:
            if cuota.pay_date <= fecha:
                pending = cuota.quota_pending()
                total_pending = pending.get('total_pending', 0)
                
                if total_pending > 0:
                    arrears_info = cuota.arrears_info(paid_day=fecha)
                    valor_mora = arrears_info.get('r_value', 0)
                    
                    if valor_mora > 0:
                        mora_total += valor_mora
                        cuotas_con_mora.append({
                            'cuota': cuota.id_quota,
                            'valor_mora': valor_mora
                        })
        
        return JsonResponse({
            'mora_actual': mora_total,
            'mora_formateada': f'${mora_total:,.0f}',
            'cuotas_con_mora': cuotas_con_mora,
            'total_cuotas_mora': len(cuotas_con_mora)
        })
        
    except Sales.DoesNotExist:
        return JsonResponse({'error': 'Venta no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@project_permission
def ajax_detalle_recaudos_gestor(request, project):
    """Vista AJAX para obtener el detalle de recaudos de un gestor específico"""
    anio = int(request.GET.get('anio'))
    mes = int(request.GET.get('mes'))
    gestor_id = request.GET.get('gestor')
    tipo_cartera = request.GET.get('tipo_cartera')
    
    # Verificar permisos
    es_gestor_cartera = any(
        all(palabra in rol.descripcion.lower() for palabra in ['gestor', 'cartera'])
        for rol in request.user.user_profile.rol.all()
    )
    
    es_lider_cartera = any(
        all(palabra in rol.descripcion.lower() for palabra in ['lider', 'cartera'])
        for rol in request.user.user_profile.rol.all()
    )
    
    # Si es gestor pero no líder, solo puede ver sus propios datos
    if es_gestor_cartera and not es_lider_cartera:
        if str(request.user.id) != str(gestor_id):
            return JsonResponse({'error': 'Sin permisos'}, status=403)
    
    try:
        # Obtener el presupuesto del período
        budget = Collection_budget.objects.get(
            project__name=project, 
            year=anio, 
            month=mes
        )
        
        # Obtener detalles del gestor y tipo de cartera
        detalles_budget = Collection_budget_detail.objects.filter(
            budget=budget,
            collector_id=gestor_id,
            portfolio_type=tipo_cartera
        ).select_related('sale', 'sale__first_owner')
        
        # Diccionario para totalizar por recibo
        recibos_dict = {}
        
        for detalle in detalles_budget:
            # Obtener todos los recaudos del período para esta venta
            pagos = Incomes_detail.objects.filter(
                quota__sale=detalle.sale,
                income__payment_date__year=anio,
                income__payment_date__month=mes
            ).select_related('income', 'quota').order_by('income__payment_date')
            
            for pago in pagos:
                recibo_key = pago.income.receipt
                
                if recibo_key not in recibos_dict:
                    # Verificar si es abono a capital
                    es_abono_capital = pago.income.description and 'abono' in pago.income.description.lower() and 'capital' in pago.income.description.lower()
                    
                    recibos_dict[recibo_key] = {
                        'contrato': f'CTR{detalle.sale.contract_number}',
                        'cliente': detalle.sale.first_owner.full_name(),  # Llamar al método aquí
                        'fecha_pago': pago.income.payment_date.strftime('%d/%m/%Y'),
                        'recibo': pago.income.receipt,
                        'capital': 0,
                        'interes': 0,
                        'otros': 0,
                        'mora': 0,
                        'total': 0,
                        'es_abono_capital': es_abono_capital,
                        'fecha_ordenar': pago.income.payment_date
                    }
                
                # Sumar los valores al recibo
                recibos_dict[recibo_key]['capital'] += float(pago.capital or 0)
                recibos_dict[recibo_key]['interes'] += float(pago.interest or 0)
                recibos_dict[recibo_key]['otros'] += float(pago.others or 0)
                recibos_dict[recibo_key]['mora'] += float(pago.arrears or 0)
                recibos_dict[recibo_key]['total'] += float((pago.capital or 0) + (pago.interest or 0) + (pago.others or 0) + (pago.arrears or 0))
        
        # Convertir a lista y ordenar por fecha
        recaudos = list(recibos_dict.values())
        recaudos.sort(key=lambda x: x['fecha_ordenar'])
        
        # Limpiar fecha_ordenar de la respuesta
        for recaudo in recaudos:
            del recaudo['fecha_ordenar']
        
        return JsonResponse({
            'recaudos': recaudos,
            'total_recaudos': len(recaudos)
        })
        
    except Collection_budget.DoesNotExist:
        return JsonResponse({'error': 'No existe presupuesto para el período seleccionado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def ajax_get_sale_status(request, project, sale_id):
    """
    Vista AJAX para obtener el status de una venta.
    Usado para ocultar/mostrar campos en el formulario de solicitud según el status.
    """
    try:
        sale = Sales.objects.get(pk=sale_id)
        return JsonResponse({
            'status': sale.status,
            'contract_number': sale.contract_number,
            'client': sale.first_owner.full_name(),
        })
    except Sales.DoesNotExist:
        return JsonResponse({'error': 'Venta no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# utils
def apply_income(income, condonate_arrears=0, apply=True, no_apply_data={}, abono_capital=False, tipo_abono=None):
    import datetime
    from decimal import Decimal
    import numpy_financial as npf
    from sales.models import Payment_plans
    from finance.models import Incomes_detail, AbonoCapital

    if apply:
        total_income = income.value
        paid_day = income.payment_date
        sale = income.sale.pk
        tasa = float(income.sale.sale_plan.rate) / 100
    else:
        sale = no_apply_data.get('sale')
        total_income = no_apply_data.get('total_income', 0)
        paid_day = no_apply_data.get('paid_day', 0)
        tasa = no_apply_data.get('rate', 0) / 100

    if isinstance(paid_day, datetime.datetime):
        paid_day = paid_day.date()
    elif not isinstance(paid_day, datetime.date):
        paid_day = datetime.date.today()

    credit = Credit_info.objects.filter(sale=sale).order_by('pay_date', 'quota_type')
    applicated = []
    remaining_value = Decimal(total_income)
    hoy = datetime.date.today()
    
    sale_obj = Sales_extra_info.objects.get(pk=sale)
    hay_ci_pendiente = sale_obj.has_pending_ci_quota()
    
    # Si el pago supera lo exigible a la fecha y el usuario no eligio un tipo
    # de abono, por defecto lo tratamos como pago anticipado de cuotas.
    if not hay_ci_pendiente and not abono_capital and not tipo_abono and remaining_value > 0:
        total_exigible_hoy = Decimal("0")
        for quota in credit:
            if quota.pay_date > paid_day:
                break
            total_exigible_hoy += Decimal(str(quota.quota_pending().get("total_pending") or 0))

        if remaining_value > total_exigible_hoy:
            abono_capital = True
            tipo_abono = "cuotas_futuras"

    if hay_ci_pendiente and abono_capital:
        if apply:
            raise ValueError("No se puede aplicar abono a capital mientras existan cuotas CI pendientes.")
        else:
            return {'error': 'No se puede aplicar abono a capital mientras existan cuotas CI pendientes.'}
    
    print(f"DEBUG: abono_capital es: {abono_capital} (Tipo: {type(abono_capital)})")
    print(f"DEBUG: tipo_abono es: '{tipo_abono}'")
    print(f"DEBUG: Condición de break futura: {not (abono_capital and tipo_abono == 'cuotas_futuras')}")
    # Fase 1: pagar cuotas vencidas (pay_date <= paid_day)
    for quota in credit:
        if remaining_value <= 0:
            break
        if hay_ci_pendiente:
            pass
        elif not (abono_capital and tipo_abono == "cuotas_futuras") and quota.pay_date > paid_day:
            break
        
        elif not abono_capital:
            total_pendiente = sum(float(q.quota_pending().get('total_pending') or 0) for q in credit)
            if total_income > total_pendiente:
                sobra = total_income - total_pendiente
                if apply:
                    raise ValueError("El valor a pagar supera el total pendiente del crédito. Sobra ${:,.2f}.".format(sobra))
                else:
                    return {'error': 'El valor a pagar supera el total pendiente del crédito. Sobra ${:,.2f}.'.format(sobra)}
                            
        pending = quota.quota_pending()
        total_pending = pending.get('total_pending') or 0

        print(f"DEBUG Cuota {quota.id_quota}: total_pending={total_pending}, remaining_value={remaining_value}")

        if total_pending > 0:

            capital = pending.get('pendient_capital')
            interest = pending.get('pendient_int')
            others = pending.get('pendient_others')
            arrears_data = quota.arrears_info(paid_day=paid_day)
            initial_arrears = arrears_data.get('r_value')
            arrears = Decimal(initial_arrears) * (Decimal('1') - Decimal(condonate_arrears) / Decimal('100'))
            arrears_days = arrears_data.get('days')

            paid_arrears = min(remaining_value, arrears)
            remaining_value -= paid_arrears

            paid_others = min(remaining_value, others)
            remaining_value -= paid_others

            paid_interest = min(remaining_value, interest)
            remaining_value -= paid_interest

            paid_capital = min(remaining_value, capital)
            remaining_value -= paid_capital

            print(f"DEBUG Cuota {quota.id_quota}: pagado capital={paid_capital}, interest={paid_interest}, others={paid_others}, arrears={paid_arrears}")
            print(f"DEBUG remaining_value después de pagar: {remaining_value}")

            if apply:
                Incomes_detail.objects.create(
                    income=income, quota=quota,
                    capital=paid_capital, interest=paid_interest,
                    others=paid_others, arrears=paid_arrears,
                    arrears_days=arrears_days,
                )
            else:
                date_pay = datetime.datetime.strftime(quota.pay_date, '%Y/%m/%d')
                applicated.append({
                    'quota': quota.id_quota,
                    'date': date_pay,
                    'capital': float(capital),
                    'interest': float(interest),
                    'others': float(others),
                    'arrears': float(initial_arrears),
                    'total': float(capital + interest + arrears),
                    'paid_capital': float(paid_capital),
                    'paid_interest': float(paid_interest),
                    'paid_others': float(paid_others),
                    'paid_arrears': float(paid_arrears),
                    'paid_total': float(paid_capital + paid_interest + paid_others + paid_arrears),
                    'arrears_days': arrears_days,
                })

    # Fase 2: aplicar abono a capital o pagar cuotas futuras
    if remaining_value > 0:
        cuotas_futuras = credit.filter(pay_date__gt=paid_day).order_by('pay_date')
        n = cuotas_futuras.count()
        saldo_capital = sum(float(q.capital) for q in cuotas_futuras)
        
        if remaining_value > saldo_capital:
            
            sobra = remaining_value - Decimal(str(saldo_capital))
            if apply:
                raise ValueError("El valor a pagar es mayor al saldo de capital pendiente. Sobran ${:,.2f}.".format(sobra))
            else:
                return {'error': 'El valor a pagar es mayor al saldo de capital pendiente. Sobran ${:,.2f}.'.format(sobra)}
        
        if not abono_capital:
            if apply:
                raise ValueError("Este pago aplica cuotas futuras, escoge una opcion de abono a capital.")
            else:
                return {'error': 'Este pago aplica a futuras, escoge una opcion de abono a capital.'}
        
                
        if abono_capital and tipo_abono and not apply:
            if tipo_abono == "reducir_tiempo":
                valor_cuota = float(cuotas_futuras.first().total_payment()) if n > 0 else 0
                nuevo_saldo_preview = float(saldo_capital) - float(remaining_value)
                nuevas_cuotas = int(round(npf.nper(float(tasa), valor_cuota, -nuevo_saldo_preview))) or 1
                nota = f'Abono capital: el plazo se reduce a {nuevas_cuotas * -1} cuotas aprox.'
                applicated.append({
                    'quota': 'ABCAP',
                    'date': '',
                    'capital': float(remaining_value),
                    'interest': 0,
                    'others': 0,
                    'arrears': 0,
                    'total': float(remaining_value),
                    'paid_capital': float(remaining_value),
                    'paid_interest': 0,
                    'paid_others': 0,
                    'paid_arrears': 0,
                    'paid_total': float(remaining_value),
                    'arrears_days': 0,
                    'nota': nota,
                    'nuevas_cuotas': nuevas_cuotas
                })
                return applicated
            elif tipo_abono == "reducir_cuota":
                nuevo_saldo_preview = float(saldo_capital) - float(remaining_value)
                nueva_cuota = int(round(npf.pmt(float(tasa), n, -nuevo_saldo_preview))) if n > 0 else 0
                nota = f'Abono capital: el valor de la cuota se reduce a ${nueva_cuota:,}.'
                applicated.append({
                    'quota': 'ABCAP',
                    'date': '',
                    'capital': float(remaining_value),
                    'interest': 0,
                    'others': 0,
                    'arrears': 0,
                    'total': float(remaining_value),
                    'paid_capital': float(remaining_value),
                    'paid_interest': 0,
                    'paid_others': 0,
                    'paid_arrears': 0,
                    'paid_total': float(remaining_value),
                    'arrears_days': 0,
                    'nota': nota,
                    'nueva_cuota': nueva_cuota
                })
                return applicated
        
        elif apply:
            
            recalcular_plan_por_abono(income, abono_capital=remaining_value, tipo_abono=tipo_abono)

    return applicated

def recalcular_plan_por_abono(income, abono_capital, tipo_abono):
    import numpy_financial as npf
    from decimal import Decimal, ROUND_HALF_UP
    from sales.models import Payment_plans, backup_payment_plans
    from finance.models import AbonoCapital, Incomes_detail

    paid_day = income.add_date
    sale = income.sale

    print(f"DEBUG RECALCULAR: Iniciando recalculo para tipo_abono='{tipo_abono}', abono_capital={abono_capital}")

    # INICIALIZAR cuotas_afectadas al principio
    cuotas_afectadas = 0

    cuotas_recalculables = Payment_plans.objects.filter(
        sale=sale,
        pay_date__gt=paid_day,
        quota_type__in=["SCR", "SCE"]
    ).order_by('pay_date')

    saldo_restante = sum(q.capital for q in cuotas_recalculables)
    nuevo_saldo = saldo_restante - Decimal(str(abono_capital))

    print(f"DEBUG RECALCULAR: Cuotas recalculables={cuotas_recalculables.count()}, saldo_restante={saldo_restante}, nuevo_saldo={nuevo_saldo}")
    
    if nuevo_saldo <= 0 or not cuotas_recalculables.exists():
        return
    
    valor_cuota = round(cuotas_recalculables.first().total_payment())
    
    # Backup completo del plan de pagos
    plan_actual = Payment_plans.objects.filter(sale=sale).order_by('pay_date')
    for cuota in plan_actual:
        backup_payment_plans.objects.create(
            backup_date=timezone.now(),
            id_payment=cuota.id_payment,
            id_quota=cuota.id_quota,
            quota_type=cuota.quota_type,
            sale=cuota.sale,
            pay_date=cuota.pay_date,
            capital=cuota.capital,
            interest=cuota.interest,
            others=cuota.others,
            project=cuota.project
        )
        
    if tipo_abono == "reducir_tiempo":
        print(f"DEBUG REDUCIR_TIEMPO: Iniciando proceso")
        tasa_mensual = Decimal(sale.sale_plan.rate) / Decimal("100")
        capital_restante = Decimal(nuevo_saldo)

        print(f"DEBUG REDUCIR_TIEMPO: tasa_mensual={tasa_mensual}, capital_restante inicial={capital_restante}")

        for cuota in cuotas_recalculables:
            if capital_restante <= 0:
                print(f"DEBUG REDUCIR_TIEMPO: Eliminando cuota {cuota.id_quota} (capital_restante <= 0)")
                cuota.delete()
                continue

            capital_original = cuota.capital
            interes_original = cuota.interest

            valor_cuota = cuota.total_payment()
            interes = (capital_restante * tasa_mensual).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            capital = min(valor_cuota - interes, capital_restante)

            cuota.capital = capital
            cuota.interest = interes
            cuota.others = 0
            cuota.save()

            print(f"DEBUG REDUCIR_TIEMPO: Cuota {cuota.id_quota} - Original: cap={capital_original}, int={interes_original} -> Nuevo: cap={capital}, int={interes}")

            capital_restante -= capital
            cuotas_afectadas += 1

        print(f"DEBUG REDUCIR_TIEMPO: Proceso terminado. Capital restante final={capital_restante}, cuotas_afectadas={cuotas_afectadas}")

                
    elif tipo_abono == "reducir_cuota":
        print(f"DEBUG REDUCIR_CUOTA: Iniciando proceso")
        tasa_mensual = Decimal(sale.sale_plan.rate) / Decimal("100")

        capital_scr = sum(c.capital for c in cuotas_recalculables.filter(quota_type="SCR"))
        capital_sce = sum(c.capital for c in cuotas_recalculables.filter(quota_type="SCE"))
        capital_total = capital_scr + capital_sce

        print(f"DEBUG REDUCIR_CUOTA: capital_scr={capital_scr}, capital_sce={capital_sce}, capital_total={capital_total}")

        saldo_por_tipo = {
            "SCR": nuevo_saldo * Decimal(str(capital_scr / capital_total)) if capital_total > 0 else Decimal("0"),
            "SCE": nuevo_saldo * Decimal(str(capital_sce / capital_total)) if capital_total > 0 else Decimal("0"),
        }

        print(f"DEBUG REDUCIR_CUOTA: saldo_por_tipo={saldo_por_tipo}")

        for tipo in ["SCR", "SCE"]:
            cuotas = list(cuotas_recalculables.filter(quota_type=tipo).order_by("pay_date"))

            if len(cuotas) < 2:
                print(f"DEBUG REDUCIR_CUOTA: Saltando tipo {tipo}, solo tiene {len(cuotas)} cuotas (necesita al menos 2)")
                continue

            saldo = saldo_por_tipo[tipo]
            n = len(cuotas)
            f1, f2 = cuotas[0].pay_date, cuotas[1].pay_date
            meses_diferencia = (f2.year - f1.year) * 12 + (f2.month - f1.month)
            tasa_efectiva = (1 + tasa_mensual) ** meses_diferencia - 1

            nueva_cuota = Decimal(npf.pmt(tasa_efectiva, n, -saldo)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
            capital_acumulado = Decimal("0")

            print(f"DEBUG REDUCIR_CUOTA: Procesando tipo {tipo}, n={n}, tasa_efectiva={tasa_efectiva}, nueva_cuota={nueva_cuota}, saldo={saldo}")

            for i, cuota in enumerate(cuotas):
                capital_original = cuota.capital
                interes_original = cuota.interest

                interes = (saldo * Decimal(tasa_efectiva)).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
                capital = nueva_cuota - interes

                if capital > saldo:
                    capital = saldo
                    interes = nueva_cuota - capital

                if i == len(cuotas) - 1 or saldo - capital <= 0:
                    capital = saldo
                    interes = nueva_cuota - capital
                    interes = max(interes, Decimal("0"))

                cuota.capital = capital
                cuota.interest = interes
                cuota.others = 0
                cuota.save()

                print(f"DEBUG REDUCIR_CUOTA: Cuota {cuota.id_quota} ({tipo}) - Original: cap={capital_original}, int={interes_original} -> Nuevo: cap={capital}, int={interes}, saldo restante={saldo}")

                saldo -= capital
                capital_acumulado += capital
                cuotas_afectadas += 1

            print(f"DEBUG REDUCIR_CUOTA: Tipo {tipo} completado. Capital acumulado={capital_acumulado}, saldo final={saldo}")
    
    # Si no se procesó ningún tipo específico, usar el conteo total como fallback
    if cuotas_afectadas == 0:
        print(f"DEBUG RECALCULAR: cuotas_afectadas era 0, usando count={cuotas_recalculables.count()}")
        cuotas_afectadas = cuotas_recalculables.count()

    print(f"DEBUG RECALCULAR: Creando cuota ABCAP con capital={abono_capital}, cuotas_afectadas={cuotas_afectadas}")

    # Calcular totales ANTES de crear ABCAP para verificación
    total_capital_antes = Decimal(str(sum(q.capital for q in Payment_plans.objects.filter(sale=sale, quota_type__in=["SCR", "SCE"])) or 0))
    print(f"DEBUG RECALCULAR: Total capital SCR/SCE ANTES de ABCAP={total_capital_antes}")

    # Crear cuota tipo ABCAP
    n_abonos = Payment_plans.objects.filter(sale=sale, quota_type='ABCAP').count()
    id_abono = f"ABCAP{n_abonos + 1}CTR{sale.contract_number}"
    cuota_abono = Payment_plans.objects.create(
        sale=sale,
        id_quota=id_abono,
        quota_type='ABCAP',
        pay_date=paid_day,
        capital=Decimal(abono_capital),
        interest=0,
        others=0,
        project=sale.project
    )

    print(f"DEBUG RECALCULAR: Cuota ABCAP creada: {id_abono}")

    AbonoCapital.objects.create(
        income=income,
        sale=sale,
        tipo=tipo_abono,
        capital_aplicado=abono_capital,
        cuotas_afectadas=cuotas_afectadas,
        nueva_cuota=valor_cuota
    )

    Incomes_detail.objects.create(
        income=income,
        quota=cuota_abono,
        capital=abono_capital,
        interest=0,
        others=0,
        arrears=0,
        arrears_days=0
    )

    # Calcular totales DESPUÉS de crear ABCAP para verificación
    total_capital_despues = Decimal(str(sum(q.capital for q in Payment_plans.objects.filter(sale=sale, quota_type__in=["SCR", "SCE", "ABCAP"])) or 0))
    total_capital_scr_sce = Decimal(str(sum(q.capital for q in Payment_plans.objects.filter(sale=sale, quota_type__in=["SCR", "SCE"])) or 0))
    print(f"DEBUG RECALCULAR: Total capital SCR/SCE DESPUÉS de ABCAP={total_capital_scr_sce}")
    print(f"DEBUG RECALCULAR: Total capital CON ABCAP={total_capital_despues}")
    print(f"DEBUG RECALCULAR: Diferencia (debería ser ~= abono_capital)={total_capital_antes - total_capital_scr_sce}")
    print(f"DEBUG RECALCULAR: VERIFICACIÓN: total_capital_antes={total_capital_antes}, saldo_restante={saldo_restante}, abono_capital={abono_capital}, nuevo_saldo esperado={nuevo_saldo}")
    
def get_gestores_comisiones(project, anio, mes, req=None):
    from finance.models import Collection_budget, Collection_budget_detail, ComisionGestorCartera, Incomes_detail, AbonoCapital
    gestor_id = req.GET.get('gestor') if req else None
    tipo_cartera = req.GET.get('tipo_cartera') if req else None

    try:
        budget = Collection_budget.objects.get(project__name=project, year=anio, month=mes)
    except Collection_budget.DoesNotExist:
        return []

    detalles = Collection_budget_detail.objects.filter(budget=budget)
    gestores = ComisionGestorCartera.objects.filter(usuario__is_active=True)
    if gestor_id:
        gestores = gestores.filter(usuario_id=gestor_id)
    if tipo_cartera:
        gestores = gestores.filter(tipo_cartera=tipo_cartera)

    res = []
    for g in gestores:
        dets = detalles.filter(collector=g.usuario, portfolio_type=g.tipo_cartera)
        presupuesto = sum((d.lt_30 or 0)+(d.lt_60 or 0)+(d.lt_90 or 0)+(d.lt_120 or 0)+(d.gt_120 or 0) for d in dets)
        cobrado = 0
        for d in dets:
            cobro = Incomes_detail.objects.filter(
                quota__sale=d.sale,
                income__payment_date__year=anio,
                income__payment_date__month=mes
            ).aggregate(s=Sum('capital')+Sum('interest')+Sum('others')+Sum('arrears'))['s'] or 0
            cobrado += cobro
        abonos_capital = AbonoCapital.objects.filter(
            sale__in=[d.sale for d in dets],
            income__payment_date__year=anio,
            income__payment_date__month=mes,
            valor_comision__isnull=False
        ).aggregate(s=Sum('capital_aplicado'))['s'] or 0
        porc = float(g.porcentaje_comision/100 or 0)
        base_comision = float(cobrado) - float(abonos_capital)
        res.append({
            'usuario': g.usuario,
            'tipo_cartera': g.tipo_cartera,
            'presupuesto': presupuesto,
            'cobrado': cobrado,
            'abonos_capital': abonos_capital,
            'porcentaje_comision': round(porc*100, 2) if g.porcentaje_comision else 0,
            'comision_total': float(base_comision) * porc,
            'base_comision': base_comision
        })

    from collections import defaultdict
    agrupado = defaultdict(list)
    for item in res:
        key = item['usuario'].id
        agrupado[key].append({
            'nombre': item['usuario'].get_full_name() or item['usuario'].username,
            'tipo_cartera': item['tipo_cartera'],
            'presupuesto': item['presupuesto'],
            'cobrado': item['cobrado'],
            'abonos_capital': item['abonos_capital'],
            'porcentaje_comision': item['porcentaje_comision'],
            'comision_total': item['comision_total'],
            'base_comision': item['base_comision']
        })
    return dict(agrupado)

def get_ci_status_notifications(project, year, month, dets):
    dets = dets.select_related('sale', 'budget')

    entering = []
    for d in dets:
        prev = (
            Collection_budget_detail.objects
            .filter(
                sale_id=d.sale_id,
                budget__project=project
            )
            .exclude(
                Q(budget__year=year, budget__month=month) |
                Q(budget__year__gt=year) |
                Q(budget__year=year, budget__month__gt=month)
            )
            .order_by('-budget__year', '-budget__month')
            .first()
        )
        
        if prev and prev.portfolio_type == 'comercial':
            entering.append({
                "contract_number": d.sale.contract_number,
                "first_owner": d.sale.first_owner.full_name(),
                "sale_id": d.sale_id
            })
    return entering

urlpattern = [
    path('projectselection', project_selection),
    path('<project>/incomes/new_sales', new_sales_incomes),
    path('<project>/incomes/adjudicated_sales', adjudicated_sales_incomes),
    path('<project>/incomes/list', incomes_list),
    path('<project>/comissions/liquidate', liquidate_comissions),
    path('<project>/comissions/list', comissions_list),
    path('<project>/comissions/export', export_paid_comissions, name='export_paid_comissions'),
    path('<project>/comissions/advances', liquidate_comissions_advances),
    path('<project>/comissions/parameters', comission_positions_parameters, name='comission_positions_parameters'),
    path('<project>/pmt', pmt),
    path('<project>/collectionbudget/<portfolio_type>', collection_budget),
    path('<project>/availablecash', incomes_division),
    path('<project>/expenses', expenses),
    path('<project>/commercialbudget', commercial_budget),
    path('<str:project>/solicitudes/', lista_solicitudes_recibo, name='lista_solicitudes_recibo'),
    path('<str:project>/solicitudes/nueva/', crear_solicitud_recibo, name='crear_solicitud_recibo'),
    path('<str:project>/solicitudes/<int:pk>/validar/', validar_solicitud_recibo, name='validar_solicitud_recibo'),
    path('<str:project>/solicitudes/<int:pk>/editar/', editar_solicitud_recibo, name='editar_solicitud_recibo'),
    path('<str:project>/solicitudes/<int:pk>/eliminar/', eliminar_solicitud_recibo, name='eliminar_solicitud_recibo'),
    path('<str:project>/solicitudes/<int:pk>/autorizar_condonacion/', autorizar_condonacion, name='autorizar_condonacion'),
    path('<str:project>/parametros/', parametros, name='parametros'),
    path('<str:project>/comisiones/', comisiones_cartera, name='comisiones_cartera'),
    path('<str:project>/abonos_capital/', abonos_capital_periodo, name='abonos_capital_periodo'),
] + [
    path('<project>/print/receipt', reprint_receipt, name='reprint_receipt'),
    path('ajax/<project>/incomes_actions', incomes_actions),
    path('ajax/<project>/comissions/positions', ajax_comission_positions, name='ajax_comission_positions'),
    path('ajax/<project>/comissions/positions/save', ajax_save_comission_position, name='ajax_save_comission_position'),
    path('ajax/<project>/comissions/positions/toggle', ajax_toggle_comission_position, name='ajax_toggle_comission_position'),
    path('ajax/<project>/comissions/minvalue', ajax_set_min_comission_value, name='ajax_set_min_comission_value'),
    path('ajax/<project>/comissions/method', ajax_set_liquidation_method, name='ajax_set_liquidation_method'),
    path('ajax/<str:project>/print_comisiones_cartera/', ajax_print_comisiones_cartera, name='ajax_print_comisiones_cartera'),
    path('ajax/<project>/collectionfeed/newcomment', ajax_new_comment_feed),
    path('ajax/<project>/comissions/detailsandactions', ajax_comissions_actions),
    path('ajax/<project>/comissions/print', ajax_print_comissions),
    path('ajax/<project>/pmt/print', ajax_print_pmt),
    path('ajax/<project>/printfilebank', ajax_print_filetobank),
    path('ajax/<project>/reapplyreceipts/<sale_pk>',ajax_reapply_rc),
    path('ajax/<project>/printdocsupport',print_comissions_ds),
    path('ajax/<str:project>/detallecliente', detalle_cliente, name='detalle_cliente'),
    path('api/cartera/data/<int:project_id>/<str:tipo_cartera>/', get_cartera_proyecto_data, name='cartera_proyecto_data'),
    path('ajax/<str:project>/set_tasa_mora/', ajax_set_tasa_mora, name='ajax_set_tasa_mora'),
    path('ajax/<str:project>/set_comision_gestor/', ajax_set_comision_gestor, name='ajax_set_comision_gestor'),
    path('ajax/<str:project>/comisiones_cartera/', ajax_comisiones_cartera, name='ajax_comisiones_cartera'),
    path('ajax/<str:project>/abonos_capital/', ajax_abonos_capital, name='ajax_abonos_capital'),
    path('<str:project>/pagar_comision_abono/', pagar_comision_abono, name='pagar_comision_abono'),
    path('ajax/<str:project>/set_rangos_comision/', set_rangos_comision, name='set_rangos_comision'),
    path('ajax/<str:project>/get_rangos_comision/', get_rangos_comision, name='get_rangos_comision'),
    path('<str:project>/set_default_admin_collector/', set_default_admin_collector, name='set_default_admin_collector'),
    path('<str:project>/abono_capital/<int:abono_id>/export_receipt/', 
         export_abono_capital_receipt, name='export_abono_capital_receipt'),
    path('ajax/<str:project>/detalle_recaudos_gestor/', ajax_detalle_recaudos_gestor, name='ajax_detalle_recaudos_gestor'),
    path('ajax/<str:project>/calcular_mora_actual/', ajax_calcular_mora_actual, name='ajax_calcular_mora_actual'),
    path('ajax/<str:project>/calcular_mora_solicitud/', ajax_calcular_mora_solicitud, name='ajax_calcular_mora_solicitud'),
    path('ajax/<str:project>/abono_capital/<int:abono_id>/preview_reversion/', preview_reversion_abono, name='preview_reversion_abono'),
    path('ajax/<str:project>/abono_capital/<int:abono_id>/revertir/', revertir_abono_capital, name='revertir_abono_capital'),
    path('ajax/<str:project>/sale_status/<int:sale_id>/', ajax_get_sale_status, name='ajax_get_sale_status'),
]
def normalize_currency(value):
    if value is None:
        return '0'
    return re.sub(r'[^\d.-]', '', str(value)) or '0'
